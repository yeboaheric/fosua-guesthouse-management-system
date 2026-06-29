import json
from datetime import date
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import AuditLog, UserAccessProfile
from inventory.models import (
    InventoryCategory,
    InventoryItem,
    InventorySubcategory,
    InventoryTransaction,
    Sale,
    SaleItem,
    Supplier,
)
from inventory.views import _inventory_items_for_stock_lock


class InventoryPermissionTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.reception_group = Group.objects.create(name="Receptionist")
        self.admin = User.objects.create_user(username="inventory-admin", password="pass123456")
        self.admin.groups.add(self.admin_group)
        self.reception = User.objects.create_user(username="inventory-reception", password="pass123456")
        self.reception.groups.add(self.reception_group)
        UserAccessProfile.objects.create(
            user=self.reception,
            dashboard_access=True,
            reservations_access=True,
            rooms_access=True,
            guests_access=True,
            payments_access=True,
            services_access=True,
            housekeeping_access=True,
            inventory_access=False,
            pos_access=True,
            notifications_access=True,
            analytics_access=True,
            reports_access=False,
            settings_access=False,
            staff_management_access=False,
            handovers_access=True,
            users_roles_access=False,
        )
        self.sale = Sale.objects.create(
            cashier=self.admin,
            customer_name="Counter Guest",
            payment_method=Sale.PaymentMethod.CASH,
            subtotal="10.00",
            grand_total="10.00",
            amount_paid="10.00",
        )

    def test_receptionist_cannot_open_inventory_dashboard(self):
        self.client.force_login(self.reception)
        response = self.client.get(reverse("inventory-dashboard"))
        self.assertEqual(response.status_code, 403)

    def test_receptionist_can_open_pos_terminal(self):
        self.client.force_login(self.reception)
        response = self.client.get(reverse("inventory-pos"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Point of Sale")

    def test_receptionist_cannot_access_pos_sale_edit_route(self):
        self.client.force_login(self.reception)
        response = self.client.get(reverse("inventory-sale-update", args=[self.sale.pk]))
        self.assertEqual(response.status_code, 403)

    def test_receptionist_cannot_see_edit_on_pos_sale_pages(self):
        self.client.force_login(self.reception)
        sale_list_response = self.client.get(reverse("inventory-sales"))
        self.assertEqual(sale_list_response.status_code, 200)
        self.assertNotContains(sale_list_response, reverse("inventory-sale-update", args=[self.sale.pk]))
        self.assertNotContains(sale_list_response, reverse("inventory-sale-delete", args=[self.sale.pk]))

        sale_detail_response = self.client.get(reverse("inventory-sale-detail", args=[self.sale.pk]))
        self.assertEqual(sale_detail_response.status_code, 200)
        self.assertNotContains(sale_detail_response, reverse("inventory-sale-update", args=[self.sale.pk]))
        self.assertNotContains(sale_detail_response, reverse("inventory-sale-delete", args=[self.sale.pk]))

    def test_receptionist_can_open_pos_analytics_and_reports(self):
        self.client.force_login(self.reception)
        analytics_response = self.client.get(reverse("inventory-pos-analytics"))
        self.assertEqual(analytics_response.status_code, 200)
        self.assertContains(analytics_response, "POS Analytics")

        reports_response = self.client.get(reverse("inventory-pos-reports"))
        self.assertEqual(reports_response.status_code, 200)
        self.assertContains(reports_response, "POS Reports")

        detail_response = self.client.get(reverse("inventory-pos-report-detail"))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "POS Report")


class InventoryPosWorkflowTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="pos-admin", password="pass123456")
        self.user.groups.add(self.admin_group)

        self.category = InventoryCategory.objects.create(
            name="Soft Drinks",
            category_group=InventoryCategory.CategoryGroup.NON_ALCOHOLIC,
            description="Chilled drinks",
        )
        self.subcategory = InventorySubcategory.objects.create(
            category=self.category,
            name="Juices",
            description="Fruit juices",
        )
        self.supplier = Supplier.objects.create(name="Fresh Supply Co")
        self.item = InventoryItem.objects.create(
            name="Mango Juice",
            category=self.category,
            subcategory=self.subcategory,
            supplier=self.supplier,
            purchase_price="6.00",
            selling_price="10.00",
            quantity_in_stock="10.000",
            minimum_stock_threshold="2.000",
            unit_of_measure=InventoryItem.UnitOfMeasure.BOTTLE,
            description="Cold mango juice",
        )
        self.item_two = InventoryItem.objects.create(
            name="Apple Juice",
            category=self.category,
            subcategory=self.subcategory,
            supplier=self.supplier,
            purchase_price="9.00",
            selling_price="15.00",
            quantity_in_stock="5.000",
            minimum_stock_threshold="1.000",
            unit_of_measure=InventoryItem.UnitOfMeasure.BOTTLE,
            description="Fresh apple juice",
        )

    def test_item_creation_records_opening_stock_transaction(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("inventory-item-create"),
            {
                "name": "Coke Bottle",
                "category": self.category.pk,
                "subcategory": self.subcategory.pk,
                "supplier": self.supplier.pk,
                "purchase_price": "5.00",
                "selling_price": "8.00",
                "quantity_in_stock": "12.000",
                "unit_of_measure": InventoryItem.UnitOfMeasure.BOTTLE,
                "minimum_stock_threshold": "3.000",
                "description": "Chilled coke bottles",
                "is_active": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        created_item = InventoryItem.objects.get(name="Coke Bottle")
        transactions = InventoryTransaction.objects.filter(
            item=created_item,
            transaction_type=InventoryTransaction.TransactionType.PURCHASE,
        )
        self.assertEqual(transactions.count(), 1)
        self.assertEqual(transactions.first().quantity_changed, created_item.quantity_in_stock)

    def test_pos_sale_stock_lock_query_does_not_join_nullable_relations(self):
        query_sql = str(_inventory_items_for_stock_lock([self.item.pk, self.item_two.pk]).query).upper()
        self.assertNotIn("JOIN", query_sql)
        self.assertIn("ORDER BY", query_sql)

    def test_pos_checkout_deducts_stock_and_creates_sale(self):
        self.client.force_login(self.user)
        cart = json.dumps([{"id": self.item.pk, "quantity": 3}])
        response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "30.00",
                "notes": "Bar sale",
                "cart": cart,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        sale = Sale.objects.latest("created_at")
        self.item.refresh_from_db()
        self.assertEqual(sale.receipt_number.startswith("POS-"), True)
        self.assertEqual(str(sale.grand_total), "30.00")
        self.assertEqual(str(self.item.quantity_in_stock), "7.000")
        self.assertEqual(
            InventoryTransaction.objects.filter(
                item=self.item,
                sale=sale,
                transaction_type=InventoryTransaction.TransactionType.SALE,
            ).count(),
            1,
        )
        self.assertRedirects(response, f"{reverse('inventory-pos')}?sale_success=1&sale_id={sale.pk}")
        self.assertContains(response, "Sale completed successfully. The terminal will reset shortly.")
        self.assertContains(response, f'data-sale-completed="true"', html=False)
        self.assertContains(response, reverse("inventory-sale-detail", args=[sale.pk]))

    def test_failed_pos_checkout_does_not_enable_reset_state(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "0.00",
                "notes": "Broken sale",
                "cart": json.dumps([]),
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, 'data-sale-completed="false"', html=False, status_code=400)
        self.assertContains(response, "Cart is empty.", status_code=400)

    def test_sale_receipt_pdf_endpoint_returns_pdf(self):
        sale = Sale.objects.create(
            cashier=self.user,
            payment_method=Sale.PaymentMethod.CASH,
            subtotal="10.00",
            grand_total="10.00",
            amount_paid="10.00",
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("inventory-sale-pdf", args=[sale.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(sale.receipt_number, response["Content-Disposition"])

    def test_item_list_stock_export_returns_item_name_and_stock_xlsx(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("inventory-items"), {"export": "stock_xlsx"})
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("inventory-stock-list-", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, "Stock List")
        self.assertEqual(worksheet["A1"].value, "Item Name")
        self.assertEqual(worksheet["B1"].value, "Quantity in Stock")
        exported_rows = {
            (worksheet["A2"].value, worksheet["B2"].value),
            (worksheet["A3"].value, worksheet["B3"].value),
        }
        self.assertIn(("Mango Juice", 10), exported_rows)

    def test_admin_can_open_and_save_pos_sale_edit(self):
        self.client.force_login(self.user)
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "30.00",
                "notes": "Original sale",
                "cart": json.dumps([{"id": self.item.pk, "quantity": 3}]),
            },
        )
        self.assertEqual(checkout_response.status_code, 302)
        sale = Sale.objects.latest("created_at")

        sale_list_response = self.client.get(reverse("inventory-sales"))
        self.assertContains(sale_list_response, reverse("inventory-sale-update", args=[sale.pk]))
        self.assertContains(sale_list_response, reverse("inventory-sale-delete", args=[sale.pk]))

        sale_detail_response = self.client.get(reverse("inventory-sale-detail", args=[sale.pk]))
        self.assertContains(sale_detail_response, reverse("inventory-sale-update", args=[sale.pk]))
        self.assertContains(sale_detail_response, reverse("inventory-sale-delete", args=[sale.pk]))

        edit_response = self.client.get(reverse("inventory-sale-update", args=[sale.pk]))
        self.assertEqual(edit_response.status_code, 200)
        self.assertContains(edit_response, "Edit POS Sale")
        self.assertContains(edit_response, 'name="sale_date"', html=False)
        self.assertContains(edit_response, 'name="customer_name"', html=False)
        self.assertContains(edit_response, 'name="items_payload"', html=False)
        original_local_timestamp = timezone.localtime(sale.created_at)

        update_response = self.client.post(
            reverse("inventory-sale-update", args=[sale.pk]),
            {
                "sale_date": "2026-06-01",
                "customer_name": "Corrected Guest",
                "customer_phone": "0244000000",
                "customer_email": "guest@example.com",
                "payment_method": Sale.PaymentMethod.MOBILE_MONEY,
                "tax_amount": "2.00",
                "discount_amount": "1.00",
                "amount_paid": "35.00",
                "notes": "Corrected after cashier review",
                "items_payload": json.dumps(
                    [
                        {
                            "item_id": self.item.pk,
                            "quantity": 2,
                            "unit_price": "12.50",
                        },
                        {
                            "item_id": self.item_two.pk,
                            "quantity": 1,
                            "unit_price": "15.00",
                        },
                    ]
                ),
            },
            follow=True,
        )
        self.assertEqual(update_response.status_code, 200)
        sale.refresh_from_db()
        self.item.refresh_from_db()
        self.item_two.refresh_from_db()
        updated_local_timestamp = timezone.localtime(sale.created_at)
        self.assertEqual(str(updated_local_timestamp.date()), "2026-06-01")
        self.assertEqual(updated_local_timestamp.time().replace(microsecond=0), original_local_timestamp.time().replace(microsecond=0))
        self.assertEqual(sale.customer_name, "Corrected Guest")
        self.assertEqual(sale.payment_method, Sale.PaymentMethod.MOBILE_MONEY)
        self.assertEqual(str(sale.subtotal), "40.00")
        self.assertEqual(str(sale.grand_total), "41.00")
        self.assertEqual(str(sale.amount_paid), "35.00")
        self.assertEqual(str(self.item.quantity_in_stock), "8.000")
        self.assertEqual(str(self.item_two.quantity_in_stock), "4.000")
        self.assertEqual(sale.items.count(), 2)
        self.assertIsNotNone(sale.edited_at)
        self.assertEqual(sale.edited_by, self.user)
        edit_log = AuditLog.objects.filter(module="pos", object_id=str(sale.pk), action=AuditLog.ActionType.UPDATE).latest("created_at")
        self.assertEqual(edit_log.details["field_changes"]["customer_name"]["to"], "Corrected Guest")
        self.assertEqual(len(edit_log.details["item_changes"]), 2)
        self.assertContains(update_response, "updated successfully")
        self.assertContains(update_response, "Edited")

    def test_admin_can_delete_sale_and_restore_stock(self):
        self.client.force_login(self.user)
        cart = json.dumps([{"id": self.item.pk, "quantity": 2}])
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "20.00",
                "notes": "Delete sale test",
                "cart": cart,
            },
        )
        self.assertEqual(checkout_response.status_code, 302)
        sale = Sale.objects.latest("created_at")
        self.item.refresh_from_db()
        self.assertEqual(str(self.item.quantity_in_stock), "8.000")

        delete_response = self.client.post(reverse("inventory-sale-delete", args=[sale.pk]), follow=True)
        self.assertEqual(delete_response.status_code, 200)
        self.item.refresh_from_db()
        self.assertFalse(Sale.objects.filter(pk=sale.pk).exists())
        self.assertEqual(str(self.item.quantity_in_stock), "10.000")
        self.assertContains(delete_response, "Sale deleted successfully.")

    def test_admin_can_remove_sale_line_and_restore_stock(self):
        self.client.force_login(self.user)
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "45.00",
                "notes": "Multi item sale",
                "cart": json.dumps(
                    [
                        {"id": self.item.pk, "quantity": 3},
                        {"id": self.item_two.pk, "quantity": 1},
                    ]
                ),
            },
        )
        self.assertEqual(checkout_response.status_code, 302)
        sale = Sale.objects.latest("created_at")

        update_response = self.client.post(
            reverse("inventory-sale-update", args=[sale.pk]),
            {
                "sale_date": timezone.localdate().isoformat(),
                "customer_name": "",
                "customer_phone": "",
                "customer_email": "",
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "15.00",
                "notes": "Removed one line",
                "items_payload": json.dumps(
                    [
                        {
                            "item_id": self.item_two.pk,
                            "quantity": 1,
                            "unit_price": "15.00",
                        },
                    ]
                ),
            },
            follow=True,
        )
        self.assertEqual(update_response.status_code, 200)
        sale.refresh_from_db()
        self.item.refresh_from_db()
        self.item_two.refresh_from_db()
        self.assertEqual(sale.items.count(), 1)
        self.assertEqual(str(self.item.quantity_in_stock), "10.000")
        self.assertEqual(str(self.item_two.quantity_in_stock), "4.000")
        self.assertEqual(str(sale.subtotal), "15.00")

    def test_admin_can_save_an_already_edited_pos_sale_again(self):
        self.client.force_login(self.user)
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "20.00",
                "notes": "First save",
                "cart": json.dumps([{"id": self.item.pk, "quantity": 2}]),
            },
        )
        self.assertEqual(checkout_response.status_code, 302)
        sale = Sale.objects.latest("created_at")

        first_edit = self.client.post(
            reverse("inventory-sale-update", args=[sale.pk]),
            {
                "sale_date": timezone.localdate().isoformat(),
                "customer_name": "First edit",
                "customer_phone": "",
                "customer_email": "",
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "20.00",
                "notes": "Edited once",
                "items_payload": json.dumps(
                    [
                        {
                            "item_id": self.item.pk,
                            "quantity": 2,
                            "unit_price": "10.00",
                        },
                    ]
                ),
            },
            follow=True,
        )
        self.assertEqual(first_edit.status_code, 200)

        second_edit = self.client.post(
            reverse("inventory-sale-update", args=[sale.pk]),
            {
                "sale_date": timezone.localdate().isoformat(),
                "customer_name": "Second edit",
                "customer_phone": "",
                "customer_email": "",
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "1.00",
                "discount_amount": "0.00",
                "amount_paid": "21.00",
                "notes": "Edited twice",
                "items_payload": json.dumps(
                    [
                        {
                            "item_id": self.item.pk,
                            "quantity": 2,
                            "unit_price": "10.00",
                        },
                    ]
                ),
            },
            follow=True,
        )
        self.assertEqual(second_edit.status_code, 200)
        sale.refresh_from_db()
        self.assertEqual(sale.customer_name, "Second edit")
        self.assertEqual(str(sale.tax_amount), "1.00")
        self.assertEqual(str(sale.grand_total), "21.00")

    def test_duplicate_pos_sale_lines_are_merged_on_save(self):
        self.client.force_login(self.user)
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "20.00",
                "notes": "Validation test sale",
                "cart": json.dumps([{"id": self.item.pk, "quantity": 2}]),
            },
        )
        self.assertEqual(checkout_response.status_code, 302)
        sale = Sale.objects.latest("created_at")

        response = self.client.post(
            reverse("inventory-sale-update", args=[sale.pk]),
            {
                "sale_date": timezone.localdate().isoformat(),
                "customer_name": "",
                "customer_phone": "",
                "customer_email": "",
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "20.00",
                "notes": "Broken edit attempt",
                "items_payload": json.dumps(
                    [
                        {"item_id": self.item.pk, "quantity": 1, "unit_price": "10.00"},
                        {"item_id": self.item.pk, "quantity": 1, "unit_price": "10.00"},
                    ]
                ),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        sale.refresh_from_db()
        self.item.refresh_from_db()
        self.assertEqual(sale.items.count(), 1)
        merged_line = sale.items.get()
        self.assertEqual(str(merged_line.quantity), "2.000")
        self.assertEqual(str(merged_line.line_total), "20.00")
        self.assertEqual(str(sale.grand_total), "20.00")
        self.assertContains(response, "updated successfully")

    def test_sale_list_csv_export_works(self):
        Sale.objects.create(
            cashier=self.user,
            payment_method=Sale.PaymentMethod.CASH,
            subtotal="10.00",
            grand_total="10.00",
            amount_paid="10.00",
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("inventory-sales"), {"export": "csv"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("Receipt Number", response.content.decode())

    def test_inventory_dashboard_and_reports_render(self):
        self.client.force_login(self.user)
        dashboard = self.client.get(reverse("inventory-dashboard"))
        self.assertEqual(dashboard.status_code, 200)
        self.assertContains(dashboard, "Monitor stock levels, item movement, and point-of-sale activity across the hotel.")

        reports = self.client.get(reverse("inventory-reports"))
        self.assertEqual(reports.status_code, 200)
        self.assertContains(reports, "Generate stock and sales reports for the selected date range.")

    def test_pos_reports_export_xlsx_uses_real_sales_data(self):
        sale = Sale.objects.create(
            cashier=self.user,
            customer_name="Report Guest",
            payment_method=Sale.PaymentMethod.CARD,
            subtotal="20.00",
            grand_total="20.00",
            amount_paid="20.00",
        )
        SaleItem.objects.create(
            sale=sale,
            item=self.item,
            quantity="2.000",
            unit_price="10.00",
            line_total="20.00",
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse("inventory-pos-reports-export"), {"period": "monthly"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("POS Sales", workbook.sheetnames)
        self.assertIn("Summary", workbook.sheetnames)
        sales_sheet = workbook["POS Sales"]
        self.assertEqual(sales_sheet["A1"].value, "POS Sales Report")
        self.assertIn(sale.receipt_number, [cell.value for row in sales_sheet.iter_rows() for cell in row])

    def test_quantity_fields_render_without_trailing_zeroes(self):
        self.client.force_login(self.user)

        item_form = self.client.get(reverse("inventory-item-update", args=[self.item.pk]))
        self.assertEqual(item_form.status_code, 200)
        self.assertContains(item_form, 'value="10"', html=False)
        self.assertContains(item_form, 'value="2"', html=False)
        self.assertNotContains(item_form, 'value="10.000"', html=False)
        self.assertNotContains(item_form, 'value="2.000"', html=False)

        adjustment_form = self.client.get(reverse("inventory-item-adjust", args=[self.item.pk]))
        self.assertEqual(adjustment_form.status_code, 200)
        self.assertContains(adjustment_form, "Adjust the stock level for Mango Juice and record the reason for the change.")
        self.assertNotContains(adjustment_form, "Adjust the stock level for Mango Juice and record the reason for the change from 10.000")

        item_list = self.client.get(reverse("inventory-items"))
        self.assertEqual(item_list.status_code, 200)
        self.assertContains(item_list, "Min 2")
        self.assertNotContains(item_list, "Min 2.000")

    def test_inventory_item_delete_removes_internal_history_only_records(self):
        self.client.force_login(self.user)

        create_response = self.client.post(
            reverse("inventory-item-create"),
            {
                "name": "Delete Me",
                "category": self.category.pk,
                "subcategory": self.subcategory.pk,
                "supplier": self.supplier.pk,
                "purchase_price": "4.00",
                "selling_price": "6.00",
                "quantity_in_stock": "5.000",
                "unit_of_measure": InventoryItem.UnitOfMeasure.BOTTLE,
                "minimum_stock_threshold": "1.000",
                "description": "Temporary stock item",
                "is_active": "on",
            },
        )
        self.assertEqual(create_response.status_code, 302)

        created_item = InventoryItem.objects.get(name="Delete Me")
        self.assertTrue(InventoryTransaction.objects.filter(item=created_item).exists())

        delete_response = self.client.post(reverse("inventory-item-delete", args=[created_item.pk]))
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(InventoryItem.objects.filter(pk=created_item.pk).exists())
        self.assertFalse(InventoryTransaction.objects.filter(item_id=created_item.pk).exists())

    def test_inventory_item_delete_stays_blocked_when_sale_history_exists(self):
        self.client.force_login(self.user)
        cart = json.dumps([{"id": self.item.pk, "quantity": 1}])
        checkout_response = self.client.post(
            reverse("inventory-pos-checkout"),
            {
                "payment_method": Sale.PaymentMethod.CASH,
                "tax_amount": "0.00",
                "discount_amount": "0.00",
                "amount_paid": "10.00",
                "notes": "Delete protection test",
                "cart": cart,
            },
        )
        self.assertEqual(checkout_response.status_code, 302)

        delete_response = self.client.post(reverse("inventory-item-delete", args=[self.item.pk]), follow=True)
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(InventoryItem.objects.filter(pk=self.item.pk).exists())
        messages = [str(message) for message in delete_response.context["messages"]]
        self.assertIn("This item is linked to stock history or sales and cannot be deleted.", messages)
