from datetime import time, timedelta

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.db.models import Q

from accounts.models import Rota
from accounts.decorators import group_required
from bookings.models import Booking
from rooms.models import Room
from shifts.forms import ShiftHandoverForm, ShiftHandoverUpdateForm, RosterFilterForm
from shifts.models import ShiftHandover, ShiftHandoverUpdate


@group_required("Admin", "Receptionist", module="handovers")
def handover_list(request):
    handovers = ShiftHandover.objects.select_related("prepared_by")
    return render(request, "shifts/handover_list.html", {"handovers": handovers})


@group_required("Admin", "Receptionist", module="handovers", action="create")
def handover_create(request):
    if request.method == "POST":
        form = ShiftHandoverForm(request.POST)
        if form.is_valid():
            handover = form.save(commit=False)
            handover.prepared_by = request.user
            _populate_shift_metrics(handover)
            handover.save()
            messages.success(request, "Shift handover report created.")
            return redirect("handover-detail", pk=handover.pk)
    else:
        now = timezone.localtime()
        default_start = now.replace(minute=0, second=0, microsecond=0)
        form = ShiftHandoverForm(
            initial={
                "started_at": (default_start - timedelta(hours=8)).strftime(
                    "%Y-%m-%dT%H:%M"
                ),
                "ended_at": default_start.strftime("%Y-%m-%dT%H:%M"),
            }
        )
    return render(
        request,
        "shifts/handover_form.html",
        {"form": form, "title": "End Shift and Create Handover"},
    )


@group_required("Admin", "Receptionist", module="handovers", action={"GET": "view", "POST": "create"})
def handover_detail(request, pk):
    handover = get_object_or_404(
        ShiftHandover.objects.select_related("prepared_by").prefetch_related("updates__author"),
        pk=pk,
    )
    if request.method == "POST":
        update_form = ShiftHandoverUpdateForm(request.POST)
        if update_form.is_valid():
            update = update_form.save(commit=False)
            update.handover = handover
            update.author = request.user
            update.save()
            messages.success(request, "Handover update added.")
            return redirect("handover-detail", pk=handover.pk)
    else:
        update_form = ShiftHandoverUpdateForm()

    return render(
        request,
        "shifts/handover_detail.html",
        {
            "handover": handover,
            "update_form": update_form,
            "updates": handover.updates.select_related("author"),
        },
    )


@group_required("Admin", "Manager", module="handovers")
def roster_report(request):
    """List and filter duty rosters for reporting and export."""
    form = RosterFilterForm(request.GET)
    rosters = (
        Rota.objects.select_related("employee")
        .filter(employee__isnull=False)
        .order_by("-period_start", "employee__last_name", "employee__first_name")
    )

    if form.is_valid():
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        department = form.cleaned_data.get("department")
        shift = form.cleaned_data.get("shift")
        employee = form.cleaned_data.get("employee")
        role = form.cleaned_data.get("role")
        status = form.cleaned_data.get("status")

        if start_date:
            rosters = rosters.filter(period_end__gte=start_date)
        if end_date:
            rosters = rosters.filter(period_start__lte=end_date)
        if department:
            rosters = rosters.filter(employee__department__iexact=department)
        if shift:
            start_value, _, end_value = shift.partition("-")
            rosters = rosters.filter(
                opening_time=time.fromisoformat(start_value),
                closing_time=time.fromisoformat(end_value),
            )
        if employee:
            rosters = rosters.filter(employee=employee)
        if role:
            rosters = rosters.filter(
                Q(employee__job_title__icontains=role)
                | Q(employee__position__icontains=role)
                | Q(period__icontains=role)
            )
        if status:
            rosters = rosters.filter(employee__employment_status=status)

    context = {
        "form": form,
        "filter_form": form,
        "rosters": rosters,
    }
    return render(request, "shifts/roster_report.html", context)


@group_required("Admin", "Manager", module="handovers")
def roster_detail(request, pk):
    """Display a detailed account for a weekly duty roster."""
    rota = get_object_or_404(Rota.objects.select_related("employee"), pk=pk)
    daily_roster = _weekly_rota_daily_breakdown(rota)

    context = {
        "rota": rota,
        "daily_roster": daily_roster,
    }
    return render(request, "shifts/roster_detail.html", context)


@group_required("Admin", "Manager", module="handovers", action="export")
def roster_export_excel(request):
    """Export roster data to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl is not installed. Please install it to export to Excel.")
        return redirect("roster-report")

    form = RosterFilterForm(request.GET)
    rosters = (
        Rota.objects.select_related("employee")
        .filter(employee__isnull=False)
        .order_by("period_start", "employee__last_name", "employee__first_name")
    )

    if form.is_valid():
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        department = form.cleaned_data.get("department")
        shift = form.cleaned_data.get("shift")
        employee = form.cleaned_data.get("employee")
        role = form.cleaned_data.get("role")
        status = form.cleaned_data.get("status")

        if start_date:
            rosters = rosters.filter(period_end__gte=start_date)
        if end_date:
            rosters = rosters.filter(period_start__lte=end_date)
        if department:
            rosters = rosters.filter(employee__department__iexact=department)
        if shift:
            start_value, _, end_value = shift.partition("-")
            rosters = rosters.filter(
                opening_time=time.fromisoformat(start_value),
                closing_time=time.fromisoformat(end_value),
            )
        if employee:
            rosters = rosters.filter(employee=employee)
        if role:
            rosters = rosters.filter(
                Q(employee__job_title__icontains=role)
                | Q(employee__position__icontains=role)
                | Q(period__icontains=role)
            )
        if status:
            rosters = rosters.filter(employee__employment_status=status)

    export_rows = []
    for rota in rosters:
        for day in _weekly_rota_daily_breakdown(rota):
            export_rows.append(
                {
                    "day_name": day["day_name"],
                    "date": day["date"],
                    "employee": rota.employee.full_name,
                    "department": rota.employee.department or "-",
                    "role": rota.employee.job_title or rota.employee.get_position_display(),
                    "start": day["start_time"].strftime("%H:%M") if day["start_time"] else "-",
                    "finish": day["end_time"].strftime("%H:%M") if day["end_time"] else "-",
                    "hours": day["hours"],
                }
            )

    export_rows.sort(
        key=lambda row: (
            row["date"],
            row["employee"].split()[-1].lower() if row["employee"] else "",
            row["employee"].lower(),
        )
    )

    week_reference = None
    if export_rows:
        week_reference = export_rows[0]["date"]
    elif form.is_valid() and form.cleaned_data.get("start_date"):
        week_reference = form.cleaned_data["start_date"]
    else:
        week_reference = timezone.localdate()
    week_start = week_reference - timedelta(days=week_reference.weekday())

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duty Roster"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers begin on the first row so exports are straightforward to consume.
    row = 1
    headers = ["Day", "Date", "Employee", "Department", "Role", "Start", "Finish", "Hours"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Data
    row = 2
    for export_row in export_rows:
        ws.cell(row=row, column=1).value = export_row["day_name"]
        ws.cell(row=row, column=2).value = export_row["date"].strftime("%d/%m/%Y")
        ws.cell(row=row, column=3).value = export_row["employee"]
        ws.cell(row=row, column=4).value = export_row["department"]
        ws.cell(row=row, column=5).value = export_row["role"]
        ws.cell(row=row, column=6).value = export_row["start"]
        ws.cell(row=row, column=7).value = export_row["finish"]
        ws.cell(row=row, column=8).value = export_row["hours"]

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12

    # Generate response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="duty-roster-week-{week_start.strftime("%d-%m-%Y")}.xlsx"'
    )
    wb.save(response)
    return response


def _populate_shift_metrics(handover):
    start = handover.started_at
    end = handover.ended_at
    start_date = timezone.localtime(start).date()
    end_date = timezone.localtime(end).date()

    handover.check_ins_count = Booking.objects.filter(check_in__range=[start_date, end_date]).count()
    handover.check_outs_count = Booking.objects.filter(
        check_out__range=[start_date, end_date]
    ).count()
    handover.reservations_count = Booking.objects.filter(created_at__range=[start, end]).count()
    handover.cancellations_count = Booking.objects.filter(
        status=Booking.BookingStatus.CANCELLED,
        updated_at__range=[start, end],
    ).count()
    handover.maintenance_rooms_count = Room.objects.filter(
        status=Room.RoomStatus.MAINTENANCE
    ).count()
    handover.cleaning_rooms_count = Room.objects.filter(status=Room.RoomStatus.CLEANING).count()


def _weekly_rota_daily_breakdown(rota):
    if not rota.period_start or not rota.period_end:
        return []

    day_count = (rota.period_end - rota.period_start).days + 1
    daily_hours = rota.daily_hours if rota.opening_time and rota.closing_time else 0
    rows = []
    for day_index in range(day_count):
        current_day = rota.period_start + timedelta(days=day_index)
        rows.append(
            {
                "date": current_day,
                "day_name": current_day.strftime("%A"),
                "start_time": rota.opening_time,
                "end_time": rota.closing_time,
                "hours": daily_hours,
            }
        )
    return rows
