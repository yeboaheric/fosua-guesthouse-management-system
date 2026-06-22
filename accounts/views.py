import csv
import json
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote_plus, unquote_plus

from django.contrib import messages
from django.contrib.auth.models import Group
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.contenttypes.models import ContentType
from django.db import IntegrityError, transaction
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Max, Q, Sum, Value
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.timesince import timesince
from django.views.decorators.http import require_POST

from accounts.audit import log_audit_event
from accounts.decorators import group_required
from accounts.formatting import format_quantity
from accounts.forms import (
    AttendanceRecordForm,
    DisciplinaryRecordForm,
    EmployeeDocumentForm,
    EmployeeForm,
    EmployeeQualificationForm,
    ExpenseForm,
    EmploymentHistoryForm,
    LeaveRequestForm,
    OwnerWithdrawalForm,
    PayrollRecordForm,
    PerformanceReviewForm,
    RoleCreateForm,
    RolePermissionForm,
    RotaForm,
    StaffRoleForm,
    StaffUserForm,
    TrainingRecordForm,
)
from accounts.models import (
    AttendanceRecord,
    AuditLog,
    DisciplinaryRecord,
    Employee,
    EmployeeDocument,
    EmployeeQualification,
    Expense,
    EmploymentHistoryEntry,
    LeaveRequest,
    PayrollRecord,
    PerformanceReview,
    Rota,
    RolePermission,
    LEAVE_TYPE_TO_EMPLOYMENT_STATUS,
    TrainingRecord,
    Notification,
    OwnerWithdrawal,
    StatusHistory,
)
from accounts.permissions import (
    ACTION_CHOICES,
    ACCESS_MODULE_CHOICES,
    default_permissions_for_role,
    seed_default_role_names,
    user_has_permission,
    user_is_admin_role,
)
from accounts.reporting import (
    booking_payment_queryset,
    booking_revenue_queryset,
    completed_pos_sales_queryset,
    daily_booking_revenue_map,
    daily_net_revenue_map,
    daily_owner_withdrawals_map,
    daily_total_revenue_map,
    event_payment_queryset,
    event_revenue_queryset,
    filter_queryset_for_local_datetime_bounds,
    money_total as shared_money_total,
    normalize_date_range,
    owner_withdrawals_queryset,
    report_window_for_period as shared_report_window_for_period,
    revenue_components,
)
from bookings.models import (
    Booking,
    EventBooking,
    EventPayment,
    Payment,
    booking_occupied_days_in_range,
    booking_occupies_day,
)
from guests.models import Guest
from inventory.models import InventoryItem, Sale
from rooms.models import HousekeepingItem, HousekeepingItemLog, MaintenanceRequest, Room


class FosuaLoginView(LoginView):
    template_name = "accounts/login.html"
    redirect_authenticated_user = True


class FosuaLogoutView(LogoutView):
    pass


def home_redirect(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    return redirect("login")


def _is_group_member(user, group_name):
    return user.groups.filter(name=group_name).exists()


def _user_can_access_module(user, module_name):
    return user_has_permission(user, module_name, "view")


def _permission_defaults_from_actions(actions):
    return {
        "can_view": "view" in actions,
        "can_create": "create" in actions,
        "can_edit": "edit" in actions,
        "can_delete": "delete" in actions,
        "can_approve": "approve" in actions,
        "can_export": "export" in actions,
        "can_print": "print" in actions,
        "can_manage": "manage" in actions,
    }


def _seed_default_roles():
    role_names = set(seed_default_role_names()) | {"Admin"}
    for role_name in role_names:
        role, _ = Group.objects.get_or_create(name=role_name)
        if role.role_permissions.exists():
            continue
        if role.name == "Admin":
            preset = {module: {action for action, _ in ACTION_CHOICES} for module, _ in ACCESS_MODULE_CHOICES}
        else:
            preset = default_permissions_for_role(role.name)
        for module_name, actions in preset.items():
            RolePermission.objects.update_or_create(
                role=role,
                module=module_name,
                defaults=_permission_defaults_from_actions(actions),
            )


def _month_start(value):
    return value.replace(day=1)


def _shift_month(value, delta):
    month_index = (value.month - 1) + delta
    year = value.year + month_index // 12
    month = (month_index % 12) + 1
    return date(year, month, 1)


def _week_start(value):
    return value - timedelta(days=value.weekday())


def _dashboard_chart_series(today):
    daily_days = _report_days(today - timedelta(days=6), today)
    weekly_starts = [_week_start(today) - timedelta(weeks=index) for index in range(7, -1, -1)]
    monthly_starts = [_shift_month(_month_start(today), -index) for index in range(5, -1, -1)]

    earliest_required_day = min(daily_days[0], weekly_starts[0], monthly_starts[0])

    booking_counts_daily = defaultdict(int)
    booking_counts_weekly = defaultdict(int)
    booking_counts_monthly = defaultdict(int)
    for created_at in Booking.objects.filter(created_at__date__gte=earliest_required_day).values_list("created_at__date", flat=True):
        booking_counts_daily[created_at] += 1
        booking_counts_weekly[_week_start(created_at)] += 1
        booking_counts_monthly[_month_start(created_at)] += 1

    revenue_daily = defaultdict(Decimal)
    revenue_weekly = defaultdict(Decimal)
    revenue_monthly = defaultdict(Decimal)
    for revenue_day, amount in daily_total_revenue_map(earliest_required_day, today).items():
        revenue_daily[revenue_day] += amount or Decimal("0")
        revenue_weekly[_week_start(revenue_day)] += amount or Decimal("0")
        revenue_monthly[_month_start(revenue_day)] += amount or Decimal("0")

    return {
        "daily": {
            "labels": [day.strftime("%d %b") for day in daily_days],
            "bookings": [booking_counts_daily.get(day, 0) for day in daily_days],
            "revenue": [float(revenue_daily.get(day, Decimal("0"))) for day in daily_days],
        },
        "weekly": {
            "labels": [f"Week of {day.strftime('%d %b')}" for day in weekly_starts],
            "bookings": [booking_counts_weekly.get(day, 0) for day in weekly_starts],
            "revenue": [float(revenue_weekly.get(day, Decimal("0"))) for day in weekly_starts],
        },
        "monthly": {
            "labels": [day.strftime("%b %Y") for day in monthly_starts],
            "bookings": [booking_counts_monthly.get(day, 0) for day in monthly_starts],
            "revenue": [float(revenue_monthly.get(day, Decimal("0"))) for day in monthly_starts],
        },
    }


def _dashboard_snapshot(user=None):
    today = timezone.localdate()
    now = timezone.localtime()
    month_start, month_end = shared_report_window_for_period("monthly", today)

    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status=Room.RoomStatus.AVAILABLE).count()
    occupied_rooms = Room.objects.filter(status=Room.RoomStatus.OCCUPIED).count()
    cleaning_rooms = Room.objects.filter(status=Room.RoomStatus.CLEANING).count()
    maintenance_rooms = Room.objects.filter(status=Room.RoomStatus.MAINTENANCE).count()
    active_bookings = Booking.objects.filter(
        status__in=[
            Booking.BookingStatus.PENDING,
            Booking.BookingStatus.CONFIRMED,
            Booking.BookingStatus.CHECKED_IN,
        ]
    ).count()
    total_bookings = Booking.objects.count()
    total_guests = Guest.objects.count()
    total_staff = Employee.objects.exclude(employment_status="terminated").count()
    staff_on_duty = Rota.objects.filter(
        period_start__lte=today,
        period_end__gte=today,
        employee__employment_status="active",
    ).count()

    daily_revenue_components = revenue_components(today, today)
    monthly_revenue_components = revenue_components(month_start, month_end)

    bookings_with_balance = Booking.objects.annotate(
        paid_total=Coalesce(
            Sum("payments__amount"),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
        ),
        balance=ExpressionWrapper(
            F("total_amount") - F("paid_total"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
    )
    pending_room_balances = bookings_with_balance.filter(balance__gt=0).count()

    event_bookings_with_balance = EventBooking.objects.annotate(
        paid_total=Coalesce(
            Sum("payments__amount"),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
        ),
        balance=ExpressionWrapper(
            F("total_amount") - F("paid_total"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
    )
    pending_event_balances = event_bookings_with_balance.filter(balance__gt=0).count()

    occupancy_rate = round((occupied_rooms / total_rooms) * 100, 1) if total_rooms else 0
    last_7_rows = _daily_report_rows(today - timedelta(days=6), today, total_rooms or 1)
    last_7_revenue_map = daily_total_revenue_map(today - timedelta(days=6), today)

    upcoming_window = now + timedelta(hours=1)
    today_check_ins = Booking.objects.filter(
        check_in=today,
        status__in=[
            Booking.BookingStatus.PENDING,
            Booking.BookingStatus.CONFIRMED,
        ],
    )
    today_check_outs = Booking.objects.filter(
        check_out=today,
        status__in=[
            Booking.BookingStatus.CHECKED_IN,
            Booking.BookingStatus.CONFIRMED,
        ],
    )
    upcoming_check_ins = Booking.objects.filter(
        check_in=today,
        check_in_time__gte=now.time(),
        check_in_time__lte=upcoming_window.time(),
        status__in=[
            Booking.BookingStatus.PENDING,
            Booking.BookingStatus.CONFIRMED,
        ],
    )
    upcoming_check_outs = Booking.objects.filter(
        check_out=today,
        check_out_time__gte=now.time(),
        check_out_time__lte=upcoming_window.time(),
        status__in=[Booking.BookingStatus.CHECKED_IN],
    )
    reserved_rooms = Booking.objects.filter(status=Booking.BookingStatus.CONFIRMED).count()
    chart_series = _dashboard_chart_series(today)
    room_status_breakdown = []
    for item in [
            {"label": "Available", "count": available_rooms},
            {"label": "Occupied", "count": occupied_rooms},
            {"label": "Cleaning", "count": cleaning_rooms},
            {"label": "Maintenance", "count": maintenance_rooms},
    ]:
        percentage = round((item["count"] / total_rooms) * 100, 1) if total_rooms else 0
        room_status_breakdown.append({**item, "percentage": percentage})
    occupancy_chart_breakdown = [
        {"label": "Occupied", "count": occupied_rooms, "percentage": round((occupied_rooms / total_rooms) * 100, 1) if total_rooms else 0},
        {"label": "Available", "count": available_rooms, "percentage": round((available_rooms / total_rooms) * 100, 1) if total_rooms else 0},
        {"label": "Under maintenance", "count": maintenance_rooms, "percentage": round((maintenance_rooms / total_rooms) * 100, 1) if total_rooms else 0},
    ]

    snapshot = {
        "today": today,
        "month_start": month_start,
        "month_end": month_end,
        "total_rooms": total_rooms,
        "available_rooms": available_rooms,
        "occupied_rooms": occupied_rooms,
        "cleaning_rooms": cleaning_rooms,
        "maintenance_rooms": maintenance_rooms,
        "active_bookings": active_bookings,
        "total_bookings": total_bookings,
        "total_guests": total_guests,
        "total_staff": total_staff,
        "staff_on_duty": staff_on_duty,
        "booking_revenue_today": daily_revenue_components["booking_revenue"],
        "event_revenue_today": daily_revenue_components["event_revenue"],
        "pos_sales_today": daily_revenue_components["pos_sales"],
        "daily_revenue": daily_revenue_components["total_revenue"],
        "booking_revenue_month": monthly_revenue_components["booking_revenue"],
        "event_revenue_month": monthly_revenue_components["event_revenue"],
        "pos_sales_month": monthly_revenue_components["pos_sales"],
        "monthly_revenue": monthly_revenue_components["total_revenue"],
        "pending_room_balances": pending_room_balances,
        "pending_event_balances": pending_event_balances,
        "pending_payments": pending_room_balances + pending_event_balances,
        "occupancy_rate": occupancy_rate,
        "reserved_rooms": reserved_rooms,
        "today_check_ins": today_check_ins,
        "today_check_outs": today_check_outs,
        "upcoming_check_ins": upcoming_check_ins,
        "upcoming_check_outs": upcoming_check_outs,
        "room_status_breakdown": room_status_breakdown,
        "recent_activity": _recent_activity_feed(),
        "recent_bookings": Booking.objects.select_related("guest", "room").order_by("-created_at")[:6],
        "recent_payments": Payment.objects.select_related("booking__guest", "booking__room", "received_by").order_by("-paid_at")[:6],
        "recent_event_bookings": EventBooking.objects.select_related("guest").order_by("-created_at")[:6],
        "dashboard_occupancy_breakdown": occupancy_chart_breakdown,
        "dashboard_chart_series_json": json.dumps(chart_series),
        "dashboard_occupancy_json": json.dumps(occupancy_chart_breakdown),
        "chart_labels_json": json.dumps([row["date"] for row in last_7_rows]),
        "revenue_data_json": json.dumps([float(last_7_revenue_map.get(date.fromisoformat(row["date"]), Decimal("0"))) for row in last_7_rows]),
        "occupancy_data_json": json.dumps([row["occupied_rooms"] for row in last_7_rows]),
    }

    if user and _user_can_access_module(user, "notifications"):
        _generate_booking_reminder_notifications(user)
        snapshot["unread_notifications_count"] = Notification.objects.filter(
            user=user,
            read_at__isnull=True,
        ).count()
    else:
        snapshot["unread_notifications_count"] = 0

    snapshot["dashboard_summary_cards"] = [
        {
            "label": "Occupancy rate",
            "value": _display_percent(occupancy_rate),
            "meta": "Rooms in active use right now",
            "href": reverse("room-availability"),
        },
        {
            "label": "Available rooms",
            "value": available_rooms,
            "meta": "Ready for new bookings",
            "href": reverse("room-list"),
        },
        {
            "label": "Active bookings",
            "value": active_bookings,
            "meta": "Confirmed or in-house stays",
            "href": reverse("booking-list"),
        },
        {
            "label": "Staff on duty",
            "value": staff_on_duty,
            "meta": "Rotas covering today",
            "href": reverse("hr-rota-list"),
        },
        {
            "label": "Daily revenue",
            "value": _display_money(daily_revenue_components["total_revenue"]),
            "meta": "Reservations, events, and POS sales today",
            "href": reverse("analytics-center"),
        },
        {
            "label": "Monthly revenue",
            "value": _display_money(monthly_revenue_components["total_revenue"]),
            "meta": "Current month bookings, events, and POS sales",
            "href": reverse("analytics-center"),
        },
        {
            "label": "Pending balances",
            "value": pending_room_balances + pending_event_balances,
            "meta": "Open booking and event accounts",
            "href": reverse("payments-center"),
        },
        {
            "label": "Total guests",
            "value": total_guests,
            "meta": "Guest profiles stored in the system",
            "href": reverse("guest-list"),
        },
        {
            "label": "Reserved rooms",
            "value": reserved_rooms,
            "meta": "Confirmed reservations",
            "href": reverse("booking-list"),
        },
        {
            "label": "Arrivals next hour",
            "value": upcoming_check_ins.count(),
            "meta": "Check-ins starting soon",
            "href": reverse("booking-list"),
        },
        {
            "label": "Departures next hour",
            "value": upcoming_check_outs.count(),
            "meta": "Check-outs due soon",
            "href": reverse("booking-list"),
        },
        {
            "label": "Unread alerts",
            "value": snapshot["unread_notifications_count"],
            "meta": "Notifications waiting for review",
            "href": reverse("notifications-center"),
        },
    ]

    return snapshot


def _generate_booking_reminder_notifications(user):
    if not user:
        return

    now = timezone.localtime()
    upcoming_window = now + timedelta(hours=1)
    today = now.date()

    check_ins = Booking.objects.filter(
        check_in=today,
        check_in_time__gte=now.time(),
        check_in_time__lte=upcoming_window.time(),
        status__in=[Booking.BookingStatus.PENDING, Booking.BookingStatus.CONFIRMED],
    )
    check_outs = Booking.objects.filter(
        check_out=today,
        check_out_time__gte=now.time(),
        check_out_time__lte=upcoming_window.time(),
        status__in=[Booking.BookingStatus.CHECKED_IN],
    )

    for booking in check_ins:
        title = f"Upcoming check-in for {booking.guest} in Room {booking.room.room_number}"
        if Notification.objects.filter(user=user, title=title).exists():
            continue
        Notification.objects.create(
            user=user,
            title=title,
            message=(
                f"Room {booking.room.room_number} is due at "
                f"{booking.check_in_time.strftime('%H:%M')} on {booking.check_in}."
            ),
            link=reverse("booking-detail", args=[booking.pk]),
            level=Notification.Level.INFO,
        )

    for booking in check_outs:
        title = f"Upcoming check-out for {booking.guest} in Room {booking.room.room_number}"
        if Notification.objects.filter(user=user, title=title).exists():
            continue
        Notification.objects.create(
            user=user,
            title=title,
            message=(
                f"Room {booking.room.room_number} is due for check-out at "
                f"{booking.check_out_time.strftime('%H:%M')} on {booking.check_out}."
            ),
            link=reverse("booking-detail", args=[booking.pk]),
            level=Notification.Level.WARNING,
        )


@login_required
def dashboard(request):
    if request.user.is_superuser or _is_group_member(request.user, "Admin"):
        return redirect("admin-dashboard")

    if _is_group_member(request.user, "Receptionist"):
        return redirect("reception-dashboard")

    return render(request, "accounts/dashboard.html")


@group_required("Admin")
def admin_dashboard(request):
    context = _dashboard_snapshot(user=request.user)
    context["role_label"] = "Admin"
    return render(request, "accounts/admin_dashboard.html", context)


@group_required("Admin")
def admin_dashboard_activity_feed(request):
    return JsonResponse({"items": _recent_activity_feed(limit=12)})


@group_required("Receptionist", "Admin", module="dashboard")
def reception_dashboard(request):
    context = _dashboard_snapshot(user=request.user)
    context["role_label"] = "Reception"
    return render(request, "accounts/reception_dashboard.html", context)


@login_required
def global_search(request):
    query = request.GET.get("q", "").strip()
    context = {"query": query}
    if not query:
        return render(request, "accounts/global_search.html", context)

    user = request.user
    context.update(
        {
            "guest_results": Guest.objects.filter(
                Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
                | Q(phone_number__icontains=query)
                | Q(email__icontains=query)
                | Q(ghana_card_number__icontains=query)
            ).order_by("last_name", "first_name")[:8]
            if _user_can_access_module(user, "guests")
            else [],
            "room_results": Room.objects.filter(
                Q(room_number__icontains=query)
                | Q(notes__icontains=query)
                | Q(room_type__icontains=query)
            ).order_by("room_number")[:8]
            if _user_can_access_module(user, "rooms")
            else [],
            "booking_results": Booking.objects.select_related("guest", "room").filter(
                Q(guest__first_name__icontains=query)
                | Q(guest__last_name__icontains=query)
                | Q(room__room_number__icontains=query)
                | Q(status__icontains=query)
            ).order_by("-created_at")[:8]
            if _user_can_access_module(user, "reservations")
            else [],
            "inventory_results": InventoryItem.objects.select_related("category", "subcategory", "supplier").filter(
                Q(name__icontains=query)
                | Q(category__name__icontains=query)
                | Q(subcategory__name__icontains=query)
                | Q(supplier__name__icontains=query)
            ).order_by("name")[:8]
            if _user_can_access_module(user, "inventory")
            else [],
            "sale_results": Sale.objects.select_related("cashier").filter(
                Q(receipt_number__icontains=query)
                | Q(customer_name__icontains=query)
                | Q(customer_phone__icontains=query)
                | Q(customer_email__icontains=query)
            ).order_by("-created_at")[:8]
            if _user_can_access_module(user, "pos")
            else [],
            "payment_results": Payment.objects.select_related("booking__guest", "booking__room").filter(
                Q(reference__icontains=query)
                | Q(notes__icontains=query)
                | Q(booking__guest__first_name__icontains=query)
                | Q(booking__guest__last_name__icontains=query)
                | Q(booking__room__room_number__icontains=query)
            ).order_by("-paid_at")[:8]
            if _user_can_access_module(user, "payments")
            else [],
            "employee_results": Employee.objects.filter(
                Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
                | Q(position__icontains=query)
                | Q(ghana_card_number__icontains=query)
            ).order_by("last_name", "first_name")[:8]
            if _user_can_access_module(user, "staff_management")
            else [],
            "event_results": EventBooking.objects.select_related("guest").filter(
                Q(event_title__icontains=query)
                | Q(event_space_name__icontains=query)
                | Q(guest__first_name__icontains=query)
                | Q(guest__last_name__icontains=query)
            ).order_by("-created_at")[:8]
            if _user_can_access_module(user, "services")
            else [],
            "event_payment_results": EventPayment.objects.select_related(
                "event_booking__guest"
            ).filter(
                Q(reference__icontains=query)
                | Q(notes__icontains=query)
                | Q(event_booking__event_title__icontains=query)
                | Q(event_booking__guest__first_name__icontains=query)
                | Q(event_booking__guest__last_name__icontains=query)
            ).order_by("-paid_at")[:8]
            if _user_can_access_module(user, "payments")
            else [],
        }
    )
    return render(request, "accounts/global_search.html", context)


@group_required("Admin", "Receptionist", module="payments")
def payments_center(request):
    today = timezone.localdate()
    month_start = today.replace(day=1)
    query = request.GET.get("q", "").strip()
    payment_method = request.GET.get("method", "")
    payment_scope = request.GET.get("scope", "all")

    room_payments = Payment.objects.select_related("booking__guest", "booking__room", "received_by").order_by("-paid_at")
    event_payments = EventPayment.objects.select_related("event_booking__guest", "received_by").order_by("-paid_at")

    if query:
        room_payments = room_payments.filter(
            Q(reference__icontains=query)
            | Q(notes__icontains=query)
            | Q(booking__guest__first_name__icontains=query)
            | Q(booking__guest__last_name__icontains=query)
            | Q(booking__room__room_number__icontains=query)
        )
        event_payments = event_payments.filter(
            Q(reference__icontains=query)
            | Q(notes__icontains=query)
            | Q(event_booking__event_title__icontains=query)
            | Q(event_booking__guest__first_name__icontains=query)
            | Q(event_booking__guest__last_name__icontains=query)
            | Q(event_booking__event_space_name__icontains=query)
        )

    if payment_method:
        room_payments = room_payments.filter(method=payment_method)
        event_payments = event_payments.filter(method=payment_method)

    if payment_scope == "room":
        event_payments = event_payments.none()
    elif payment_scope == "event":
        room_payments = room_payments.none()

    room_payments = room_payments[:20]
    event_payments = event_payments[:20]

    context = {
        "today_total": Payment.objects.filter(paid_at__date=today).aggregate(
            total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)))
        )["total"]
        + EventPayment.objects.filter(paid_at__date=today).aggregate(
            total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)))
        )["total"],
        "month_total": Payment.objects.filter(paid_at__date__gte=month_start).aggregate(
            total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)))
        )["total"]
        + EventPayment.objects.filter(paid_at__date__gte=month_start).aggregate(
            total=Coalesce(Sum("amount"), Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)))
        )["total"],
        "room_payments": room_payments,
        "event_payments": event_payments,
        "query": query,
        "payment_method": payment_method,
        "payment_scope": payment_scope,
        "payment_methods": Payment.PaymentMethod.choices,
        "can_manage_payment_records": user_is_admin_role(request.user),
    }
    return render(request, "accounts/payments_center.html", context)


def _sales_deposit_date_value(raw_value):
    raw_value = (raw_value or "").strip()
    if not raw_value:
        return None
    try:
        return date.fromisoformat(raw_value)
    except ValueError:
        return None


def _sales_deposit_summary_cards():
    today = timezone.localdate()
    week_start, week_end = shared_report_window_for_period("weekly", today)
    month_start, month_end = shared_report_window_for_period("monthly", today)
    year_start, year_end = shared_report_window_for_period("yearly", today)
    return [
        {"label": "Collected today", "value": _display_money(_money_total(owner_withdrawals_queryset(today, today), "amount"))},
        {"label": "This week", "value": _display_money(_money_total(owner_withdrawals_queryset(week_start, week_end), "amount"))},
        {"label": "This month", "value": _display_money(_money_total(owner_withdrawals_queryset(month_start, month_end), "amount"))},
        {"label": "This year", "value": _display_money(_money_total(owner_withdrawals_queryset(year_start, year_end), "amount"))},
    ]


def _sales_deposit_filtered_queryset(request):
    query = request.GET.get("q", "").strip()
    start_date = _sales_deposit_date_value(request.GET.get("start_date"))
    end_date = _sales_deposit_date_value(request.GET.get("end_date"))
    if start_date and end_date:
        start_date, end_date = normalize_date_range(start_date, end_date)
    withdrawals = OwnerWithdrawal.objects.select_related("recorded_by").order_by("-created_at", "-pk")

    if query:
        withdrawals = withdrawals.filter(
            Q(collection_method__icontains=query)
            | Q(collected_by__icontains=query)
            | Q(recorded_by__username__icontains=query)
            | Q(recorded_by__first_name__icontains=query)
            | Q(recorded_by__last_name__icontains=query)
        )

    withdrawals = filter_queryset_for_local_datetime_bounds(
        withdrawals,
        "created_at",
        start_date,
        end_date,
    )

    return {
        "withdrawals": withdrawals,
        "query": query,
        "start_date": start_date.isoformat() if start_date else "",
        "end_date": end_date.isoformat() if end_date else "",
    }


def _sales_deposit_export_range(request):
    start_date = _sales_deposit_date_value(request.GET.get("start_date"))
    end_date = _sales_deposit_date_value(request.GET.get("end_date"))
    if start_date and end_date:
        return normalize_date_range(start_date, end_date)
    if start_date:
        return start_date, start_date
    if end_date:
        return end_date, end_date

    today = timezone.localdate()
    earliest = OwnerWithdrawal.objects.order_by("created_at").values_list("created_at", flat=True).first()
    latest = OwnerWithdrawal.objects.order_by("-created_at").values_list("created_at", flat=True).first()
    if earliest and latest:
        return timezone.localtime(earliest).date(), timezone.localtime(latest).date()
    return today, today


def _sales_deposit_export_querystring(start_date, end_date):
    query = QueryDict("", mutable=True)
    query["start_date"] = start_date.isoformat()
    query["end_date"] = end_date.isoformat()
    return query.urlencode()


def _sales_deposit_admin_redirect(request):
    messages.error(request, "Access Denied: only admin accounts can edit or delete sales deposits.")
    return redirect("sales-deposits-center")


@group_required("Admin", "Receptionist", module="payments", action={"GET": "view", "POST": "create"})
def sales_deposits_center(request):
    form = OwnerWithdrawalForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        withdrawal = form.save(commit=False)
        withdrawal.recorded_by = request.user
        withdrawal.save()
        log_audit_event(
            request=request,
            user=request.user,
            action=AuditLog.ActionType.CREATE,
            module="payments",
            object_repr=f"Sales deposit {withdrawal.pk}",
            object_id=withdrawal.pk,
            details={
                "event": "sales_deposit_created",
                "amount": str(withdrawal.amount),
                "collection_method": withdrawal.collection_method,
                "collected_by": withdrawal.collected_by,
            },
        )
        messages.success(request, "Sales deposit logged successfully.")
        return redirect("sales-deposits-center")

    filtered = _sales_deposit_filtered_queryset(request)
    export_start_date, export_end_date = _sales_deposit_export_range(request)
    return render(
        request,
        "accounts/sales_deposits_center.html",
        {
            "form": form,
            "form_title": "Log collection",
            "summary_cards": _sales_deposit_summary_cards(),
            "withdrawals": filtered["withdrawals"],
            "query": filtered["query"],
            "start_date": filtered["start_date"],
            "end_date": filtered["end_date"],
            "can_manage_sales_deposits": user_is_admin_role(request.user),
            "can_export_sales_deposits": user_has_permission(request.user, "payments", "export"),
            "export_url": (
                f"{reverse('sales-deposits-export-xlsx')}?"
                f"{_sales_deposit_export_querystring(export_start_date, export_end_date)}"
            ),
            "export_start_date": export_start_date.isoformat(),
            "export_end_date": export_end_date.isoformat(),
            "is_editing": False,
        },
    )


@group_required("Admin", "Receptionist", module="payments", action="view")
def sales_deposit_update(request, pk):
    if not user_is_admin_role(request.user):
        return _sales_deposit_admin_redirect(request)

    withdrawal = get_object_or_404(OwnerWithdrawal.objects.select_related("recorded_by"), pk=pk)
    form = OwnerWithdrawalForm(request.POST or None, instance=withdrawal)
    if request.method == "POST" and form.is_valid():
        updated_withdrawal = form.save(commit=False)
        if not updated_withdrawal.recorded_by_id:
            updated_withdrawal.recorded_by = request.user
        updated_withdrawal.save()
        log_audit_event(
            request=request,
            user=request.user,
            action=AuditLog.ActionType.UPDATE,
            module="payments",
            object_repr=f"Sales deposit {updated_withdrawal.pk}",
            object_id=updated_withdrawal.pk,
            details={
                "event": "sales_deposit_updated",
                "amount": str(updated_withdrawal.amount),
                "collection_method": updated_withdrawal.collection_method,
                "collected_by": updated_withdrawal.collected_by,
            },
        )
        messages.success(request, "Sales deposit updated successfully.")
        return redirect("sales-deposits-center")

    filtered = _sales_deposit_filtered_queryset(request)
    export_start_date, export_end_date = _sales_deposit_export_range(request)
    return render(
        request,
        "accounts/sales_deposits_center.html",
        {
            "form": form,
            "form_title": "Edit collection",
            "editing_withdrawal": withdrawal,
            "summary_cards": _sales_deposit_summary_cards(),
            "withdrawals": filtered["withdrawals"],
            "query": filtered["query"],
            "start_date": filtered["start_date"],
            "end_date": filtered["end_date"],
            "can_manage_sales_deposits": True,
            "can_export_sales_deposits": user_has_permission(request.user, "payments", "export"),
            "export_url": (
                f"{reverse('sales-deposits-export-xlsx')}?"
                f"{_sales_deposit_export_querystring(export_start_date, export_end_date)}"
            ),
            "export_start_date": export_start_date.isoformat(),
            "export_end_date": export_end_date.isoformat(),
            "is_editing": True,
        },
    )


@require_POST
@group_required("Admin", "Receptionist", module="payments", action="view")
def sales_deposit_delete(request, pk):
    if not user_is_admin_role(request.user):
        return _sales_deposit_admin_redirect(request)

    withdrawal = get_object_or_404(OwnerWithdrawal, pk=pk)
    withdrawal_id = withdrawal.pk
    withdrawal_amount = withdrawal.amount
    withdrawal.delete()
    log_audit_event(
        request=request,
        user=request.user,
        action=AuditLog.ActionType.DELETE,
        module="payments",
        object_repr=f"Sales deposit {withdrawal_id}",
        object_id=withdrawal_id,
        details={
            "event": "sales_deposit_deleted",
            "amount": str(withdrawal_amount),
        },
    )
    messages.success(request, "Sales deposit deleted successfully.")
    return redirect("sales-deposits-center")


@group_required("Admin", "Receptionist", module="payments", action="export")
def sales_deposits_export_xlsx(request):
    workbook = _create_reports_workbook_or_none(request)
    if workbook is None:
        return redirect("sales-deposits-center")

    start_date, end_date = _sales_deposit_export_range(request)
    withdrawals = list(
        owner_withdrawals_queryset(start_date, end_date)
        .select_related("recorded_by")
        .order_by("-created_at", "-pk")
    )

    sheet = workbook.active
    sheet.title = "Collections Log"
    _write_owner_withdrawals_log_sheet(sheet, withdrawals, start_date, end_date)

    summary_sheet = workbook.create_sheet(title="Financial Summary")
    _write_owner_withdrawals_summary_sheet(summary_sheet, start_date, end_date)

    log_audit_event(
        request=request,
        user=request.user,
        action=AuditLog.ActionType.EXPORT,
        module="payments",
        object_repr="Sales deposits export",
        details={
            "event": "sales_deposit_exported",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
        },
    )

    return _xlsx_response(
        workbook,
        f"owner-withdrawals-report-{start_date.strftime('%d-%m-%Y')}-to-{end_date.strftime('%d-%m-%Y')}.xlsx",
    )


def _finance_filters_querystring(filters):
    query = QueryDict("", mutable=True)
    query["period"] = filters["report_window"]["period"]
    query["start_date"] = filters["report_window"]["start_date"].isoformat()
    query["end_date"] = filters["report_window"]["end_date"].isoformat()
    if filters["query"]:
        query["q"] = filters["query"]
    if filters["category"]:
        query["category"] = filters["category"]
    return query.urlencode()


def _finance_export_querystring(filters):
    query = QueryDict("", mutable=True)
    query["period"] = filters["report_window"]["period"]
    query["start_date"] = filters["report_window"]["start_date"].isoformat()
    query["end_date"] = filters["report_window"]["end_date"].isoformat()
    return query.urlencode()


def _finance_filters_from_request(request):
    report_window = _report_window_from_request(request)
    return {
        "report_window": report_window,
        "query": request.GET.get("q", "").strip(),
        "category": request.GET.get("category", "").strip(),
    }


def _finance_category_options():
    categories = list(Expense.DEFAULT_CATEGORIES)
    existing = list(
        Expense.objects.exclude(category="")
        .values_list("category", flat=True)
        .distinct()
    )
    for category in sorted(existing):
        if category not in categories:
            categories.append(category)
    return categories


def _finance_pnl_group_labels():
    return [group_label for group_label, _categories in Expense.DEFAULT_CATEGORY_GROUPS]


def _finance_expenses_queryset(start_date, end_date, query="", category=""):
    expenses = Expense.objects.select_related("recorded_by").filter(
        date__range=[start_date, end_date]
    ).order_by("-date", "-created_at", "-pk")
    if query:
        expenses = expenses.filter(
            Q(category__icontains=query)
            | Q(description__icontains=query)
            | Q(recorded_by__username__icontains=query)
            | Q(recorded_by__first_name__icontains=query)
            | Q(recorded_by__last_name__icontains=query)
        )
    if category:
        expenses = expenses.filter(category=category)
    return expenses


def _finance_summary_cards():
    today = timezone.localdate()
    week_start, week_end = shared_report_window_for_period("weekly", today)
    month_start, month_end = shared_report_window_for_period("monthly", today)
    year_start, year_end = shared_report_window_for_period("yearly", today)
    return [
        {"label": "Expenses today", "value": _display_money(_money_total(Expense.objects.filter(date=today), "amount"))},
        {"label": "This week", "value": _display_money(_money_total(Expense.objects.filter(date__range=[week_start, week_end]), "amount"))},
        {"label": "This month", "value": _display_money(_money_total(Expense.objects.filter(date__range=[month_start, month_end]), "amount"))},
        {"label": "This year", "value": _display_money(_money_total(Expense.objects.filter(date__range=[year_start, year_end]), "amount"))},
    ]


def _finance_revenue_source_breakdown(start_date, end_date):
    revenue = revenue_components(start_date, end_date)
    return [
        {"label": "Room Bookings Revenue", "amount": Decimal(str(revenue["booking_revenue"] or 0))},
        {"label": "POS Sales Revenue", "amount": Decimal(str(revenue["pos_sales"] or 0))},
        {"label": "Event Reservations Revenue", "amount": Decimal(str(revenue["event_revenue"] or 0))},
        {"label": "Other Revenue", "amount": Decimal("0.00")},
    ]


def _finance_pnl_expense_totals(expenses_queryset):
    category_to_group = {}
    for group_label, categories in Expense.DEFAULT_CATEGORY_GROUPS:
        for category in categories:
            category_to_group[category] = group_label

    category_totals = {group_label: Decimal("0.00") for group_label in _finance_pnl_group_labels()}
    category_totals["Other"] = Decimal("0.00")
    for row in expenses_queryset.values("category").annotate(
        total=Coalesce(
            Sum("amount"),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        )
    ):
        category_name = (row["category"] or "").strip()
        total = Decimal(str(row["total"] or 0))
        group_name = category_to_group.get(category_name)
        if group_name:
            category_totals[group_name] += total
        else:
            category_totals["Other"] += total
    return category_totals


def _finance_pnl_rows(revenue_breakdown, expense_category_totals, gross_revenue, total_expenses, sales_deposits_total, net_profit):
    rows = [
        ["Room Bookings Revenue", _display_money(revenue_breakdown[0]["amount"])],
        ["POS Sales Revenue", _display_money(revenue_breakdown[1]["amount"])],
        ["Event Reservations Revenue", _display_money(revenue_breakdown[2]["amount"])],
        ["Other Revenue", _display_money(revenue_breakdown[3]["amount"])],
        ["TOTAL REVENUE", _display_money(gross_revenue)],
    ]
    for group_label in _finance_pnl_group_labels():
        rows.append([group_label, _display_money(expense_category_totals[group_label])])
    rows.extend(
        [
            ["Other Expenses", _display_money(expense_category_totals["Other"])],
            ["TOTAL EXPENSES", _display_money(total_expenses)],
            ["SALES DEPOSITS", _display_money(sales_deposits_total)],
            ["NET PROFIT / LOSS", _display_money(net_profit)],
        ]
    )
    return rows


def _finance_expense_category_breakdown(expenses_queryset):
    rows = []
    export_rows = []
    for row in expenses_queryset.values("category").annotate(
        total=Coalesce(
            Sum("amount"),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
        ),
        entries=Count("id"),
    ).order_by("-total", "category"):
        category_name = row["category"] or "Uncategorised"
        rows.append([category_name, row["entries"], _display_money(row["total"])])
        export_rows.append([category_name, row["entries"], float(row["total"] or 0)])
    return rows, export_rows


def _finance_daily_money_map_from_queryset(queryset, field_name, amount_field):
    rows = (
        queryset.values(field_name)
        .annotate(
            total=Coalesce(
                Sum(amount_field),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            )
        )
        .order_by(field_name)
    )
    return {row[field_name]: Decimal(str(row["total"] or 0)) for row in rows}


def _finance_daily_datetime_money_map_from_queryset(queryset, field_name, amount_field):
    rows = (
        queryset.annotate(day=TruncDate(field_name))
        .values("day")
        .annotate(
            total=Coalesce(
                Sum(amount_field),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            )
        )
        .order_by("day")
    )
    return {row["day"]: Decimal(str(row["total"] or 0)) for row in rows}


def _finance_daily_breakdown(start_date, end_date):
    room_map = _finance_daily_money_map_from_queryset(
        booking_revenue_queryset(start_date, end_date),
        "check_in",
        "total_amount",
    )
    pos_map = _finance_daily_datetime_money_map_from_queryset(
        completed_pos_sales_queryset(start_date, end_date),
        "created_at",
        "grand_total",
    )
    event_map = _finance_daily_datetime_money_map_from_queryset(
        event_revenue_queryset(start_date, end_date),
        "event_start",
        "total_amount",
    )
    expense_map = _finance_daily_money_map_from_queryset(
        Expense.objects.filter(date__range=[start_date, end_date]),
        "date",
        "amount",
    )
    deposit_map = daily_owner_withdrawals_map(start_date, end_date)

    rows = []
    for current_day in _report_days(start_date, end_date):
        room_total = room_map.get(current_day, Decimal("0"))
        pos_total = pos_map.get(current_day, Decimal("0"))
        event_total = event_map.get(current_day, Decimal("0"))
        other_total = Decimal("0")
        gross_total = room_total + pos_total + event_total + other_total
        expense_total = expense_map.get(current_day, Decimal("0"))
        deposit_total = deposit_map.get(current_day, Decimal("0"))
        net_total = gross_total - expense_total - deposit_total
        rows.append(
            {
                "date": current_day,
                "room_revenue": room_total,
                "pos_revenue": pos_total,
                "event_revenue": event_total,
                "other_revenue": other_total,
                "gross_revenue": gross_total,
                "expenses": expense_total,
                "sales_deposits": deposit_total,
                "net_profit": net_total,
            }
        )
    return rows


def _finance_cash_on_hand(snapshot_date):
    booking_cash = _money_total(
        Payment.objects.filter(
            method=Payment.PaymentMethod.CASH,
            paid_at__date__lte=snapshot_date,
        ),
        "amount",
    )
    event_cash = _money_total(
        EventPayment.objects.filter(
            method=EventPayment.PaymentMethod.CASH,
            paid_at__date__lte=snapshot_date,
        ),
        "amount",
    )
    pos_cash = _money_total(
        Sale.objects.filter(
            status=Sale.SaleStatus.COMPLETED,
            payment_method=Sale.PaymentMethod.CASH,
            created_at__date__lte=snapshot_date,
        ),
        "grand_total",
    )
    withdrawals = _money_total(
        OwnerWithdrawal.objects.filter(created_at__date__lte=snapshot_date),
        "amount",
    )
    cash_expenses = _money_total(
        Expense.objects.filter(
            date__lte=snapshot_date,
            payment_method=Expense.PaymentMethod.CASH,
        ),
        "amount",
    )
    return Decimal(str(booking_cash + event_cash + pos_cash - withdrawals - cash_expenses))


def _finance_retained_earnings(snapshot_date):
    system_start = date(2000, 1, 1)
    revenue = revenue_components(system_start, snapshot_date)
    expenses_total = _money_total(Expense.objects.filter(date__lte=snapshot_date), "amount")
    return Decimal(str(revenue["gross_revenue"] - expenses_total - revenue["owner_withdrawals"]))


def _finance_balance_sheet_snapshot(snapshot_date):
    inventory_items = list(InventoryItem.objects.filter(is_active=True))
    inventory_value = sum(
        (Decimal(str(item.purchase_price or 0)) * Decimal(str(item.quantity_in_stock or 0)) for item in inventory_items),
        Decimal("0.00"),
    )
    return {
        "snapshot_date": snapshot_date,
        "cash_on_hand": _finance_cash_on_hand(snapshot_date),
        "inventory_value": inventory_value,
        "inventory_item_count": len(inventory_items),
        "liabilities_note": "No liability data tracked yet",
        "retained_earnings": _finance_retained_earnings(snapshot_date),
    }


@group_required(
    "Admin",
    denied_redirect="dashboard",
    denied_message="Access Denied: Finance is available to admin accounts only.",
)
def finance_center(request):
    filters = _finance_filters_from_request(request)
    report_window = filters["report_window"]
    start_date = report_window["start_date"]
    end_date = report_window["end_date"]

    expense_form = ExpenseForm(request.POST or None, request.FILES or None)
    editing_expense = None

    if request.method == "POST" and expense_form.is_valid():
        expense = expense_form.save(commit=False)
        expense.recorded_by = request.user
        expense.save()
        log_audit_event(
            request=request,
            user=request.user,
            action=AuditLog.ActionType.CREATE,
            module="finance",
            object_repr=f"Expense {expense.pk}",
            object_id=expense.pk,
            details={"event": "expense_created", "amount": str(expense.amount), "category": expense.category},
        )
        messages.success(request, "Expense logged successfully.")
        return redirect("finance-center")

    all_range_expenses = Expense.objects.filter(date__range=[start_date, end_date])
    filtered_expenses = _finance_expenses_queryset(
        start_date,
        end_date,
        query=filters["query"],
        category=filters["category"],
    )
    revenue_breakdown = _finance_revenue_source_breakdown(start_date, end_date)
    gross_revenue = sum((row["amount"] for row in revenue_breakdown), Decimal("0.00"))
    total_expenses = _money_total(all_range_expenses, "amount")
    sales_deposits_total = _money_total(owner_withdrawals_queryset(start_date, end_date), "amount")
    net_profit = gross_revenue - total_expenses - sales_deposits_total
    expense_category_totals = _finance_pnl_expense_totals(all_range_expenses)
    expense_breakdown_rows, expense_breakdown_export_rows = _finance_expense_category_breakdown(all_range_expenses)
    balance_sheet = _finance_balance_sheet_snapshot(end_date)

    daily_rows = _finance_daily_breakdown(start_date, end_date)
    chart_series = _analytics_aggregate_points(
        [
            {
                "date": row["date"],
                "gross_revenue": row["gross_revenue"],
                "expenses": row["expenses"],
                "net_profit": row["net_profit"],
            }
            for row in daily_rows
        ],
        report_window["period"],
        start_date,
        end_date,
        sum_fields=("gross_revenue", "expenses", "net_profit"),
    )

    expense_donut_labels = [row[0] for row in expense_breakdown_export_rows] or ["No expenses"]
    expense_donut_values = [row[2] for row in expense_breakdown_export_rows] or [0]
    revenue_donut_labels = [row["label"] for row in revenue_breakdown]
    revenue_donut_values = [float(row["amount"]) for row in revenue_breakdown]
    charts = [
        _analytics_chart(
            "finance-revenue-expense-chart",
            "Revenue vs expenses",
            "line",
            chart_series["labels"],
            [
                {"label": "Gross Revenue", "data": chart_series["values"]["gross_revenue"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.12)", "fill": True, "tension": 0.35},
                {"label": "Expenses", "data": chart_series["values"]["expenses"], "borderColor": "#CFAE84", "backgroundColor": "rgba(207,174,132,0.10)", "fill": False, "tension": 0.35},
            ],
        ),
        _analytics_chart(
            "finance-net-profit-chart",
            "Net profit trend",
            "line",
            chart_series["labels"],
            [
                {"label": "Net Profit / Loss", "data": chart_series["values"]["net_profit"], "borderColor": "#3D7DFF", "backgroundColor": "rgba(61,125,255,0.12)", "fill": True, "tension": 0.35},
            ],
        ),
        _analytics_chart(
            "finance-expense-breakdown-chart",
            "Expense breakdown by category",
            "doughnut",
            expense_donut_labels,
            [
                {"label": "Expenses", "data": expense_donut_values, "backgroundColor": ["#23444B", "#CFAE84", "#3D7DFF", "#78C0A8", "#D95D39", "#6C757D", "#A06CD5"]},
            ],
        ),
        _analytics_chart(
            "finance-revenue-breakdown-chart",
            "Revenue breakdown by source",
            "doughnut",
            revenue_donut_labels,
            [
                {"label": "Revenue", "data": revenue_donut_values, "backgroundColor": ["#23444B", "#3D7DFF", "#78C0A8", "#CFAE84"]},
            ],
        ),
    ]

    revenue_breakdown_rows = [
        [row["label"], _display_money(row["amount"])]
        for row in revenue_breakdown
    ]
    pnl_rows = _finance_pnl_rows(
        revenue_breakdown,
        expense_category_totals,
        gross_revenue,
        total_expenses,
        sales_deposits_total,
        net_profit,
    )

    return render(
        request,
        "accounts/finance_center.html",
        {
            "report_window": report_window,
            "selected_period": report_window["period"],
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "query": filters["query"],
            "selected_category": filters["category"],
            "category_options": _finance_category_options(),
            "expense_form": expense_form,
            "editing_expense": editing_expense,
            "summary_cards": _finance_summary_cards(),
            "gross_revenue": gross_revenue,
            "total_expenses": total_expenses,
            "sales_deposits_total": sales_deposits_total,
            "net_profit": net_profit,
            "revenue_breakdown_rows": revenue_breakdown_rows,
            "expense_breakdown_rows": expense_breakdown_rows,
            "pnl_rows": pnl_rows,
            "balance_sheet": balance_sheet,
            "expenses": filtered_expenses,
            "charts": charts,
            "charts_json": json.dumps(charts),
            "filter_query_string": _finance_filters_querystring(filters),
            "export_url": f"{reverse('finance-export-xlsx')}?{_finance_export_querystring(filters)}",
            "custom_category_sentinel": ExpenseForm.CUSTOM_CATEGORY_VALUE,
        },
    )


@group_required(
    "Admin",
    denied_redirect="dashboard",
    denied_message="Access Denied: Finance is available to admin accounts only.",
)
def finance_expense_update(request, pk):
    expense = get_object_or_404(Expense.objects.select_related("recorded_by"), pk=pk)
    filters = _finance_filters_from_request(request)
    report_window = filters["report_window"]
    start_date = report_window["start_date"]
    end_date = report_window["end_date"]

    expense_form = ExpenseForm(request.POST or None, request.FILES or None, instance=expense)
    if request.method == "POST" and expense_form.is_valid():
        updated_expense = expense_form.save(commit=False)
        if not updated_expense.recorded_by_id:
            updated_expense.recorded_by = request.user
        updated_expense.save()
        log_audit_event(
            request=request,
            user=request.user,
            action=AuditLog.ActionType.UPDATE,
            module="finance",
            object_repr=f"Expense {updated_expense.pk}",
            object_id=updated_expense.pk,
            details={"event": "expense_updated", "amount": str(updated_expense.amount), "category": updated_expense.category},
        )
        messages.success(request, "Expense updated successfully.")
        return redirect(f"{reverse('finance-center')}?{_finance_filters_querystring(filters)}")

    all_range_expenses = Expense.objects.filter(date__range=[start_date, end_date])
    filtered_expenses = _finance_expenses_queryset(start_date, end_date, filters["query"], filters["category"])
    revenue_breakdown = _finance_revenue_source_breakdown(start_date, end_date)
    gross_revenue = sum((row["amount"] for row in revenue_breakdown), Decimal("0.00"))
    total_expenses = _money_total(all_range_expenses, "amount")
    sales_deposits_total = _money_total(owner_withdrawals_queryset(start_date, end_date), "amount")
    net_profit = gross_revenue - total_expenses - sales_deposits_total
    expense_category_totals = _finance_pnl_expense_totals(all_range_expenses)
    expense_breakdown_rows, expense_breakdown_export_rows = _finance_expense_category_breakdown(all_range_expenses)
    balance_sheet = _finance_balance_sheet_snapshot(end_date)
    daily_rows = _finance_daily_breakdown(start_date, end_date)
    chart_series = _analytics_aggregate_points(
        [
            {
                "date": row["date"],
                "gross_revenue": row["gross_revenue"],
                "expenses": row["expenses"],
                "net_profit": row["net_profit"],
            }
            for row in daily_rows
        ],
        report_window["period"],
        start_date,
        end_date,
        sum_fields=("gross_revenue", "expenses", "net_profit"),
    )
    charts = [
        _analytics_chart(
            "finance-revenue-expense-chart",
            "Revenue vs expenses",
            "line",
            chart_series["labels"],
            [
                {"label": "Gross Revenue", "data": chart_series["values"]["gross_revenue"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.12)", "fill": True, "tension": 0.35},
                {"label": "Expenses", "data": chart_series["values"]["expenses"], "borderColor": "#CFAE84", "backgroundColor": "rgba(207,174,132,0.10)", "fill": False, "tension": 0.35},
            ],
        ),
        _analytics_chart(
            "finance-net-profit-chart",
            "Net profit trend",
            "line",
            chart_series["labels"],
            [
                {"label": "Net Profit / Loss", "data": chart_series["values"]["net_profit"], "borderColor": "#3D7DFF", "backgroundColor": "rgba(61,125,255,0.12)", "fill": True, "tension": 0.35},
            ],
        ),
        _analytics_chart(
            "finance-expense-breakdown-chart",
            "Expense breakdown by category",
            "doughnut",
            [row[0] for row in expense_breakdown_export_rows] or ["No expenses"],
            [
                {"label": "Expenses", "data": [row[2] for row in expense_breakdown_export_rows] or [0], "backgroundColor": ["#23444B", "#CFAE84", "#3D7DFF", "#78C0A8", "#D95D39", "#6C757D", "#A06CD5"]},
            ],
        ),
        _analytics_chart(
            "finance-revenue-breakdown-chart",
            "Revenue breakdown by source",
            "doughnut",
            [row["label"] for row in revenue_breakdown],
            [
                {"label": "Revenue", "data": [float(row["amount"]) for row in revenue_breakdown], "backgroundColor": ["#23444B", "#3D7DFF", "#78C0A8", "#CFAE84"]},
            ],
        ),
    ]
    pnl_rows = _finance_pnl_rows(
        revenue_breakdown,
        expense_category_totals,
        gross_revenue,
        total_expenses,
        sales_deposits_total,
        net_profit,
    )
    return render(
        request,
        "accounts/finance_center.html",
        {
            "report_window": report_window,
            "selected_period": report_window["period"],
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "query": filters["query"],
            "selected_category": filters["category"],
            "category_options": _finance_category_options(),
            "expense_form": expense_form,
            "editing_expense": expense,
            "summary_cards": _finance_summary_cards(),
            "gross_revenue": gross_revenue,
            "total_expenses": total_expenses,
            "sales_deposits_total": sales_deposits_total,
            "net_profit": net_profit,
            "revenue_breakdown_rows": [[row["label"], _display_money(row["amount"])] for row in revenue_breakdown],
            "expense_breakdown_rows": expense_breakdown_rows,
            "pnl_rows": pnl_rows,
            "balance_sheet": balance_sheet,
            "expenses": filtered_expenses,
            "charts": charts,
            "charts_json": json.dumps(charts),
            "filter_query_string": _finance_filters_querystring(filters),
            "export_url": f"{reverse('finance-export-xlsx')}?{_finance_export_querystring(filters)}",
            "custom_category_sentinel": ExpenseForm.CUSTOM_CATEGORY_VALUE,
        },
    )


@require_POST
@group_required(
    "Admin",
    denied_redirect="dashboard",
    denied_message="Access Denied: Finance is available to admin accounts only.",
)
def finance_expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    expense_id = expense.pk
    expense_amount = expense.amount
    expense_category = expense.category
    expense.delete()
    log_audit_event(
        request=request,
        user=request.user,
        action=AuditLog.ActionType.DELETE,
        module="finance",
        object_repr=f"Expense {expense_id}",
        object_id=expense_id,
        details={"event": "expense_deleted", "amount": str(expense_amount), "category": expense_category},
    )
    messages.success(request, "Expense deleted successfully.")
    return redirect("finance-center")


@group_required(
    "Admin",
    denied_redirect="dashboard",
    denied_message="Access Denied: Finance is available to admin accounts only.",
)
def finance_export_xlsx(request):
    workbook = _create_reports_workbook_or_none(request)
    if workbook is None:
        return redirect("finance-center")

    report_window = _report_window_from_request(request)
    start_date = report_window["start_date"]
    end_date = report_window["end_date"]
    revenue_breakdown = _finance_revenue_source_breakdown(start_date, end_date)
    daily_rows = _finance_daily_breakdown(start_date, end_date)
    expenses = Expense.objects.select_related("recorded_by").filter(date__range=[start_date, end_date]).order_by("date", "created_at")
    expense_breakdown_rows, expense_breakdown_export_rows = _finance_expense_category_breakdown(expenses)
    deposits = list(owner_withdrawals_queryset(start_date, end_date).select_related("recorded_by").order_by("-created_at", "-pk"))
    gross_revenue = sum((row["amount"] for row in revenue_breakdown), Decimal("0.00"))
    total_expenses = _money_total(expenses, "amount")
    sales_deposits_total = _money_total(owner_withdrawals_queryset(start_date, end_date), "amount")
    net_profit = gross_revenue - total_expenses - sales_deposits_total
    expense_category_totals = _finance_pnl_expense_totals(expenses)
    balance_sheet = _finance_balance_sheet_snapshot(end_date)

    revenue_sheet = workbook.active
    revenue_sheet.title = "Revenue Breakdown"
    _write_finance_revenue_sheet(revenue_sheet, revenue_breakdown, daily_rows, start_date, end_date)

    expense_sheet = workbook.create_sheet(title="Expense Breakdown")
    _write_finance_expense_sheet(expense_sheet, expense_breakdown_export_rows, expenses, start_date, end_date)

    deposit_sheet = workbook.create_sheet(title="Sales Deposit Log")
    _write_owner_withdrawals_log_sheet(deposit_sheet, deposits, start_date, end_date)

    pnl_sheet = workbook.create_sheet(title="Profit Loss")
    _write_finance_pnl_sheet(
        pnl_sheet,
        revenue_breakdown,
        expense_category_totals,
        gross_revenue,
        total_expenses,
        sales_deposits_total,
        net_profit,
        start_date,
        end_date,
    )

    balance_sheet_ws = workbook.create_sheet(title="Balance Sheet")
    _write_finance_balance_sheet(balance_sheet_ws, balance_sheet)

    log_audit_event(
        request=request,
        user=request.user,
        action=AuditLog.ActionType.EXPORT,
        module="finance",
        object_repr="Finance workbook export",
        details={"event": "finance_exported", "start_date": start_date.isoformat(), "end_date": end_date.isoformat()},
    )
    return _xlsx_response(
        workbook,
        f"finance-report-{start_date.strftime('%d-%m-%Y')}-to-{end_date.strftime('%d-%m-%Y')}.xlsx",
    )


@group_required("Admin", "Receptionist", module="services")
def services_center(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "")
    events = EventBooking.objects.select_related("guest").order_by("-created_at")

    if query:
        events = events.filter(
            Q(event_title__icontains=query)
            | Q(event_space_name__icontains=query)
            | Q(purpose__icontains=query)
            | Q(guest__first_name__icontains=query)
            | Q(guest__last_name__icontains=query)
        )
    if status:
        events = events.filter(status=status)

    return render(
        request,
        "accounts/services_center.html",
        {
            "events": events[:12],
            "query": query,
            "selected_status": status,
            "status_choices": EventBooking.EventBookingStatus.choices,
        },
    )


@group_required(
    "Admin",
    "Receptionist",
    module="housekeeping",
    denied_redirect="dashboard",
    denied_message="You are not authorized to access Housekeeping.",
)
def housekeeping_center(request):
    report_range = request.GET.get("report", "daily")
    return redirect(f"{reverse('housekeeping-dashboard')}?report={report_range}")


@group_required("Admin", "Receptionist", module="notifications")
def notifications_center(request):
    _generate_booking_reminder_notifications(request.user)
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")
    unread_count = notifications.filter(read_at__isnull=True).count()
    return render(
        request,
        "accounts/notifications_center.html",
        {
            "notifications": notifications,
            "unread_count": unread_count,
        },
    )


@require_POST
@group_required("Admin", "Receptionist", module="notifications")
def notification_mark_read(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    if notification.read_at is None:
        notification.read_at = timezone.now()
        notification.save(update_fields=["read_at"])
        messages.success(request, "Notification marked read.")
    return redirect("notifications-center")


@group_required("Admin")
def settings_center(request):
    return render(
        request,
        "accounts/settings_center.html",
        {
            "room_categories": [
                {"label": "Deluxe", "count": Room.objects.filter(room_type=Room.RoomType.DELUXE).count()},
                {"label": "Standard", "count": Room.objects.filter(room_type=Room.RoomType.STANDARD).count()},
            ],
            "payment_methods": ["Cash", "Mobile Money", "Card", "Bank Transfer"],
        },
    )


@group_required("Admin", "Super Administrator", module="users_roles", action={"GET": "view", "POST": "manage"})
def users_roles_center(request):
    _seed_default_roles()

    users = User.objects.select_related("staff_profile", "access_profile").prefetch_related("groups", "groups__role_permissions").order_by("username")
    roles = Group.objects.prefetch_related("role_permissions").order_by("name")
    create_form = StaffUserForm()
    role_create_form = RoleCreateForm()
    user_forms = {user.pk: StaffRoleForm(user=user, initial={"user_id": user.pk}) for user in users}
    role_permission_forms = {role.pk: RolePermissionForm(role=role, initial={"role_id": role.pk}) for role in roles}

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_user":
            create_form = StaffUserForm(request.POST, request.FILES)
            if create_form.is_valid():
                user = create_form.save()
                log_audit_event(
                    request=request,
                    user=request.user,
                    action=AuditLog.ActionType.CREATE,
                    module="users_roles",
                    object_repr=user.username,
                    object_id=user.pk,
                    details={
                        "event": "user_created",
                        "roles": list(user.groups.values_list("name", flat=True)),
                    },
                )
                messages.success(request, f"User '{user.username}' created successfully.")
                return redirect("users-roles-center")
        elif action == "update_user":
            user = get_object_or_404(User, pk=request.POST.get("user_id"))
            role_form = StaffRoleForm(request.POST, request.FILES, user=user)
            user_forms[user.pk] = role_form
            if role_form.is_valid():
                updated_user = role_form.save()
                log_audit_event(
                    request=request,
                    user=request.user,
                    action=AuditLog.ActionType.UPDATE,
                    module="users_roles",
                    object_repr=updated_user.username,
                    object_id=updated_user.pk,
                    details={
                        "event": "user_updated",
                        "roles": list(updated_user.groups.values_list("name", flat=True)),
                        "is_active": updated_user.is_active,
                        "is_staff": updated_user.is_staff,
                    },
                )
                messages.success(request, f"Roles updated for '{user.username}'.")
                return redirect("users-roles-center")
        elif action == "delete_user":
            user = get_object_or_404(User, pk=request.POST.get("user_id"))
            if user.pk == request.user.pk:
                messages.error(request, "You cannot delete your own account while logged in.")
            else:
                username = user.username
                user_id = user.pk
                try:
                    with transaction.atomic():
                        user.delete()
                except ProtectedError:
                    messages.error(request, f"User '{username}' could not be deleted because the account is linked to other records.")
                except IntegrityError:
                    messages.error(request, f"User '{username}' could not be deleted because the database rejected the operation.")
                else:
                    try:
                        log_audit_event(
                            request=request,
                            user=request.user,
                            action=AuditLog.ActionType.DELETE,
                            module="users_roles",
                            object_repr=username,
                            object_id=user_id,
                            details={"event": "user_deleted"},
                        )
                    except Exception:
                        pass
                    messages.success(request, f"User '{username}' deleted successfully.")
                    return redirect("users-roles-center")
        elif action == "create_role":
            role_create_form = RoleCreateForm(request.POST)
            if role_create_form.is_valid():
                role = role_create_form.save()
                log_audit_event(
                    request=request,
                    user=request.user,
                    action=AuditLog.ActionType.CREATE,
                    module="users_roles",
                    object_repr=role.name,
                    object_id=role.pk,
                    details={"event": "role_created"},
                )
                messages.success(request, f"Role '{role.name}' created successfully.")
                return redirect("users-roles-center")
        elif action == "save_role_permissions":
            role = get_object_or_404(Group, pk=request.POST.get("role_id"))
            role_form = RolePermissionForm(request.POST, role=role)
            role_permission_forms[role.pk] = role_form
            if role_form.is_valid():
                role = role_form.save()
                log_audit_event(
                    request=request,
                    user=request.user,
                    action=AuditLog.ActionType.PERMISSION_CHANGE,
                    module="users_roles",
                    object_repr=role.name,
                    object_id=role.pk,
                    details={"event": "role_permissions_updated"},
                )
                messages.success(request, f"Permissions updated for role '{role.name}'.")
                return redirect("users-roles-center")

    recent_audit_logs = AuditLog.objects.select_related("user").order_by("-created_at")[:15]
    return render(
        request,
        "accounts/users_roles_center.html",
        {
            "create_form": create_form,
            "role_create_form": role_create_form,
            "users": users,
            "role_forms": user_forms,
            "role_permission_forms": role_permission_forms,
            "available_roles": roles,
            "recent_audit_logs": recent_audit_logs,
            "permission_modules": ACCESS_MODULE_CHOICES,
            "permission_actions": ACTION_CHOICES,
        },
    )


@group_required("Admin", "Receptionist", module="analytics")
def analytics_center(request):
    filters = _analytics_filters_from_request(request)
    sections = _build_analytics_sections(filters)
    charts = [chart for section in sections for chart in section["charts"]]
    return render(
        request,
        "accounts/analytics_center.html",
        {
            "report_window": filters["report_window"],
            "start_date": filters["start_date"].isoformat(),
            "end_date": filters["end_date"].isoformat(),
            "selected_period": filters["period"],
            "selected_department": filters["department"],
            "selected_room_type": filters["room_type"],
            "selected_staff_role": filters["staff_role"],
            "department_options": _analytics_department_options(),
            "room_type_options": list(Room.RoomType.choices),
            "staff_role_options": _system_role_filter_options(),
            "summary_metrics": _build_analytics_summary_metrics(sections),
            "sections": sections,
            "charts_json": json.dumps(charts),
            "export_xlsx_url": f"{reverse('analytics-export', args=['xlsx'])}?{filters['query_string']}",
            "export_pdf_url": f"{reverse('analytics-export', args=['pdf'])}?{filters['query_string']}",
        },
    )


@group_required("Admin", "Receptionist", module="analytics", action="export")
def analytics_export(request, fmt):
    filters = _analytics_filters_from_request(request)
    sections = _build_analytics_sections(filters)
    filename_base = f"analytics-{filters['report_window']['filename_range']}"

    if fmt == "xlsx":
        workbook = _create_reports_workbook_or_none(request)
        if workbook is None:
            return redirect("analytics-center")
        overview_sheet = workbook.active
        overview_sheet.title = "Overview"
        _write_report_overview_sheet(overview_sheet, sections, filters["report_window"]["display_range"])
        for section in sections:
            sheet = workbook.create_sheet(title=section["sheet_title"])
            _write_report_section_sheet(sheet, section, filters["report_window"]["display_range"])
        return _xlsx_response(workbook, f"{filename_base}.xlsx")

    if fmt == "pdf":
        return _analytics_pdf_response(sections, filters, f"{filename_base}.pdf")

    messages.error(request, "Unsupported analytics export format.")
    return redirect("analytics-center")


STAFF_FILTER_QUERY_KEYS = (
    "staff_view",
    "q",
    "department",
    "status",
    "role",
    "certifications",
    "roster",
)

LEAVE_EMPLOYMENT_STATUSES = tuple(dict(LEAVE_TYPE_TO_EMPLOYMENT_STATUS).values())
EMPLOYMENT_STATUS_TO_LEAVE_TYPE = {
    employment_status: leave_type
    for leave_type, employment_status in LEAVE_TYPE_TO_EMPLOYMENT_STATUS.items()
}


def _build_staff_filters_token(params):
    query = QueryDict("", mutable=True)
    staff_view = (params.get("staff_view", "").strip() or "active")
    if staff_view not in {"active", "terminated"}:
        staff_view = "active"

    for key in STAFF_FILTER_QUERY_KEYS:
        value = params.get(key, "").strip()
        if value and (key != "staff_view" or value != "active"):
            query[key] = value

    encoded = query.urlencode()
    return quote_plus(encoded) if encoded else ""


def _decode_staff_filters_token(token):
    if not token:
        return "", QueryDict("", mutable=False)
    query_string = unquote_plus(token).strip()
    return query_string, QueryDict(query_string, mutable=False)


def _append_query_param(url, name, value):
    if not value:
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}{name}={value}"


def _current_leave_q():
    today = timezone.localdate()
    return Q(
        leave_requests__approval_status=LeaveRequest.ApprovalStatus.APPROVED,
        leave_requests__start_date__lte=today,
        leave_requests__end_date__gte=today,
    )


def _system_status_filter_options():
    status_labels = dict(Employee.EMPLOYMENT_STATUS_CHOICES)
    options = [("active", "Active"), ("terminated", "Terminated")]
    options.extend(
        (employment_status, dict(LeaveRequest.LeaveType.choices).get(leave_type, status_labels.get(employment_status)))
        for leave_type, employment_status in LEAVE_TYPE_TO_EMPLOYMENT_STATUS.items()
    )

    existing_values = {value for value, _label in options}
    available_statuses = sorted(
        set(
            Employee.objects.exclude(employment_status="")
            .values_list("employment_status", flat=True)
            .distinct()
        )
    )
    for value in available_statuses:
        if value in existing_values:
            continue
        options.append((value, status_labels.get(value, value.replace("_", " ").title())))
    return options


def _system_role_filter_options():
    position_labels = dict(Employee.POSITION_CHOICES)
    positions = sorted(
        set(
            Employee.objects.exclude(position="")
            .values_list("position", flat=True)
            .distinct()
        )
    )
    return [
        (value, position_labels.get(value, value.replace("_", " ").title()))
        for value in positions
    ]


def _system_certification_filter_options():
    options = []
    qualification_names = sorted(
        set(
            EmployeeQualification.objects.exclude(qualification_name="")
            .values_list("qualification_name", flat=True)
            .distinct()
        )
    )
    if EmployeeQualification.objects.exists():
        options.extend(
            [
                ("expiring", "Expiring Soon"),
                ("expired", "Expired"),
            ]
        )
    options.extend((name, name) for name in qualification_names)
    return options


def _system_roster_filter_options():
    rotas = (
        Rota.objects.select_related("employee")
        .prefetch_related("staff_members")
        .filter(Q(employee__isnull=False) | Q(staff_members__isnull=False))
        .order_by("-period_start", "-created_at")
        .distinct()
    )
    options = []
    for rota in rotas:
        assigned_names = []
        if rota.employee_id:
            assigned_names.append(rota.employee.full_name)
        for member in rota.staff_members.all():
            if member.full_name not in assigned_names:
                assigned_names.append(member.full_name)
        label_parts = [
            rota.period or f"{rota.period_start or '-'} to {rota.period_end or '-'}",
        ]
        if rota.operating_hours != "-":
            label_parts.append(rota.operating_hours)
        if assigned_names:
            label_parts.append(", ".join(assigned_names))
        options.append((f"rota:{rota.pk}", " · ".join(label_parts)))
    return options


def _employee_section_filters(section, params):
    filters = []

    def add_select(name, label, choices):
        filters.append(
            {
                "name": name,
                "label": label,
                "type": "select",
                "choices": [("", "All")] + list(choices),
                "value": params.get(name, "").strip(),
            }
        )

    def add_date(name, label):
        filters.append(
            {
                "name": name,
                "label": label,
                "type": "date",
                "value": params.get(name, "").strip(),
            }
        )

    def add_text(name, label, placeholder):
        filters.append(
            {
                "name": name,
                "label": label,
                "type": "text",
                "placeholder": placeholder,
                "value": params.get(name, "").strip(),
            }
        )

    if section == "leave":
        add_select("approval_status", "Status", LeaveRequest.ApprovalStatus.choices)
        add_select("leave_type", "Leave Type", LeaveRequest.LeaveType.choices)
        add_date("date_from", "Start From")
        add_date("date_to", "End To")
    elif section == "attendance":
        add_select("attendance_status", "Status", AttendanceRecord.AttendanceStatus.choices)
        add_date("date_from", "Date From")
        add_date("date_to", "Date To")
    elif section == "certifications":
        add_select(
            "certification_status",
            "Status",
            [
                ("current", "Current"),
                ("expiring", "Expiring Soon"),
                ("expired", "Expired"),
            ],
        )
        add_text("query", "Search", "Qualification or institution")
        add_date("date_from", "Issued From")
        add_date("date_to", "Issued To")
    elif section == "documents":
        add_select("document_type", "Type", EmployeeDocument.DocumentType.choices)
        add_text("query", "Search", "Title or description")
    elif section == "payroll":
        add_select("payment_status", "Status", PayrollRecord.PaymentStatus.choices)
        add_date("date_from", "Period From")
        add_date("date_to", "Period To")
    elif section == "performance":
        add_select(
            "rating",
            "Rating",
            [(str(value), str(value)) for value in range(1, 6)],
        )
        add_date("date_from", "Review From")
        add_date("date_to", "Review To")
    elif section == "disciplinary":
        add_select("record_type", "Type", DisciplinaryRecord.RecordType.choices)
        add_select(
            "resolved",
            "Resolved",
            [("yes", "Yes"), ("no", "No")],
        )
        add_date("date_from", "Incident From")
        add_date("date_to", "Incident To")
    elif section == "training":
        add_text("query", "Search", "Training or provider")
        add_date("date_from", "Completed From")
        add_date("date_to", "Completed To")
    elif section == "history":
        add_select("change_type", "Change Type", EmploymentHistoryEntry.ChangeType.choices)
        add_date("date_from", "Effective From")
        add_date("date_to", "Effective To")

    return filters


def _apply_employee_section_filters(records, section, params):
    if section == "leave":
        approval_status = params.get("approval_status", "").strip()
        leave_type = params.get("leave_type", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if approval_status:
            records = records.filter(approval_status=approval_status)
        if leave_type:
            records = records.filter(leave_type=leave_type)
        if date_from:
            records = records.filter(start_date__gte=date_from)
        if date_to:
            records = records.filter(end_date__lte=date_to)
        return records

    if section == "attendance":
        attendance_status = params.get("attendance_status", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if attendance_status:
            records = records.filter(status=attendance_status)
        if date_from:
            records = records.filter(work_date__gte=date_from)
        if date_to:
            records = records.filter(work_date__lte=date_to)
        return records

    if section == "certifications":
        certification_status = params.get("certification_status", "").strip()
        query = params.get("query", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        today = timezone.localdate()
        if certification_status == "current":
            records = records.filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=today))
        elif certification_status == "expiring":
            records = records.filter(
                expiry_date__isnull=False,
                expiry_date__gte=today,
                expiry_date__lte=today + timedelta(days=30),
            )
        elif certification_status == "expired":
            records = records.filter(expiry_date__isnull=False, expiry_date__lt=today)
        if query:
            records = records.filter(
                Q(qualification_name__icontains=query)
                | Q(institution__icontains=query)
                | Q(certificate_number__icontains=query)
            )
        if date_from:
            records = records.filter(certification_date__gte=date_from)
        if date_to:
            records = records.filter(certification_date__lte=date_to)
        return records

    if section == "documents":
        document_type = params.get("document_type", "").strip()
        query = params.get("query", "").strip()
        if document_type:
            records = records.filter(document_type=document_type)
        if query:
            records = records.filter(
                Q(title__icontains=query) | Q(description__icontains=query)
            )
        return records

    if section == "payroll":
        payment_status = params.get("payment_status", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if payment_status:
            records = records.filter(payment_status=payment_status)
        if date_from:
            records = records.filter(pay_period_start__gte=date_from)
        if date_to:
            records = records.filter(pay_period_end__lte=date_to)
        return records

    if section == "performance":
        rating = params.get("rating", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if rating:
            records = records.filter(rating=rating)
        if date_from:
            records = records.filter(review_date__gte=date_from)
        if date_to:
            records = records.filter(review_date__lte=date_to)
        return records

    if section == "disciplinary":
        record_type = params.get("record_type", "").strip()
        resolved = params.get("resolved", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if record_type:
            records = records.filter(record_type=record_type)
        if resolved == "yes":
            records = records.filter(resolved=True)
        elif resolved == "no":
            records = records.filter(resolved=False)
        if date_from:
            records = records.filter(incident_date__gte=date_from)
        if date_to:
            records = records.filter(incident_date__lte=date_to)
        return records

    if section == "training":
        query = params.get("query", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if query:
            records = records.filter(
                Q(training_name__icontains=query) | Q(provider__icontains=query)
            )
        if date_from:
            records = records.filter(completion_date__gte=date_from)
        if date_to:
            records = records.filter(completion_date__lte=date_to)
        return records

    if section == "history":
        change_type = params.get("change_type", "").strip()
        date_from = params.get("date_from", "").strip()
        date_to = params.get("date_to", "").strip()
        if change_type:
            records = records.filter(change_type=change_type)
        if date_from:
            records = records.filter(effective_date__gte=date_from)
        if date_to:
            records = records.filter(effective_date__lte=date_to)
        return records

    return records


def _employee_section_links(employee, staff_filters_token="", active_section=None):
    section_specs = [
        ("leave", "Annual Leave", employee.leave_requests.count()),
        ("certifications", "Certifications", employee.qualifications.count()),
        ("documents", "Documents", employee.documents.count()),
        ("attendance", "Attendance History", employee.attendance_records.count()),
        ("payroll", "Payroll Information", employee.payroll_records.count()),
        ("performance", "Performance Reviews", employee.performance_reviews.count()),
        ("disciplinary", "Disciplinary Records", employee.disciplinary_records.count()),
        ("training", "Training Records", employee.training_records.count()),
        ("history", "Employment History", employee.history_entries.count()),
    ]
    links = []
    for key, label, count in section_specs:
        url = reverse("hr-employee-section", args=[employee.pk, key])
        url = _append_query_param(url, "staff_filters", staff_filters_token)
        links.append(
            {
                "key": key,
                "label": label,
                "count": count,
                "url": url,
                "active": key == active_section,
            }
        )

    roster_url = reverse("hr-rota-list")
    roster_url = _append_query_param(roster_url, "employee", employee.pk)
    links.append(
        {
            "key": "roster",
            "label": "Roster Assignments",
            "count": employee.rota_entries.count(),
            "url": roster_url,
            "active": active_section == "roster",
        }
    )
    return links


def _employee_section_headers(section):
    return {
        "leave": ["Type", "Dates", "Days", "Return", "Status", "Manager"],
        "certifications": ["Qualification", "Institution", "Certificate #", "Issued", "Expires"],
        "documents": ["Type", "Title", "File", "Created"],
        "attendance": ["Date", "Shift", "Status", "Check-in", "Check-out"],
        "payroll": ["Pay period", "Gross", "Net", "Status", "Paid at"],
        "performance": ["Review date", "Reviewer", "Rating", "Next review", "Summary"],
        "disciplinary": ["Date", "Type", "Resolved", "Action taken", "Notes"],
        "training": ["Training", "Provider", "Completed", "Expires", "Notes"],
        "history": ["Change", "Effective date", "Description", "Created by"],
    }.get(section, [])


def _employee_section_records(employee, section):
    if section == "leave":
        return employee.leave_requests.select_related("approving_manager").order_by("-created_at")
    if section == "certifications":
        return employee.qualifications.order_by("-certification_date")
    if section == "documents":
        return employee.documents.order_by("-created_at")
    if section == "attendance":
        return employee.attendance_records.order_by("-work_date")
    if section == "payroll":
        return employee.payroll_records.order_by("-pay_period_start")
    if section == "performance":
        return employee.performance_reviews.select_related("reviewer").order_by("-review_date")
    if section == "disciplinary":
        return employee.disciplinary_records.order_by("-incident_date")
    if section == "training":
        return employee.training_records.order_by("-completion_date")
    if section == "history":
        return employee.history_entries.select_related("created_by").order_by("-effective_date")
    return []


def _employee_section_form_class(section):
    return {
        "leave": LeaveRequestForm,
        "certifications": EmployeeQualificationForm,
        "documents": EmployeeDocumentForm,
        "attendance": AttendanceRecordForm,
        "payroll": PayrollRecordForm,
        "performance": PerformanceReviewForm,
        "disciplinary": DisciplinaryRecordForm,
        "training": TrainingRecordForm,
        "history": EmploymentHistoryForm,
    }.get(section)


def _employee_section_title(section):
    return {
        "leave": "Annual Leave",
        "certifications": "Qualifications & Certifications",
        "documents": "Employee Documents",
        "attendance": "Attendance History",
        "payroll": "Payroll Information",
        "performance": "Performance Reviews",
        "disciplinary": "Disciplinary Records",
        "training": "Training Records",
        "history": "Employment History",
    }.get(section, "Staff Management")


def _employee_section_row(record, section):
    if section == "leave":
        return [
            record.get_leave_type_display(),
            f"{record.start_date:%Y-%m-%d} to {record.end_date:%Y-%m-%d}",
            str(record.days),
            record.return_to_work_date.strftime("%d/%m/%Y") if record.return_to_work_date else "-",
            record.get_approval_status_display(),
            record.approving_manager.full_name if record.approving_manager else "-",
        ]
    if section == "certifications":
        return [
            record.qualification_name,
            record.institution,
            record.certificate_number or "-",
            record.certification_date.strftime("%d/%m/%Y"),
            record.expiry_date.strftime("%d/%m/%Y") if record.expiry_date else "-",
        ]
    if section == "documents":
        return [
            record.get_document_type_display(),
            record.title,
            record.file.name.split("/")[-1],
            record.created_at.strftime("%d/%m/%Y"),
        ]
    if section == "attendance":
        return [
            record.work_date.strftime("%d/%m/%Y"),
            record.get_shift_type_display(),
            record.get_status_display(),
            record.check_in.strftime("%d/%m/%Y %H:%M") if record.check_in else "-",
            record.check_out.strftime("%d/%m/%Y %H:%M") if record.check_out else "-",
        ]
    if section == "payroll":
        return [
            f"{record.pay_period_start:%Y-%m-%d} to {record.pay_period_end:%Y-%m-%d}",
            f"GHS {record.basic_salary}",
            f"GHS {record.net_pay}",
            record.get_payment_status_display(),
            record.paid_at.strftime("%d/%m/%Y %H:%M") if record.paid_at else "-",
        ]
    if section == "performance":
        return [
            record.review_date.strftime("%d/%m/%Y"),
            record.reviewer.get_username() if record.reviewer else "-",
            str(record.rating),
            record.next_review_date.strftime("%d/%m/%Y") if record.next_review_date else "-",
            record.summary[:80] if record.summary else "-",
        ]
    if section == "disciplinary":
        return [
            record.incident_date.strftime("%d/%m/%Y"),
            record.get_record_type_display(),
            "Yes" if record.resolved else "No",
            record.action_taken[:80] if record.action_taken else "-",
            record.notes[:80] if record.notes else "-",
        ]
    if section == "training":
        return [
            record.training_name,
            record.provider or "-",
            record.completion_date.strftime("%d/%m/%Y") if record.completion_date else "-",
            record.expiry_date.strftime("%d/%m/%Y") if record.expiry_date else "-",
            record.notes[:80] if record.notes else "-",
        ]
    if section == "history":
        return [
            record.get_change_type_display(),
            record.effective_date.strftime("%d/%m/%Y"),
            record.description[:80],
            record.created_by.get_username() if record.created_by else "-",
        ]
    return []


def _staff_management_queryset(request):
    query = request.GET.get("q", "").strip()
    department = request.GET.get("department", "").strip()
    employment_status = request.GET.get("status", "").strip()
    role = request.GET.get("role", "").strip()
    certification_status = request.GET.get("certifications", "").strip()
    roster_employee = request.GET.get("roster", "").strip()
    staff_view = request.GET.get("staff_view", "active").strip() or "active"
    if staff_view not in {"active", "terminated"}:
        staff_view = "active"

    employees = (
        Employee.objects.select_related("supervisor")
        .prefetch_related("qualifications", "leave_requests", "rota_entries", "rotas")
        .order_by("last_name", "first_name")
    )

    show_terminated_only = (
        staff_view == "terminated"
        or employment_status == "terminated"
    )

    if show_terminated_only:
        employees = employees.filter(employment_status="terminated")
    else:
        employees = employees.exclude(employment_status="terminated")

    if query:
        employees = employees.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(employee_id__icontains=query)
            | Q(position__icontains=query)
            | Q(job_title__icontains=query)
            | Q(department__icontains=query)
            | Q(ghana_card_number__icontains=query)
        )
    if department:
        employees = employees.filter(department=department)
    if employment_status:
        if employment_status == "active":
            employees = employees.filter(employment_status="active").exclude(_current_leave_q())
        elif employment_status in EMPLOYMENT_STATUS_TO_LEAVE_TYPE:
            leave_type = EMPLOYMENT_STATUS_TO_LEAVE_TYPE[employment_status]
            employees = employees.filter(
                Q(employment_status=employment_status)
                | Q(
                    leave_requests__leave_type=leave_type,
                    leave_requests__approval_status=LeaveRequest.ApprovalStatus.APPROVED,
                    leave_requests__start_date__lte=timezone.localdate(),
                    leave_requests__end_date__gte=timezone.localdate(),
                )
            )
        else:
            employees = employees.filter(employment_status=employment_status)
    if role:
        employees = employees.filter(position=role)
    if roster_employee:
        if roster_employee.startswith("rota:"):
            rota_id = roster_employee.split(":", 1)[1]
            employees = employees.filter(Q(rota_entries__pk=rota_id) | Q(rotas__pk=rota_id))
        else:
            employees = employees.filter(rota_entries__isnull=False, pk=roster_employee)
    if certification_status == "expiring":
        today = timezone.localdate()
        employees = employees.filter(
            qualifications__expiry_date__isnull=False,
            qualifications__expiry_date__gte=today,
            qualifications__expiry_date__lte=today + timedelta(days=30),
        )
    elif certification_status == "expired":
        employees = employees.filter(
            qualifications__expiry_date__isnull=False,
            qualifications__expiry_date__lt=timezone.localdate(),
        )
    elif certification_status:
        employees = employees.filter(
            qualifications__qualification_name=certification_status
        )

    employees = employees.distinct()

    return {
        "employees": employees,
        "query": query,
        "department": department,
        "employment_status": employment_status,
        "role": role,
        "certification_status": certification_status,
        "selected_roster_employee": roster_employee,
        "staff_view": staff_view,
    }


@group_required(
    "Admin",
    "Super Administrator",
    module="staff_management",
    action={"GET": "view", "POST": "edit"},
)
def hr_employee_section(request, pk, section):
    employee = get_object_or_404(Employee, pk=pk)
    form_class = _employee_section_form_class(section)
    if form_class is None:
        return redirect("hr-detail", pk=employee.pk)

    staff_filters_token = request.GET.get("staff_filters", "").strip()
    back_to_list_url = reverse("hr-list")
    if staff_filters_token:
        staff_filters_query, _ = _decode_staff_filters_token(staff_filters_token)
        if staff_filters_query:
            back_to_list_url = f"{back_to_list_url}?{staff_filters_query}"

    form = form_class(request.POST or None, request.FILES or None)
    records = _apply_employee_section_filters(
        _employee_section_records(employee, section),
        section,
        request.GET,
    )
    if request.method == "POST" and form.is_valid():
        record = form.save(commit=False)
        record.employee = employee
        if hasattr(record, "approving_manager_id") and not record.approving_manager_id and section == "leave":
            if record.approval_status == LeaveRequest.ApprovalStatus.APPROVED and employee.supervisor_id:
                record.approving_manager = employee.supervisor
        if hasattr(record, "approved_at") and section == "leave":
            if record.approval_status == LeaveRequest.ApprovalStatus.APPROVED and not record.approved_at:
                record.approved_at = timezone.now()
            elif record.approval_status != LeaveRequest.ApprovalStatus.APPROVED:
                record.approved_at = None
        if hasattr(record, "reviewer_id") and not record.reviewer_id and section == "performance":
            record.reviewer = request.user
        if hasattr(record, "created_by_id") and not record.created_by_id and section == "history":
            record.created_by = request.user
        if hasattr(record, "paid_at") and record.payment_status == PayrollRecord.PaymentStatus.PAID and not record.paid_at:
            record.paid_at = timezone.now()
        record.save()
        log_audit_event(
            request=request,
            user=request.user,
            action=AuditLog.ActionType.CREATE,
            module="staff_management",
            object_repr=str(employee),
            object_id=employee.pk,
            details={"section": section, "record": str(record)},
        )
        messages.success(request, f"{_employee_section_title(section)} saved successfully.")
        redirect_url = reverse("hr-employee-section", args=[employee.pk, section])
        if staff_filters_token:
            redirect_url = _append_query_param(
                redirect_url,
                "staff_filters",
                staff_filters_token,
            )
        return redirect(redirect_url)

    return render(
        request,
        "accounts/hr_employee_section.html",
        {
            "employee": employee,
            "section": section,
            "section_title": _employee_section_title(section),
            "headers": _employee_section_headers(section),
            "rows": [_employee_section_row(record, section) for record in records],
            "records": records,
            "record_count": records.count(),
            "form": form,
            "section_filters": _employee_section_filters(section, request.GET),
            "staff_filters_token": staff_filters_token,
            "back_to_list_url": back_to_list_url,
            "section_links": _employee_section_links(
                employee,
                staff_filters_token=staff_filters_token,
                active_section=section,
            ),
        },
    )


@group_required("Admin", "Super Administrator", module="staff_management")
def hr_employee_list(request):
    context = _staff_management_queryset(request)
    context.update(
        {
            "departments": sorted(
                set(
                    Employee.objects.exclude(department="")
                    .values_list("department", flat=True)
                    .distinct()
                )
            ),
            "role_options": _system_role_filter_options(),
            "status_choices": _system_status_filter_options(),
            "certification_filter_options": _system_certification_filter_options(),
            "roster_options": _system_roster_filter_options(),
            "active_count": Employee.objects.exclude(employment_status="terminated").count(),
            "terminated_count": Employee.objects.filter(employment_status="terminated").count(),
            "staff_filters_token": _build_staff_filters_token(request.GET),
        }
    )
    return render(request, "accounts/hr_employee_list.html", context)


@group_required("Admin", "Super Administrator", module="staff_management", action="create")
def hr_employee_create(request):
    form = EmployeeForm(request.POST or None, request.FILES or None, for_create=True)
    if form.is_valid():
        form.save()
        messages.success(request, "Employee saved successfully.")
        return redirect("hr-list")
    return render(
        request,
        "accounts/hr_employee_form.html",
        {"form": form, "form_title": "Add New Employee", "is_create": True},
    )


@group_required("Admin", "Super Administrator", module="staff_management", action="edit")
def hr_employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    form = EmployeeForm(request.POST or None, request.FILES or None, instance=employee)
    if form.is_valid():
        form.save()
        messages.success(request, "Employee updated successfully.")
        return redirect("hr-list")
    return render(
        request,
        "accounts/hr_employee_form.html",
        {"form": form, "form_title": "Update Employee", "is_create": False},
    )


@group_required("Admin", "Super Administrator", module="staff_management")
def hr_employee_detail(request, pk):
    staff_filters_token = request.GET.get("staff_filters", "").strip()
    back_to_list_url = reverse("hr-list")
    if staff_filters_token:
        staff_filters_query, _ = _decode_staff_filters_token(staff_filters_token)
        if staff_filters_query:
            back_to_list_url = f"{back_to_list_url}?{staff_filters_query}"

    employee = get_object_or_404(
        Employee.objects.select_related("supervisor", "termination_approved_by")
        .prefetch_related("rota_entries", "documents", "qualifications", "leave_requests"),
        pk=pk,
    )
    rotas = employee.rota_entries.order_by("-period_start", "opening_time")
    return render(
        request,
        "accounts/hr_employee_detail.html",
        {
            "employee": employee,
            "rotas": rotas,
            "leave_balance": employee.annual_leave_balance,
            "expiring_certifications_count": employee.expiring_certifications_count,
            "staff_filters_token": staff_filters_token,
            "back_to_list_url": back_to_list_url,
            "section_links": _employee_section_links(
                employee,
                staff_filters_token=staff_filters_token,
            ),
        },
    )


@group_required("Admin", "Super Administrator", module="staff_management", action="delete")
def hr_employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        employee.delete()
        return redirect("hr-list")
    return render(request, "accounts/hr_employee_confirm_delete.html", {"employee": employee})


@group_required("Admin", "Super Administrator", module="staff_management")
def hr_rota_list(request):
    period, reference_date, start_date, end_date = _parse_rota_range(request)
    query = request.GET.get("q", "").strip()
    employee_id = request.GET.get("employee", "").strip()
    rotas_qs = (
        Rota.objects.select_related("employee")
        .filter(employee__isnull=False, period_start__lte=end_date, period_end__gte=start_date)
        .order_by("-period_start", "employee__last_name", "employee__first_name")
    )
    if query:
        rotas_qs = rotas_qs.filter(
            Q(employee__first_name__icontains=query)
            | Q(employee__last_name__icontains=query)
            | Q(employee__position__icontains=query)
            | Q(period__icontains=query)
        )
    if employee_id:
        rotas_qs = rotas_qs.filter(employee_id=employee_id)
    rotas = list(rotas_qs)

    summary = {
        "total_rotas": len(rotas),
        "total_hours": sum((rota.total_hours for rota in rotas), 0),
        "active_employees": len({rota.employee_id for rota in rotas}),
    }
    return render(
        request,
        "accounts/hr_rota_list.html",
        {
            "rotas": rotas,
            "query": query,
            "period": period,
            "reference_date": reference_date,
            "start_date": start_date,
            "end_date": end_date,
            "summary": summary,
            "day_reports": _rota_day_reports(rotas, start_date, end_date),
            "selected_employee": employee_id,
            "employees": Employee.objects.order_by("last_name", "first_name"),
        },
    )


@group_required("Admin", "Super Administrator", module="staff_management", action="create")
def hr_rota_create(request):
    form = RotaForm(request.POST or None)
    if form.is_valid():
        rota = form.save(commit=False)
        if not rota.period:
            rota.period = f"{rota.employee} weekly duty roster"
        rota.save()
        rota.staff_members.set([rota.employee])
        messages.success(request, "Duty roster created successfully.")
        return redirect("hr-rota-list")
    return render(
        request,
        "accounts/hr_rota_form.html",
        {"form": form, "form_title": "Create Weekly Duty Roster"},
    )


@group_required("Admin", "Super Administrator", module="staff_management", action="edit")
def hr_rota_update(request, pk):
    rota = get_object_or_404(Rota, pk=pk)
    form = RotaForm(request.POST or None, instance=rota)
    if form.is_valid():
        rota = form.save()
        if rota.employee_id:
            rota.staff_members.set([rota.employee])
        messages.success(request, "Duty roster updated successfully.")
        return redirect("hr-rota-list")
    return render(
        request,
        "accounts/hr_rota_form.html",
        {"form": form, "form_title": "Edit Weekly Duty Roster"},
    )


@group_required("Admin", "Super Administrator", module="staff_management")
def hr_rota_detail(request, pk):
    rota = get_object_or_404(Rota.objects.select_related("employee"), pk=pk)
    return render(
        request,
        "accounts/hr_rota_detail.html",
        {
            "rota": rota,
            "daily_roster": _rota_daily_breakdown(rota),
        },
    )


@group_required("Admin")
def admin_reports(request):
    report_window = _report_window_from_request(request)
    sections = _build_admin_report_sections(
        report_window["start_date"],
        report_window["end_date"],
    )
    occupancy_rows = _daily_report_rows(
        report_window["start_date"],
        report_window["end_date"],
        Room.objects.count(),
    )
    revenue_rows = _daily_total_revenue_rows(
        report_window["start_date"],
        report_window["end_date"],
    )

    context = {
        "today": timezone.localdate(),
        "report_window": report_window,
        "report_period": report_window["period"],
        "report_label": report_window["label"],
        "display_range": report_window["display_range"],
        "start_date": report_window["start_date"].isoformat(),
        "end_date": report_window["end_date"].isoformat(),
        "sections": sections,
        "summary_metrics": _build_admin_report_summary_metrics(sections),
        "chart_labels_json": json.dumps([row["date"] for row in occupancy_rows]),
        "revenue_data_json": json.dumps([float(row["revenue_total"]) for row in revenue_rows]),
        "occupancy_data_json": json.dumps([row["occupied_rooms"] for row in occupancy_rows]),
    }
    return render(request, "accounts/admin_reports.html", context)


@group_required("Admin")
def admin_reports_export_daily_csv(request):
    start_date, end_date = _parse_report_range(request)
    total_rooms = Room.objects.count()
    daily_rows = _daily_report_rows(start_date, end_date, total_rooms)
    revenue_lookup = {
        date.fromisoformat(row["date"]): row["revenue_total"]
        for row in _daily_total_revenue_rows(start_date, end_date)
    }

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="daily-report-{start_date}-{end_date}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(["date", "occupied_rooms", "occupancy_percent", "revenue_collected"])
    for row in daily_rows:
        writer.writerow(
            [
                row["date"],
                row["occupied_rooms"],
                row["occupancy_percent"],
                revenue_lookup.get(date.fromisoformat(row["date"]), Decimal("0")),
            ]
        )
    return response


@group_required("Admin")
def admin_reports_export_balances_csv(request):
    bookings_with_balance = Booking.objects.annotate(
        paid_total=Coalesce(
            Sum("payments__amount"),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
        ),
        balance=ExpressionWrapper(
            F("total_amount") - F("paid_total"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
    ).filter(balance__gt=0).select_related("guest", "room")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="outstanding-balances.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "booking_id",
            "guest_name",
            "room_number",
            "status",
            "total_amount",
            "amount_paid",
            "balance_due",
        ]
    )
    for booking in bookings_with_balance:
        writer.writerow(
            [
                booking.id,
                f"{booking.guest.first_name} {booking.guest.last_name}",
                booking.room.room_number,
                booking.get_status_display(),
                booking.total_amount,
                booking.paid_total,
                booking.balance,
            ]
        )
    return response


def _parse_report_range(request):
    report_window = _report_window_from_request(request)
    return report_window["start_date"], report_window["end_date"]


@group_required("Admin")
def admin_reports_export_section_excel(request, section):
    workbook = _create_reports_workbook_or_none(request)
    if workbook is None:
        return redirect("admin-reports")

    report_window = _report_window_from_request(request)
    sections = _build_admin_report_sections(
        report_window["start_date"],
        report_window["end_date"],
    )
    section_map = {section_data["key"]: section_data for section_data in sections}
    section_data = section_map.get(section)
    if section_data is None:
        messages.error(request, "That report section could not be found.")
        return redirect("admin-reports")

    workbook.active.title = section_data["sheet_title"]
    _write_report_section_sheet(workbook.active, section_data, report_window["display_range"])
    return _xlsx_response(
        workbook,
        f"{section_data['filename_prefix']}-report-{report_window['filename_range']}.xlsx",
    )


@group_required("Admin")
def admin_reports_export_all_excel(request):
    workbook = _create_reports_workbook_or_none(request)
    if workbook is None:
        return redirect("admin-reports")

    report_window = _report_window_from_request(request)
    sections = _build_admin_report_sections(
        report_window["start_date"],
        report_window["end_date"],
    )

    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    _write_report_overview_sheet(overview_sheet, sections, report_window["display_range"])
    for section in sections:
        sheet = workbook.create_sheet(title=section["sheet_title"])
        _write_report_section_sheet(sheet, section, report_window["display_range"])

    return _xlsx_response(
        workbook,
        f"full-report-{report_window['filename_range']}.xlsx",
    )


REPORT_PRESET_LABELS = {
    "daily": "Daily",
    "weekly": "Weekly",
    "monthly": "Monthly",
    "yearly": "Yearly",
    "custom": "Custom",
}


def _report_window_from_request(request):
    today = timezone.localdate()
    period = (request.GET.get("period") or "weekly").strip().lower()
    if period not in {"daily", "weekly", "monthly", "yearly"}:
        period = "weekly"

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if start_date_str or end_date_str:
        start_date = today
        end_date = today
        try:
            if start_date_str:
                start_date = date.fromisoformat(start_date_str)
            if end_date_str:
                end_date = date.fromisoformat(end_date_str)
        except ValueError:
            start_date = today
            end_date = today
        period = "custom"
    else:
        start_date, end_date = _report_window_for_period(period, today)

    start_date, end_date = normalize_date_range(start_date, end_date)

    display_range = f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    label = f"{REPORT_PRESET_LABELS.get(period, 'Custom')} report for {display_range}"

    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "display_range": display_range,
        "label": label,
        "filename_range": f"{start_date.strftime('%d-%m-%Y')}-to-{end_date.strftime('%d-%m-%Y')}",
    }


def _report_window_for_period(period, today):
    return shared_report_window_for_period(period, today)


def _report_days(start_date, end_date):
    return [start_date + timedelta(days=index) for index in range((end_date - start_date).days + 1)]


def _display_date(value):
    return value.strftime("%d/%m/%Y") if value else "-"


def _display_money(value):
    return f"GHS {Decimal(str(value or 0)):,.2f}"


def _display_quantity(value):
    return format_quantity(value or 0) or "0"


def _display_percent(value):
    return f"{float(value or 0):.2f}%"


def _relative_time(value):
    if not value:
        return "-"
    local_value = timezone.localtime(value)
    now = timezone.localtime()
    if local_value.date() == now.date():
        delta = now - local_value
        if delta < timedelta(minutes=1):
            return "Just now"
        if delta < timedelta(hours=1):
            minutes = max(int(delta.total_seconds() // 60), 1)
            return f"{minutes} min{'s' if minutes != 1 else ''} ago"
        return f"Today {local_value.strftime('%I:%M%p').lstrip('0').lower()}"
    return f"{timesince(local_value, now)} ago"


def _money_total(queryset, field_name):
    return shared_money_total(queryset, field_name)


def _quantity_total(queryset, field_name):
    return queryset.aggregate(
        total=Coalesce(
            Sum(field_name),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=3)),
        )
    )["total"]


def _build_admin_report_sections(start_date, end_date):
    total_rooms = Room.objects.count()
    return [
        _build_bookings_report_section(start_date, end_date, total_rooms),
        _build_revenue_report_section(start_date, end_date),
        _build_housekeeping_report_section(start_date, end_date),
        _build_roster_report_section(start_date, end_date),
        _build_rooms_report_section(start_date, end_date, total_rooms),
        _build_staff_report_section(start_date, end_date),
    ]


def _build_admin_report_summary_metrics(sections):
    section_map = {section["key"]: section for section in sections}
    return [
        {"label": "Reservations", "value": section_map["bookings"]["summary"]["reservations"]},
        {
            "label": "Revenue",
            "value": _display_money(section_map["revenue-payments"]["summary"]["total_revenue"]),
        },
        {
            "label": "Items Used",
            "value": _display_quantity(section_map["housekeeping"]["summary"]["items_used"]),
        },
        {
            "label": "Avg Occupancy",
            "value": _display_percent(section_map["rooms"]["summary"]["average_occupancy"]),
        },
    ]


def _analytics_filters_from_request(request):
    report_window = _report_window_from_request(request)
    query = QueryDict("", mutable=True)
    query["period"] = report_window["period"]
    query["start_date"] = report_window["start_date"].isoformat()
    query["end_date"] = report_window["end_date"].isoformat()

    department = request.GET.get("department", "").strip()
    room_type = request.GET.get("room_type", "").strip()
    staff_role = request.GET.get("staff_role", "").strip()

    if department:
        query["department"] = department
    if room_type:
        query["room_type"] = room_type
    if staff_role:
        query["staff_role"] = staff_role

    return {
        "report_window": report_window,
        "period": report_window["period"],
        "start_date": report_window["start_date"],
        "end_date": report_window["end_date"],
        "department": department,
        "room_type": room_type,
        "staff_role": staff_role,
        "query_string": query.urlencode(),
    }


def _analytics_department_options():
    return sorted(
        set(
            Employee.objects.exclude(department="")
            .values_list("department", flat=True)
            .distinct()
        )
    )


def _analytics_room_floor(room_number):
    digits = "".join(char for char in str(room_number or "") if char.isdigit())
    if not digits:
        return "Unknown"
    if len(digits) >= 3:
        floor_number = int(digits[:-2] or 0)
    else:
        floor_number = int(digits[0] or 0)
    return "Ground Floor" if floor_number == 0 else f"Floor {floor_number}"


def _analytics_granularity(period, start_date, end_date):
    day_span = max((end_date - start_date).days + 1, 1)
    if period == "yearly" or day_span > 180:
        return "month"
    if period == "monthly" or day_span > 45:
        return "week"
    return "day"


def _analytics_bucket_key(day, granularity):
    if granularity == "month":
        return day.replace(day=1)
    if granularity == "week":
        return day - timedelta(days=day.weekday())
    return day


def _analytics_bucket_label(bucket_key, granularity):
    if granularity == "month":
        return bucket_key.strftime("%b %Y")
    if granularity == "week":
        return f"Week of {bucket_key.strftime('%d %b')}"
    return bucket_key.strftime("%d %b")


def _analytics_aggregate_points(points, period, start_date, end_date, sum_fields, avg_fields=()):
    granularity = _analytics_granularity(period, start_date, end_date)
    buckets = {}
    counts = defaultdict(int)
    field_names = list(sum_fields) + list(avg_fields)

    for point in points:
        bucket_key = _analytics_bucket_key(point["date"], granularity)
        if bucket_key not in buckets:
            buckets[bucket_key] = {field_name: Decimal("0") for field_name in field_names}
        counts[bucket_key] += 1
        for field_name in field_names:
            buckets[bucket_key][field_name] += Decimal(str(point.get(field_name, 0) or 0))

    labels = []
    values = {field_name: [] for field_name in field_names}
    for bucket_key in sorted(buckets):
        labels.append(_analytics_bucket_label(bucket_key, granularity))
        for field_name in sum_fields:
            values[field_name].append(float(buckets[bucket_key][field_name]))
        for field_name in avg_fields:
            divisor = counts[bucket_key] or 1
            values[field_name].append(float(buckets[bucket_key][field_name] / Decimal(divisor)))
    return {"labels": labels, "values": values, "granularity": granularity}


def _analytics_room_queryset(room_type=""):
    queryset = Room.objects.all().order_by("room_number")
    if room_type:
        queryset = queryset.filter(room_type=room_type)
    return queryset


def _analytics_employee_queryset(department="", staff_role=""):
    queryset = Employee.objects.all().order_by("last_name", "first_name")
    if department:
        queryset = queryset.filter(department=department)
    if staff_role:
        queryset = queryset.filter(position=staff_role)
    return queryset


def _analytics_booking_creation_queryset(start_date, end_date, room_type=""):
    queryset = Booking.objects.select_related("guest", "room").filter(
        created_at__date__range=[start_date, end_date]
    )
    if room_type:
        queryset = queryset.filter(room__room_type=room_type)
    return queryset


def _analytics_booking_stay_queryset(start_date, end_date, room_type=""):
    queryset = Booking.objects.select_related("guest", "room").filter(
        check_in__lte=end_date,
        check_out__gte=start_date,
    )
    if room_type:
        queryset = queryset.filter(room__room_type=room_type)
    return queryset


def _analytics_room_status_at(room, target_moment):
    history_entries = sorted(
        list(room.status_history.all()),
        key=lambda entry: entry.changed_at,
    )
    relevant_history = [entry for entry in history_entries if entry.changed_at <= target_moment]
    if relevant_history:
        return relevant_history[-1].new_status
    if history_entries and history_entries[0].changed_at > target_moment:
        return history_entries[0].previous_status or None
    if room.created_at and room.created_at <= target_moment:
        return room.status
    return None


def _analytics_room_status_points(start_date, end_date, room_type=""):
    room_type_ct = ContentType.objects.get_for_model(Room)
    rooms = list(
        _analytics_room_queryset(room_type).prefetch_related("status_history").filter(
            Q(created_at__date__lte=end_date)
            | Q(status_history__content_type=room_type_ct)
        ).distinct()
    )
    points = []
    for current_day in _report_days(start_date, end_date):
        target_moment = timezone.make_aware(
            datetime.combine(current_day, time.max)
        )
        status_counts = {
            Room.RoomStatus.AVAILABLE: 0,
            Room.RoomStatus.OCCUPIED: 0,
            Room.RoomStatus.MAINTENANCE: 0,
            Room.RoomStatus.CLEANING: 0,
        }
        for room in rooms:
            status_value = _analytics_room_status_at(room, target_moment)
            if status_value in status_counts:
                status_counts[status_value] += 1
        points.append(
            {
                "date": current_day,
                "available": status_counts[Room.RoomStatus.AVAILABLE],
                "occupied": status_counts[Room.RoomStatus.OCCUPIED],
                "maintenance": status_counts[Room.RoomStatus.MAINTENANCE],
                "cleaning": status_counts[Room.RoomStatus.CLEANING],
            }
        )
    return points


def _analytics_daily_occupancy_rows(start_date, end_date, room_type=""):
    rooms = list(_analytics_room_queryset(room_type))
    room_ids = {room.pk for room in rooms}
    total_rooms = len(room_ids)
    active_statuses = [
        Booking.BookingStatus.PENDING,
        Booking.BookingStatus.CONFIRMED,
        Booking.BookingStatus.CHECKED_IN,
    ]
    bookings = list(
        Booking.objects.filter(
            room_id__in=room_ids,
            status__in=active_statuses,
            check_in__lte=end_date,
            check_out__gte=start_date,
        ).values("room_id", "check_in", "check_out")
    ) if room_ids else []
    occupied_sets = {current_day: set() for current_day in _report_days(start_date, end_date)}
    for booking in bookings:
        overlap_start = max(booking["check_in"], start_date)
        overlap_end = min(
            end_date,
            booking["check_out"] if booking["check_in"] == booking["check_out"] else booking["check_out"] - timedelta(days=1),
        )
        current_day = overlap_start
        while current_day <= overlap_end:
            occupied_sets[current_day].add(booking["room_id"])
            current_day += timedelta(days=1)

    revenue_by_day = daily_booking_revenue_map(start_date, end_date, room_type) if room_ids else {}

    rows = []
    for current_day in _report_days(start_date, end_date):
        occupied_rooms = len(occupied_sets[current_day])
        occupancy_percent = round((occupied_rooms / total_rooms) * 100, 2) if total_rooms else 0
        rows.append(
            {
                "date": current_day,
                "occupied_rooms": occupied_rooms,
                "occupancy_percent": occupancy_percent,
                "revenue_collected": revenue_by_day.get(current_day, Decimal("0")),
            }
        )
    return rows


def _analytics_chart(chart_id, title, chart_type, labels, datasets):
    return {
        "id": chart_id,
        "title": title,
        "type": chart_type,
        "labels": labels,
        "datasets": datasets,
    }


def _build_rooms_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    room_type = filters["room_type"]

    rooms = list(_analytics_room_queryset(room_type))
    bookings = list(_analytics_booking_stay_queryset(start_date, end_date, room_type))
    occupancy_rows = _analytics_daily_occupancy_rows(start_date, end_date, room_type)
    status_points = _analytics_room_status_points(start_date, end_date, room_type)

    type_rollup = defaultdict(int)
    floor_rollup = defaultdict(int)
    for room in rooms:
        type_rollup[room.get_room_type_display()] += 1
        floor_rollup[_analytics_room_floor(room.room_number)] += 1

    room_type_booking_rollup = defaultdict(int)
    total_nights = 0
    for booking in bookings:
        stay_nights = booking_occupied_days_in_range(
            booking.check_in,
            booking.check_out,
            start_date,
            end_date,
        )
        total_nights += stay_nights
        room_type_booking_rollup[booking.room.get_room_type_display()] += 1

    average_occupancy = round(
        sum((row["occupancy_percent"] for row in occupancy_rows), 0) / len(occupancy_rows),
        2,
    ) if occupancy_rows else 0
    average_length_of_stay = round(total_nights / len(bookings), 2) if bookings else 0

    occupancy_series = _analytics_aggregate_points(
        [{"date": row["date"], "occupancy_percent": row["occupancy_percent"]} for row in occupancy_rows],
        filters["period"],
        start_date,
        end_date,
        sum_fields=(),
        avg_fields=("occupancy_percent",),
    )
    status_series = _analytics_aggregate_points(
        status_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("available", "occupied", "maintenance", "cleaning"),
    )

    type_rows = [[label, count] for label, count in sorted(type_rollup.items())]
    floor_rows = [[label, count] for label, count in sorted(floor_rollup.items())]
    booked_type_rows = [[label, count] for label, count in sorted(room_type_booking_rollup.items(), key=lambda item: item[1], reverse=True)]

    return {
        "key": "rooms-analytics",
        "title": "Rooms Analytics",
        "sheet_title": "Rooms Analytics",
        "filename_prefix": "rooms-analytics",
        "subtitle": "Room inventory, occupancy, booking mix, maintenance, and average stay length across the selected period.",
        "summary": {
            "rooms": len(rooms),
            "average_occupancy": average_occupancy,
            "average_length_of_stay": average_length_of_stay,
        },
        "metrics": [
            {"label": "Total rooms", "value": len(rooms), "export_value": len(rooms)},
            {"label": "Average occupancy", "value": _display_percent(average_occupancy), "export_value": float(average_occupancy)},
            {"label": "Average stay", "value": f"{average_length_of_stay:.2f} nights", "export_value": float(average_length_of_stay)},
            {"label": "Most booked room type", "value": booked_type_rows[0][0] if booked_type_rows else "-", "export_value": booked_type_rows[0][0] if booked_type_rows else ""},
        ],
        "charts": [
            _analytics_chart(
                "rooms-by-type-chart",
                "Rooms by type",
                "doughnut",
                [row[0] for row in type_rows],
                [{"label": "Rooms", "data": [row[1] for row in type_rows], "backgroundColor": ["#23444B", "#3D7DFF", "#F4B942", "#78C0A8"]}],
            ),
            _analytics_chart(
                "rooms-occupancy-chart",
                "Occupancy trend",
                "line",
                occupancy_series["labels"],
                [{"label": "Occupancy %", "data": occupancy_series["values"]["occupancy_percent"], "borderColor": "#3D7DFF", "backgroundColor": "rgba(61,125,255,0.14)", "fill": True, "tension": 0.35}],
            ),
            _analytics_chart(
                "rooms-status-chart",
                "Room status over time",
                "bar",
                status_series["labels"],
                [
                    {"label": "Available", "data": status_series["values"]["available"], "backgroundColor": "#78C0A8"},
                    {"label": "Occupied", "data": status_series["values"]["occupied"], "backgroundColor": "#3D7DFF"},
                    {"label": "Maintenance", "data": status_series["values"]["maintenance"], "backgroundColor": "#D95D39"},
                    {"label": "Cleaning", "data": status_series["values"]["cleaning"], "backgroundColor": "#F4B942"},
                ],
            ),
        ],
        "tables": [
            {
                "title": "Rooms by type",
                "headers": ["Room Type", "Total Rooms"],
                "rows": type_rows,
                "export_rows": type_rows,
                "summary_row": ["TOTALS", len(rooms)],
                "export_summary_row": ["TOTALS", len(rooms)],
            },
            {
                "title": "Rooms by floor",
                "headers": ["Floor", "Total Rooms"],
                "rows": floor_rows,
                "export_rows": floor_rows,
                "summary_row": ["TOTALS", len(rooms)],
                "export_summary_row": ["TOTALS", len(rooms)],
            },
            {
                "title": "Most booked room types",
                "headers": ["Room Type", "Bookings"],
                "rows": booked_type_rows,
                "export_rows": booked_type_rows,
                "summary_row": ["TOTALS", len(bookings)],
                "export_summary_row": ["TOTALS", len(bookings)],
            },
        ],
    }


def _build_bookings_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    room_type = filters["room_type"]

    bookings_created = list(_analytics_booking_creation_queryset(start_date, end_date, room_type))
    stay_bookings = list(_analytics_booking_stay_queryset(start_date, end_date, room_type))

    status_rollup = defaultdict(int)
    bookings_by_day = defaultdict(int)
    cancellations = 0
    weekday_rollup = defaultdict(int)
    month_rollup = defaultdict(int)
    new_guest_ids = set()
    returning_guest_ids = set()

    for booking in bookings_created:
        status_rollup[booking.get_status_display()] += 1
        bookings_by_day[booking.created_at.date()] += 1
        weekday_rollup[booking.created_at.strftime("%A")] += 1
        month_rollup[booking.created_at.strftime("%b %Y")] += 1
        if booking.status == Booking.BookingStatus.CANCELLED:
            cancellations += 1
        prior_exists = Booking.objects.filter(
            guest_id=booking.guest_id,
            created_at__lt=booking.created_at,
        ).exists()
        if prior_exists:
            returning_guest_ids.add(booking.guest_id)
        else:
            new_guest_ids.add(booking.guest_id)

    day_points = [
        {"date": current_day, "bookings": bookings_by_day.get(current_day, 0)}
        for current_day in _report_days(start_date, end_date)
    ]
    booking_series = _analytics_aggregate_points(
        day_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("bookings",),
    )

    cancellation_rate = round((cancellations / len(bookings_created)) * 100, 2) if bookings_created else 0
    average_length_of_stay = round(
        sum((booking.nights for booking in stay_bookings), 0) / len(stay_bookings),
        2,
    ) if stay_bookings else 0
    status_rows = [[label, count] for label, count in sorted(status_rollup.items())]
    peak_period_rows = sorted(weekday_rollup.items(), key=lambda item: item[1], reverse=True)
    peak_season_rows = sorted(month_rollup.items(), key=lambda item: item[1], reverse=True)

    return {
        "key": "bookings-analytics",
        "title": "Bookings Analytics",
        "sheet_title": "Bookings Analytics",
        "filename_prefix": "bookings-analytics",
        "subtitle": "Booking volumes, statuses, cancellations, guest mix, and peak periods based on live reservation data.",
        "summary": {
            "total_bookings": len(bookings_created),
            "cancellation_rate": cancellation_rate,
            "new_guests": len(new_guest_ids),
            "returning_guests": len(returning_guest_ids),
        },
        "metrics": [
            {"label": "Total bookings", "value": len(bookings_created), "export_value": len(bookings_created)},
            {"label": "Cancellation rate", "value": _display_percent(cancellation_rate), "export_value": float(cancellation_rate)},
            {"label": "New guests", "value": len(new_guest_ids), "export_value": len(new_guest_ids)},
            {"label": "Returning guests", "value": len(returning_guest_ids), "export_value": len(returning_guest_ids)},
            {"label": "Average stay", "value": f"{average_length_of_stay:.2f} nights", "export_value": float(average_length_of_stay)},
        ],
        "charts": [
            _analytics_chart(
                "bookings-volume-chart",
                "Bookings over time",
                "line",
                booking_series["labels"],
                [{"label": "Bookings", "data": booking_series["values"]["bookings"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.12)", "fill": True, "tension": 0.3}],
            ),
            _analytics_chart(
                "bookings-status-chart",
                "Bookings by status",
                "pie",
                [row[0] for row in status_rows],
                [{"label": "Bookings", "data": [row[1] for row in status_rows], "backgroundColor": ["#3D7DFF", "#F4B942", "#D95D39", "#78C0A8", "#6C757D"]}],
            ),
            _analytics_chart(
                "bookings-guest-mix-chart",
                "New vs returning guests",
                "bar",
                ["Guests"],
                [
                    {"label": "New", "data": [len(new_guest_ids)], "backgroundColor": "#78C0A8"},
                    {"label": "Returning", "data": [len(returning_guest_ids)], "backgroundColor": "#3D7DFF"},
                ],
            ),
        ],
        "tables": [
            {
                "title": "Booking statuses",
                "headers": ["Status", "Bookings"],
                "rows": status_rows,
                "export_rows": status_rows,
                "summary_row": ["TOTALS", len(bookings_created)],
                "export_summary_row": ["TOTALS", len(bookings_created)],
            },
            {
                "title": "Peak booking periods",
                "headers": ["Day / Period", "Bookings"],
                "rows": [[label, count] for label, count in peak_period_rows[:7]],
                "export_rows": [[label, count] for label, count in peak_period_rows[:7]],
                "summary_row": ["TOTALS", len(bookings_created)],
                "export_summary_row": ["TOTALS", len(bookings_created)],
            },
            {
                "title": "Peak booking seasons",
                "headers": ["Month", "Bookings"],
                "rows": [[label, count] for label, count in peak_season_rows[:12]],
                "export_rows": [[label, count] for label, count in peak_season_rows[:12]],
                "summary_row": ["TOTALS", len(bookings_created)],
                "export_summary_row": ["TOTALS", len(bookings_created)],
            },
        ],
    }


def _build_revenue_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    room_type = filters["room_type"]
    money_field = DecimalField(max_digits=12, decimal_places=2)

    revenue_breakdown = revenue_components(start_date, end_date, room_type)
    booking_stays = booking_revenue_queryset(start_date, end_date, room_type)

    booking_total = revenue_breakdown["booking_revenue"]
    event_total = revenue_breakdown["event_revenue"]
    pos_total = revenue_breakdown["pos_sales"]
    total_revenue = revenue_breakdown["gross_revenue"]
    owner_withdrawals_total = revenue_breakdown["owner_withdrawals"]
    net_revenue_total = revenue_breakdown["net_revenue"]

    fully_paid_bookings = booking_stays.annotate(
        paid_total=Coalesce(Sum("payments__amount"), Value(0, output_field=money_field)),
        balance=ExpressionWrapper(F("total_amount") - F("paid_total"), output_field=money_field),
    )
    outstanding_total = fully_paid_bookings.aggregate(
        total=Coalesce(Sum("balance", filter=Q(balance__gt=0)), Value(0, output_field=money_field))
    )["total"]
    fully_paid_count = fully_paid_bookings.filter(balance__lte=0).count()

    revenue_points = []
    room_type_rollup = defaultdict(Decimal)
    revenue_by_day = daily_total_revenue_map(start_date, end_date, room_type)
    withdrawals_by_day = daily_owner_withdrawals_map(start_date, end_date)
    net_revenue_by_day = daily_net_revenue_map(start_date, end_date, room_type)
    for current_day in _report_days(start_date, end_date):
        revenue_points.append(
            {
                "date": current_day,
                "gross_revenue": revenue_by_day.get(current_day, Decimal("0")),
                "owner_withdrawals": withdrawals_by_day.get(current_day, Decimal("0")),
                "net_revenue": net_revenue_by_day.get(current_day, Decimal("0")),
            }
        )

    for row in (
        booking_stays.values("room__room_type")
        .annotate(total=Coalesce(Sum("total_amount"), Value(0, output_field=money_field)))
        .order_by("room__room_type")
    ):
        label = dict(Room.RoomType.choices).get(row["room__room_type"], row["room__room_type"] or "Unknown")
        room_type_rollup[label] += Decimal(str(row["total"] or 0))

    revenue_series = _analytics_aggregate_points(
        revenue_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("gross_revenue", "owner_withdrawals", "net_revenue"),
    )

    average_revenue_per_booking = round(
        float(booking_total) / booking_stays.count(),
        2,
    ) if booking_stays.exists() else 0

    room_type_rows = [
        [label, _display_money(total)]
        for label, total in sorted(room_type_rollup.items(), key=lambda item: item[1], reverse=True)
    ]
    room_type_export_rows = [
        [label, float(total)]
        for label, total in sorted(room_type_rollup.items(), key=lambda item: item[1], reverse=True)
    ]

    expense_available = False
    expense_note = "No expense data available in the current system."

    return {
        "key": "revenue-analytics",
        "title": "Revenue Analytics",
        "sheet_title": "Revenue Analytics",
        "filename_prefix": "revenue-analytics",
        "subtitle": "Gross revenue, owner withdrawals, net revenue, room-type performance, POS sales, and payment completion for the selected period.",
        "summary": {
            "total_revenue": total_revenue,
            "net_revenue": net_revenue_total,
            "owner_withdrawals": owner_withdrawals_total,
            "outstanding_total": outstanding_total,
            "fully_paid_count": fully_paid_count,
        },
        "metrics": [
            {"label": "Gross revenue", "value": _display_money(total_revenue), "export_value": float(total_revenue)},
            {"label": "Net revenue", "value": _display_money(net_revenue_total), "export_value": float(net_revenue_total)},
            {"label": "Owner withdrawals", "value": _display_money(owner_withdrawals_total), "export_value": float(owner_withdrawals_total)},
            {"label": "Average revenue per booking", "value": _display_money(average_revenue_per_booking), "export_value": float(average_revenue_per_booking)},
            {"label": "Outstanding payments", "value": _display_money(outstanding_total), "export_value": float(outstanding_total)},
            {"label": "Fully paid bookings", "value": fully_paid_count, "export_value": fully_paid_count},
            {"label": "Expenses", "value": expense_note if not expense_available else "-", "export_value": expense_note if not expense_available else ""},
        ],
        "charts": [
            _analytics_chart(
                "revenue-trend-chart",
                "Gross vs net revenue",
                "line",
                revenue_series["labels"],
                [
                    {"label": "Gross Revenue", "data": revenue_series["values"]["gross_revenue"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.14)", "fill": True, "tension": 0.35},
                    {"label": "Net Revenue", "data": revenue_series["values"]["net_revenue"], "borderColor": "#CFAE84", "backgroundColor": "rgba(207,174,132,0.08)", "fill": False, "tension": 0.35},
                ],
            ),
            _analytics_chart(
                "revenue-room-type-chart",
                "Revenue by room type",
                "doughnut",
                list(room_type_rollup.keys()),
                [{"label": "Revenue", "data": [float(total) for total in room_type_rollup.values()], "backgroundColor": ["#23444B", "#3D7DFF", "#F4B942", "#78C0A8"]}],
            ),
            _analytics_chart(
                "revenue-payment-status-chart",
                "Payment completion",
                "pie",
                ["Outstanding", "Fully Paid"],
                [{"label": "Bookings", "data": [float(outstanding_total), fully_paid_count], "backgroundColor": ["#D95D39", "#78C0A8"]}],
            ),
        ],
        "tables": [
            {
                "title": "Revenue by room type",
                "headers": ["Room Type", "Revenue"],
                "rows": room_type_rows,
                "export_rows": room_type_export_rows,
                "summary_row": ["TOTALS", _display_money(booking_total)],
                "export_summary_row": ["TOTALS", float(booking_total)],
            },
            {
                "title": "Revenue overview",
                "headers": ["Metric", "Value"],
                "rows": [
                    ["Gross revenue", _display_money(total_revenue)],
                    ["Owner withdrawals", _display_money(owner_withdrawals_total)],
                    ["Net revenue", _display_money(net_revenue_total)],
                    ["Booking revenue", _display_money(booking_total)],
                    ["Event revenue", _display_money(event_total)],
                    ["POS sales", _display_money(pos_total)],
                    ["Payments received", _display_money(revenue_breakdown["payments_received"])],
                    ["Outstanding balances", _display_money(outstanding_total)],
                    ["Revenue vs expenses", expense_note],
                ],
                "export_rows": [
                    ["Gross revenue", float(total_revenue)],
                    ["Owner withdrawals", float(owner_withdrawals_total)],
                    ["Net revenue", float(net_revenue_total)],
                    ["Booking revenue", float(booking_total)],
                    ["Event revenue", float(event_total)],
                    ["POS sales", float(pos_total)],
                    ["Payments received", float(revenue_breakdown["payments_received"])],
                    ["Outstanding balances", float(outstanding_total)],
                    ["Revenue vs expenses", expense_note],
                ],
                "summary_row": ["TOTALS", _display_money(net_revenue_total)],
                "export_summary_row": ["TOTALS", float(net_revenue_total)],
            },
        ],
    }


def _build_staff_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    department = filters["department"]
    staff_role = filters["staff_role"]

    employees = list(_analytics_employee_queryset(department, staff_role))
    employee_ids = [employee.pk for employee in employees]
    status_rollup = defaultdict(int)
    department_rollup = defaultdict(int)
    role_rollup = defaultdict(int)
    for employee in employees:
        status_rollup[employee.get_employment_status_display()] += 1
        department_rollup[employee.department or "Unassigned"] += 1
        role_rollup[employee.get_position_display()] += 1

    leave_requests = LeaveRequest.objects.filter(
        employee_id__in=employee_ids,
        start_date__lte=end_date,
        end_date__gte=start_date,
    )
    hires_points = [
        {"date": current_day, "hires": 0, "terminations": 0}
        for current_day in _report_days(start_date, end_date)
    ]
    hire_lookup = {row["date"]: row for row in hires_points}
    for employee in employees:
        if start_date <= employee.start_date <= end_date:
            hire_lookup[employee.start_date]["hires"] += 1
        if employee.termination_date and start_date <= employee.termination_date <= end_date:
            hire_lookup[employee.termination_date]["terminations"] += 1

    hire_series = _analytics_aggregate_points(
        hires_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("hires", "terminations"),
    )

    leave_employee_rows = []
    leave_department_rollup = defaultdict(lambda: {"requests": 0, "days": 0})
    for row in (
        leave_requests.values("employee__first_name", "employee__last_name", "employee__department")
        .annotate(total_requests=Count("id"), total_days=Coalesce(Sum("days"), Value(0)))
        .order_by("-total_requests", "employee__last_name")
    ):
        full_name = f"{row['employee__first_name']} {row['employee__last_name']}".strip()
        leave_employee_rows.append([full_name, row["employee__department"] or "Unassigned", row["total_requests"], row["total_days"]])
        department_key = row["employee__department"] or "Unassigned"
        leave_department_rollup[department_key]["requests"] += row["total_requests"]
        leave_department_rollup[department_key]["days"] += row["total_days"]

    leave_department_rows = [
        [dept, values["requests"], values["days"]]
        for dept, values in sorted(leave_department_rollup.items())
    ]

    return {
        "key": "staff-analytics",
        "title": "Staff Analytics",
        "sheet_title": "Staff Analytics",
        "filename_prefix": "staff-analytics",
        "subtitle": "Headcount, employment status, leave frequency, hires, and terminations driven directly from HR records.",
        "summary": {
            "staff_total": len(employees),
            "leave_requests": leave_requests.count(),
        },
        "metrics": [
            {"label": "Total staff", "value": len(employees), "export_value": len(employees)},
            {"label": "Departments", "value": len(department_rollup), "export_value": len(department_rollup)},
            {"label": "Roles", "value": len(role_rollup), "export_value": len(role_rollup)},
            {"label": "Leave requests", "value": leave_requests.count(), "export_value": leave_requests.count()},
        ],
        "charts": [
            _analytics_chart(
                "staff-department-chart",
                "Staff by department",
                "bar",
                list(department_rollup.keys()),
                [{"label": "Employees", "data": list(department_rollup.values()), "backgroundColor": "#23444B"}],
            ),
            _analytics_chart(
                "staff-status-chart",
                "Staff status breakdown",
                "doughnut",
                list(status_rollup.keys()),
                [{"label": "Employees", "data": list(status_rollup.values()), "backgroundColor": ["#78C0A8", "#D95D39", "#3D7DFF", "#F4B942", "#6C757D", "#A06CD5"]}],
            ),
            _analytics_chart(
                "staff-hires-chart",
                "New hires vs terminations",
                "line",
                hire_series["labels"],
                [
                    {"label": "New hires", "data": hire_series["values"]["hires"], "borderColor": "#78C0A8", "backgroundColor": "rgba(120,192,168,0.16)", "fill": False, "tension": 0.3},
                    {"label": "Terminations", "data": hire_series["values"]["terminations"], "borderColor": "#D95D39", "backgroundColor": "rgba(217,93,57,0.16)", "fill": False, "tension": 0.3},
                ],
            ),
        ],
        "tables": [
            {
                "title": "Staff by role",
                "headers": ["Role", "Employees"],
                "rows": [[role, count] for role, count in sorted(role_rollup.items(), key=lambda item: item[1], reverse=True)],
                "export_rows": [[role, count] for role, count in sorted(role_rollup.items(), key=lambda item: item[1], reverse=True)],
                "summary_row": ["TOTALS", len(employees)],
                "export_summary_row": ["TOTALS", len(employees)],
            },
            {
                "title": "Leave frequency per employee",
                "headers": ["Employee", "Department", "Requests", "Days"],
                "rows": leave_employee_rows[:20],
                "export_rows": leave_employee_rows,
                "summary_row": ["TOTALS", "", leave_requests.count(), sum((row[3] for row in leave_employee_rows), 0)],
                "export_summary_row": ["TOTALS", "", leave_requests.count(), sum((row[3] for row in leave_employee_rows), 0)],
            },
            {
                "title": "Leave frequency per department",
                "headers": ["Department", "Requests", "Days"],
                "rows": leave_department_rows,
                "export_rows": leave_department_rows,
                "summary_row": ["TOTALS", leave_requests.count(), sum((row[2] for row in leave_department_rows), 0)],
                "export_summary_row": ["TOTALS", leave_requests.count(), sum((row[2] for row in leave_department_rows), 0)],
            },
        ],
    }


def _build_housekeeping_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    room_type = filters["room_type"]

    logs = HousekeepingItemLog.objects.select_related("item", "room").filter(
        used_at__date__range=[start_date, end_date]
    )
    if room_type:
        logs = logs.filter(Q(room__room_type=room_type) | Q(room__isnull=True))

    items = list(HousekeepingItem.objects.order_by("name"))
    low_stock_items = [item for item in items if item.is_low_stock]
    usage_points = []
    usage_by_item = defaultdict(lambda: {"used": Decimal("0"), "entries": 0, "initial": Decimal("0"), "stock": Decimal("0")})

    for current_day in _report_days(start_date, end_date):
        day_total = logs.filter(used_at__date=current_day).aggregate(
            total=Coalesce(Sum("quantity_used"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=3)))
        )["total"]
        usage_points.append({"date": current_day, "used": Decimal(str(day_total or 0))})

    for row in (
        logs.values("item_name", "unit")
        .annotate(
            total_used=Coalesce(Sum("quantity_used"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=3))),
            entries=Count("id"),
            initial_quantity=Coalesce(Max("initial_quantity"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=3))),
            stock=Coalesce(Max("item__quantity_in_stock"), Value(0, output_field=DecimalField(max_digits=12, decimal_places=3))),
        )
        .order_by("-total_used", "item_name")
    ):
        usage_by_item[row["item_name"]] = {
            "used": Decimal(str(row["total_used"] or 0)),
            "entries": row["entries"],
            "initial": Decimal(str(row["initial_quantity"] or 0)),
            "stock": Decimal(str(row["stock"] or 0)),
            "unit": row["unit"],
        }

    total_used = sum((values["used"] for values in usage_by_item.values()), Decimal("0"))
    total_initial = sum((Decimal(str(item.initial_quantity or 0)) for item in items), Decimal("0"))
    total_stock = sum((Decimal(str(item.quantity_in_stock or 0)) for item in items), Decimal("0"))

    depletion_rows = []
    for item_name, values in usage_by_item.items():
        depletion_rate = round((float(values["used"]) / float(values["initial"])) * 100, 2) if values["initial"] else 0
        depletion_rows.append([item_name, _display_quantity(values["initial"]), _display_quantity(values["used"]), _display_quantity(values["stock"]), values["unit"], _display_percent(depletion_rate)])

    usage_series = _analytics_aggregate_points(
        usage_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("used",),
    )

    most_consumed = list(usage_by_item.items())[:10]

    return {
        "key": "housekeeping-analytics",
        "title": "Housekeeping Analytics",
        "sheet_title": "Housekeeping Analytics",
        "filename_prefix": "housekeeping-analytics",
        "subtitle": "Consumption trends, low-stock items, depletion rate, and the overall stock picture from housekeeping logs.",
        "summary": {
            "used": total_used,
            "low_stock": len(low_stock_items),
        },
        "metrics": [
            {"label": "Initial stock", "value": _display_quantity(total_initial), "export_value": float(total_initial)},
            {"label": "Current stock", "value": _display_quantity(total_stock), "export_value": float(total_stock)},
            {"label": "Items used", "value": _display_quantity(total_used), "export_value": float(total_used)},
            {"label": "Low stock alerts", "value": len(low_stock_items), "export_value": len(low_stock_items)},
        ],
        "charts": [
            _analytics_chart(
                "housekeeping-usage-chart",
                "Items used over time",
                "line",
                usage_series["labels"],
                [{"label": "Items used", "data": usage_series["values"]["used"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.12)", "fill": True, "tension": 0.35}],
            ),
            _analytics_chart(
                "housekeeping-consumed-chart",
                "Most consumed items",
                "bar",
                [item_name for item_name, _values in most_consumed],
                [{"label": "Quantity used", "data": [float(values["used"]) for _item_name, values in most_consumed], "backgroundColor": "#3D7DFF"}],
            ),
            _analytics_chart(
                "housekeeping-stock-chart",
                "Initial vs current vs used",
                "pie",
                ["Initial Stock", "Current Stock", "Used"],
                [{"label": "Stock", "data": [float(total_initial), float(total_stock), float(total_used)], "backgroundColor": ["#23444B", "#78C0A8", "#F4B942"]}],
            ),
        ],
        "tables": [
            {
                "title": "Items running low",
                "headers": ["Item", "Current Stock", "Threshold", "Unit"],
                "rows": [[item.name, _display_quantity(item.quantity_in_stock), _display_quantity(item.effective_low_stock_threshold), item.unit] for item in low_stock_items],
                "export_rows": [[item.name, float(item.quantity_in_stock or 0), float(item.effective_low_stock_threshold or 0), item.unit] for item in low_stock_items],
                "summary_row": ["TOTALS", len(low_stock_items), "", ""],
                "export_summary_row": ["TOTALS", len(low_stock_items), "", ""],
            },
            {
                "title": "Stock depletion rate per item",
                "headers": ["Item", "Initial Qty", "Used", "Current Stock", "Unit", "Depletion Rate"],
                "rows": depletion_rows,
                "export_rows": [
                    [item_name, float(values["initial"]), float(values["used"]), float(values["stock"]), values["unit"], round((float(values["used"]) / float(values["initial"])) * 100, 2) if values["initial"] else 0]
                    for item_name, values in usage_by_item.items()
                ],
                "summary_row": ["TOTALS", _display_quantity(total_initial), _display_quantity(total_used), _display_quantity(total_stock), "", ""],
                "export_summary_row": ["TOTALS", float(total_initial), float(total_used), float(total_stock), "", ""],
            },
        ],
    }


def _build_roster_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    department = filters["department"]
    staff_role = filters["staff_role"]

    rotas = list(
        _analytics_employee_queryset(department, staff_role)
        .prefetch_related("rota_entries")
    )
    daily_counts = {current_day: 0 for current_day in _report_days(start_date, end_date)}
    shift_rollup = defaultdict(int)
    department_rollup = defaultdict(int)
    employee_count_by_day = defaultdict(set)

    rota_entries = list(
        Rota.objects.select_related("employee").filter(
            employee__in=[employee.pk for employee in rotas],
            period_end__gte=start_date,
            period_start__lte=end_date,
        )
    )
    for rota in rota_entries:
        shift_label = rota.operating_hours if rota.operating_hours != "-" else "Unspecified shift"
        department_label = rota.employee.department or "Unassigned"
        overlap_start = max(rota.period_start, start_date)
        overlap_end = min(rota.period_end, end_date)
        current_day = overlap_start
        while current_day <= overlap_end:
            daily_counts[current_day] += 1
            employee_count_by_day[current_day].add(rota.employee_id)
            current_day += timedelta(days=1)
        shift_rollup[shift_label] += 1
        department_rollup[department_label] += 1

    coverage_points = [
        {"date": current_day, "coverage": len(employee_count_by_day[current_day])}
        for current_day in _report_days(start_date, end_date)
    ]
    coverage_series = _analytics_aggregate_points(
        coverage_points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("coverage",),
    )
    average_coverage = (
        sum((len(employee_count_by_day[current_day]) for current_day in _report_days(start_date, end_date)), 0)
        / max(len(daily_counts), 1)
    )
    gap_rows = []
    for current_day in _report_days(start_date, end_date):
        employees_count = len(employee_count_by_day[current_day])
        if employees_count == 0 or employees_count < average_coverage:
            gap_rows.append([
                _display_date(current_day),
                employees_count,
                "Gap" if employees_count == 0 else "Understaffed",
            ])

    return {
        "key": "roster-analytics",
        "title": "Duty Roster Analytics",
        "sheet_title": "Duty Roster Analytics",
        "filename_prefix": "duty-roster-analytics",
        "subtitle": "Coverage by day and week, employee distribution by shift and department, plus gap and understaffing signals.",
        "summary": {
            "rota_entries": len(rota_entries),
            "gap_days": len([row for row in gap_rows if row[2] == "Gap"]),
        },
        "metrics": [
            {"label": "Rota entries", "value": len(rota_entries), "export_value": len(rota_entries)},
            {"label": "Distinct shifts", "value": len(shift_rollup), "export_value": len(shift_rollup)},
            {"label": "Departments covered", "value": len(department_rollup), "export_value": len(department_rollup)},
            {"label": "Coverage gaps", "value": len([row for row in gap_rows if row[2] == "Gap"]), "export_value": len([row for row in gap_rows if row[2] == "Gap"])},
        ],
        "charts": [
            _analytics_chart(
                "roster-coverage-chart",
                "Shift coverage by day",
                "line",
                coverage_series["labels"],
                [{"label": "Employees scheduled", "data": coverage_series["values"]["coverage"], "borderColor": "#23444B", "backgroundColor": "rgba(35,68,75,0.12)", "fill": True, "tension": 0.3}],
            ),
            _analytics_chart(
                "roster-shift-chart",
                "Employees per shift",
                "bar",
                list(shift_rollup.keys()),
                [{"label": "Rota entries", "data": list(shift_rollup.values()), "backgroundColor": "#3D7DFF"}],
            ),
            _analytics_chart(
                "roster-department-chart",
                "Department coverage",
                "doughnut",
                list(department_rollup.keys()),
                [{"label": "Entries", "data": list(department_rollup.values()), "backgroundColor": ["#23444B", "#3D7DFF", "#F4B942", "#78C0A8", "#D95D39"]}],
            ),
        ],
        "tables": [
            {
                "title": "Employees per shift",
                "headers": ["Shift", "Rota Entries"],
                "rows": [[shift, count] for shift, count in sorted(shift_rollup.items(), key=lambda item: item[1], reverse=True)],
                "export_rows": [[shift, count] for shift, count in sorted(shift_rollup.items(), key=lambda item: item[1], reverse=True)],
                "summary_row": ["TOTALS", len(rota_entries)],
                "export_summary_row": ["TOTALS", len(rota_entries)],
            },
            {
                "title": "Roster gaps or understaffed days",
                "headers": ["Date", "Employees Scheduled", "Status"],
                "rows": gap_rows,
                "export_rows": gap_rows,
                "summary_row": ["TOTALS", len(gap_rows), ""],
                "export_summary_row": ["TOTALS", len(gap_rows), ""],
            },
        ],
    }


def _build_operations_analytics_section(filters):
    start_date = filters["start_date"]
    end_date = filters["end_date"]
    room_type = filters["room_type"]
    department = filters["department"]
    staff_role = filters["staff_role"]

    room_filter = Q()
    if room_type:
        room_filter = Q(room__room_type=room_type)

    employee_filter = Q()
    if department:
        employee_filter &= Q(employee__department=department)
    if staff_role:
        employee_filter &= Q(employee__position=staff_role)

    points = []
    for current_day in _report_days(start_date, end_date):
        reservations = Booking.objects.filter(created_at__date=current_day).filter(room_filter).count()
        check_ins = Booking.objects.filter(check_in=current_day).filter(room_filter).count()
        check_outs = Booking.objects.filter(check_out=current_day).filter(room_filter).count()
        cancellations = Booking.objects.filter(status=Booking.BookingStatus.CANCELLED, updated_at__date=current_day).filter(room_filter).count()
        housekeeping_entries = HousekeepingItemLog.objects.filter(used_at__date=current_day)
        if room_type:
            housekeeping_entries = housekeeping_entries.filter(Q(room__room_type=room_type) | Q(room__isnull=True))
        housekeeping_count = housekeeping_entries.count()
        payments_count = Payment.objects.filter(paid_at__date=current_day).filter(Q() if not room_type else Q(booking__room__room_type=room_type)).count()
        staff_on_rota = Rota.objects.filter(period_start__lte=current_day, period_end__gte=current_day).filter(employee_filter).values("employee_id").distinct().count()
        activity_score = reservations + check_ins + check_outs + cancellations + housekeeping_count + payments_count + staff_on_rota
        points.append(
            {
                "date": current_day,
                "reservations": reservations,
                "check_ins": check_ins,
                "check_outs": check_outs,
                "housekeeping": housekeeping_count,
                "payments": payments_count,
                "staff": staff_on_rota,
                "activity": activity_score,
            }
        )

    busiest_rows = sorted(points, key=lambda row: row["activity"], reverse=True)
    activity_series = _analytics_aggregate_points(
        points,
        filters["period"],
        start_date,
        end_date,
        sum_fields=("reservations", "check_ins", "check_outs", "housekeeping", "payments", "staff", "activity"),
    )

    return {
        "key": "operations-analytics",
        "title": "Operations Overview Analytics",
        "sheet_title": "Operations Analytics",
        "filename_prefix": "operations-analytics",
        "subtitle": "Cross-module operating load trends, busiest periods, and the daily activity mix across reservations, housekeeping, payments, and staffing.",
        "summary": {
            "busiest_day": busiest_rows[0]["date"] if busiest_rows else None,
            "busiest_score": busiest_rows[0]["activity"] if busiest_rows else 0,
        },
        "metrics": [
            {"label": "Busiest day", "value": _display_date(busiest_rows[0]["date"]) if busiest_rows else "-", "export_value": busiest_rows[0]["date"] if busiest_rows else ""},
            {"label": "Peak activity score", "value": busiest_rows[0]["activity"] if busiest_rows else 0, "export_value": busiest_rows[0]["activity"] if busiest_rows else 0},
            {"label": "Total reservation actions", "value": sum((row["reservations"] + row["check_ins"] + row["check_outs"] for row in points), 0), "export_value": sum((row["reservations"] + row["check_ins"] + row["check_outs"] for row in points), 0)},
            {"label": "Housekeeping actions", "value": sum((row["housekeeping"] for row in points), 0), "export_value": sum((row["housekeeping"] for row in points), 0)},
        ],
        "charts": [
            _analytics_chart(
                "operations-activity-chart",
                "Operations trend over time",
                "line",
                activity_series["labels"],
                [{"label": "Activity score", "data": activity_series["values"]["activity"], "borderColor": "#D95D39", "backgroundColor": "rgba(217,93,57,0.14)", "fill": True, "tension": 0.35}],
            ),
            _analytics_chart(
                "operations-module-chart",
                "Module activity mix",
                "bar",
                activity_series["labels"],
                [
                    {"label": "Reservations", "data": activity_series["values"]["reservations"], "backgroundColor": "#23444B"},
                    {"label": "Check-ins", "data": activity_series["values"]["check_ins"], "backgroundColor": "#3D7DFF"},
                    {"label": "Check-outs", "data": activity_series["values"]["check_outs"], "backgroundColor": "#78C0A8"},
                    {"label": "Housekeeping", "data": activity_series["values"]["housekeeping"], "backgroundColor": "#F4B942"},
                    {"label": "Payments", "data": activity_series["values"]["payments"], "backgroundColor": "#A06CD5"},
                ],
            ),
        ],
        "tables": [
            {
                "title": "Busiest days and periods",
                "headers": ["Date", "Activity Score", "Reservations", "Check-ins", "Check-outs", "Housekeeping", "Payments", "Staff on Rota"],
                "rows": [
                    [_display_date(row["date"]), row["activity"], row["reservations"], row["check_ins"], row["check_outs"], row["housekeeping"], row["payments"], row["staff"]]
                    for row in busiest_rows[:15]
                ],
                "export_rows": [
                    [row["date"], row["activity"], row["reservations"], row["check_ins"], row["check_outs"], row["housekeeping"], row["payments"], row["staff"]]
                    for row in busiest_rows
                ],
                "summary_row": ["TOTALS", sum((row["activity"] for row in points), 0), "", "", "", "", "", ""],
                "export_summary_row": ["TOTALS", sum((row["activity"] for row in points), 0), "", "", "", "", "", ""],
            }
        ],
    }


def _build_analytics_sections(filters):
    return [
        _build_rooms_analytics_section(filters),
        _build_bookings_analytics_section(filters),
        _build_revenue_analytics_section(filters),
        _build_staff_analytics_section(filters),
        _build_housekeeping_analytics_section(filters),
        _build_roster_analytics_section(filters),
        _build_operations_analytics_section(filters),
    ]


def _build_analytics_summary_metrics(sections):
    section_map = {section["key"]: section for section in sections}
    return [
        {"label": "Revenue", "value": section_map["revenue-analytics"]["metrics"][0]["value"]},
        {"label": "Bookings", "value": section_map["bookings-analytics"]["metrics"][0]["value"]},
        {"label": "Avg Occupancy", "value": section_map["rooms-analytics"]["metrics"][1]["value"]},
        {"label": "Staff", "value": section_map["staff-analytics"]["metrics"][0]["value"]},
        {"label": "Low Stock", "value": section_map["housekeeping-analytics"]["metrics"][3]["value"]},
    ]


def _analytics_pdf_response(sections, filters, filename):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return redirect("analytics-center")

    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("System Analytics", styles["Title"]),
        Paragraph(filters["report_window"]["display_range"], styles["Normal"]),
        Spacer(1, 12),
    ]

    for index, section in enumerate(sections):
        story.append(Paragraph(section["title"], styles["Heading2"]))
        story.append(Paragraph(section["subtitle"], styles["BodyText"]))
        story.append(Spacer(1, 8))

        metric_table = Table(
            [["Metric", "Value"]] + [[metric["label"], str(metric["value"])] for metric in section["metrics"]],
            repeatRows=1,
        )
        metric_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#23444B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2E5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(metric_table)
        story.append(Spacer(1, 10))

        for table in section["tables"]:
            story.append(Paragraph(table["title"], styles["Heading3"]))
            table_rows = [table["headers"]]
            table_rows.extend([[str(cell) for cell in row] for row in table["rows"][:20]])
            if table.get("summary_row"):
                table_rows.append([str(cell) for cell in table["summary_row"]])
            pdf_table = Table(table_rows, repeatRows=1)
            pdf_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#23444B")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2E5")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(pdf_table)
            story.append(Spacer(1, 12))
        if index < len(sections) - 1:
            story.append(PageBreak())

    document.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def _build_bookings_report_section(start_date, end_date, total_rooms):
    money_field = DecimalField(max_digits=12, decimal_places=2)
    reservations_by_day = {
        row["day"]: row["count"]
        for row in Booking.objects.filter(created_at__date__range=[start_date, end_date])
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(count=Count("id"))
        .order_by("day")
    }
    check_ins_by_day = {
        row["check_in"]: row["count"]
        for row in Booking.objects.filter(check_in__range=[start_date, end_date])
        .values("check_in")
        .annotate(count=Count("id"))
        .order_by("check_in")
    }
    check_outs_by_day = {
        row["check_out"]: row["count"]
        for row in Booking.objects.filter(check_out__range=[start_date, end_date])
        .values("check_out")
        .annotate(count=Count("id"))
        .order_by("check_out")
    }
    cancellations_by_day = {
        row["day"]: row["count"]
        for row in Booking.objects.filter(
            status=Booking.BookingStatus.CANCELLED,
            updated_at__date__range=[start_date, end_date],
        )
        .annotate(day=TruncDate("updated_at"))
        .values("day")
        .annotate(count=Count("id"))
        .order_by("day")
    }

    occupancy_lookup = {
        date.fromisoformat(row["date"]): row for row in _daily_report_rows(start_date, end_date, total_rooms)
    }

    display_rows = []
    export_rows = []
    total_occupied_room_days = 0
    total_occupancy_percent = 0
    for current_day in _report_days(start_date, end_date):
        occupied_rooms = occupancy_lookup.get(current_day, {}).get("occupied_rooms", 0)
        occupancy_percent = occupancy_lookup.get(current_day, {}).get("occupancy_percent", 0)
        reservations = reservations_by_day.get(current_day, 0)
        check_ins = check_ins_by_day.get(current_day, 0)
        check_outs = check_outs_by_day.get(current_day, 0)
        cancellations = cancellations_by_day.get(current_day, 0)
        total_occupied_room_days += occupied_rooms
        total_occupancy_percent += occupancy_percent

        display_rows.append(
            [
                _display_date(current_day),
                reservations,
                check_ins,
                check_outs,
                cancellations,
                occupied_rooms,
                _display_percent(occupancy_percent),
            ]
        )
        export_rows.append(
            [
                current_day,
                reservations,
                check_ins,
                check_outs,
                cancellations,
                occupied_rooms,
                float(occupancy_percent),
            ]
        )

    report_days = _report_days(start_date, end_date)
    average_occupancy = round(total_occupancy_percent / len(report_days), 2) if report_days else 0
    reservations_total = Booking.objects.filter(created_at__date__range=[start_date, end_date]).count()
    check_ins_total = Booking.objects.filter(check_in__range=[start_date, end_date]).count()
    check_outs_total = Booking.objects.filter(check_out__range=[start_date, end_date]).count()
    cancellations_total = Booking.objects.filter(
        status=Booking.BookingStatus.CANCELLED,
        updated_at__date__range=[start_date, end_date],
    ).count()
    event_bookings_total = EventBooking.objects.filter(event_start__date__range=[start_date, end_date]).count()
    booking_revenue = _money_total(
        booking_revenue_queryset(start_date, end_date),
        "total_amount",
    )

    return {
        "key": "bookings",
        "title": "Bookings",
        "sheet_title": "Bookings",
        "filename_prefix": "bookings",
        "subtitle": "Reservations, arrivals, departures, cancellations, and occupancy for the selected period.",
        "summary": {
            "reservations": reservations_total,
            "check_ins": check_ins_total,
            "check_outs": check_outs_total,
            "cancellations": cancellations_total,
            "average_occupancy": average_occupancy,
            "event_bookings": event_bookings_total,
            "booking_revenue": booking_revenue,
        },
        "metrics": [
            {"label": "Reservations", "value": reservations_total, "export_value": reservations_total},
            {"label": "Check-ins", "value": check_ins_total, "export_value": check_ins_total},
            {"label": "Check-outs", "value": check_outs_total, "export_value": check_outs_total},
            {"label": "Cancellations", "value": cancellations_total, "export_value": cancellations_total},
            {
                "label": "Average occupancy",
                "value": _display_percent(average_occupancy),
                "export_value": float(average_occupancy),
            },
            {"label": "Event bookings", "value": event_bookings_total, "export_value": event_bookings_total},
        ],
        "tables": [
            {
                "title": "Daily booking activity",
                "headers": [
                    "Date",
                    "Reservations",
                    "Check-ins",
                    "Check-outs",
                    "Cancellations",
                    "Occupied Rooms",
                    "Occupancy %",
                ],
                "rows": display_rows,
                "export_rows": export_rows,
                "summary_row": [
                    "TOTALS",
                    reservations_total,
                    check_ins_total,
                    check_outs_total,
                    cancellations_total,
                    total_occupied_room_days,
                    _display_percent(average_occupancy),
                ],
                "export_summary_row": [
                    "TOTALS",
                    reservations_total,
                    check_ins_total,
                    check_outs_total,
                    cancellations_total,
                    total_occupied_room_days,
                    float(average_occupancy),
                ],
            }
        ],
    }


def _build_revenue_report_section(start_date, end_date):
    money_field = DecimalField(max_digits=12, decimal_places=2)
    revenue_breakdown = revenue_components(start_date, end_date)
    booking_payments = booking_payment_queryset(start_date, end_date)
    event_payments = event_payment_queryset(start_date, end_date)
    pos_sales = completed_pos_sales_queryset(start_date, end_date)

    booking_payments_total = _money_total(booking_payments, "amount")
    event_payments_total = _money_total(event_payments, "amount")
    booking_revenue_total = revenue_breakdown["booking_revenue"]
    event_revenue_total = revenue_breakdown["event_revenue"]
    pos_sales_total = revenue_breakdown["pos_sales"]
    total_revenue = revenue_breakdown["total_revenue"]
    payments_received = revenue_breakdown["payments_received"]

    bookings_with_balance = booking_revenue_queryset(start_date, end_date).annotate(
        paid_total=Coalesce(Sum("payments__amount"), Value(0, output_field=money_field)),
        balance=ExpressionWrapper(F("total_amount") - F("paid_total"), output_field=money_field),
    )
    event_bookings_with_balance = event_revenue_queryset(start_date, end_date).annotate(
        paid_total=Coalesce(Sum("payments__amount"), Value(0, output_field=money_field)),
        balance=ExpressionWrapper(F("total_amount") - F("paid_total"), output_field=money_field),
    )

    outstanding_total = bookings_with_balance.aggregate(
        total=Coalesce(Sum("balance", filter=Q(balance__gt=0)), Value(0, output_field=money_field))
    )["total"] + event_bookings_with_balance.aggregate(
        total=Coalesce(Sum("balance", filter=Q(balance__gt=0)), Value(0, output_field=money_field))
    )["total"]

    payment_labels = {
        **dict(Payment.PaymentMethod.choices),
        **dict(EventPayment.PaymentMethod.choices),
        **dict(Sale.PaymentMethod.choices),
    }
    payment_breakdown = defaultdict(
        lambda: {"booking": Decimal("0.00"), "event": Decimal("0.00"), "pos": Decimal("0.00")}
    )

    for row in booking_payments.values("method").annotate(total=Coalesce(Sum("amount"), Value(0, output_field=money_field))):
        payment_breakdown[payment_labels.get(row["method"], row["method"] or "Other")]["booking"] += Decimal(str(row["total"] or 0))
    for row in event_payments.values("method").annotate(total=Coalesce(Sum("amount"), Value(0, output_field=money_field))):
        payment_breakdown[payment_labels.get(row["method"], row["method"] or "Other")]["event"] += Decimal(str(row["total"] or 0))
    for row in pos_sales.values("payment_method").annotate(total=Coalesce(Sum("grand_total"), Value(0, output_field=money_field))):
        payment_breakdown[payment_labels.get(row["payment_method"], row["payment_method"] or "Other")]["pos"] += Decimal(str(row["total"] or 0))

    payment_rows = []
    payment_export_rows = []
    for method_name in sorted(payment_breakdown):
        booking_total = payment_breakdown[method_name]["booking"]
        event_total = payment_breakdown[method_name]["event"]
        pos_total = payment_breakdown[method_name]["pos"]
        grand_total = booking_total + event_total + pos_total
        payment_rows.append(
            [
                method_name,
                _display_money(booking_total),
                _display_money(event_total),
                _display_money(pos_total),
                _display_money(grand_total),
            ]
        )
        payment_export_rows.append([method_name, float(booking_total), float(event_total), float(pos_total), float(grand_total)])

    outstanding_rows = []
    outstanding_export_rows = []
    outstanding_items = []
    for booking in bookings_with_balance.filter(balance__gt=0).select_related("guest", "room"):
        outstanding_items.append(
            {
                "source": "Reservation",
                "reference": f"Booking #{booking.pk}",
                "subject": str(booking.guest),
                "status": booking.get_status_display(),
                "total_amount": booking.total_amount,
                "paid_total": booking.paid_total,
                "balance": booking.balance,
            }
        )
    for event_booking in event_bookings_with_balance.filter(balance__gt=0).select_related("guest"):
        outstanding_items.append(
            {
                "source": "Event",
                "reference": f"Event #{event_booking.pk}",
                "subject": event_booking.event_title,
                "status": event_booking.get_status_display(),
                "total_amount": event_booking.total_amount,
                "paid_total": event_booking.paid_total,
                "balance": event_booking.balance,
            }
        )
    outstanding_items.sort(key=lambda row: row["balance"], reverse=True)
    for row in outstanding_items[:20]:
        outstanding_rows.append(
            [
                row["source"],
                row["reference"],
                row["subject"],
                row["status"],
                _display_money(row["total_amount"]),
                _display_money(row["paid_total"]),
                _display_money(row["balance"]),
            ]
        )
        outstanding_export_rows.append(
            [
                row["source"],
                row["reference"],
                row["subject"],
                row["status"],
                float(row["total_amount"] or 0),
                float(row["paid_total"] or 0),
                float(row["balance"] or 0),
            ]
        )

    return {
        "key": "revenue-payments",
        "title": "Revenue & Payments",
        "sheet_title": "Revenue Payments",
        "filename_prefix": "revenue-payments",
        "subtitle": "Booking revenue, event revenue, POS sales, payment collections, and balances still outstanding across the selected period.",
        "summary": {
            "total_revenue": total_revenue,
            "payments_received": payments_received,
            "pos_sales_total": pos_sales_total,
            "outstanding_total": outstanding_total,
        },
        "metrics": [
            {"label": "Total revenue", "value": _display_money(total_revenue), "export_value": float(total_revenue)},
            {"label": "Payments received", "value": _display_money(payments_received), "export_value": float(payments_received)},
            {"label": "POS sales", "value": _display_money(pos_sales_total), "export_value": float(pos_sales_total)},
            {"label": "Outstanding balances", "value": _display_money(outstanding_total), "export_value": float(outstanding_total)},
        ],
        "tables": [
            {
                "title": "Collections by payment method",
                "headers": ["Method", "Reservation Payments", "Event Payments", "POS Sales", "Total"],
                "rows": payment_rows,
                "export_rows": payment_export_rows,
                "summary_row": [
                    "TOTALS",
                    _display_money(booking_payments_total),
                    _display_money(event_payments_total),
                    _display_money(pos_sales_total),
                    _display_money(total_revenue),
                ],
                "export_summary_row": ["TOTALS", float(booking_payments_total), float(event_payments_total), float(pos_sales_total), float(total_revenue)],
            },
            {
                "title": "Revenue overview",
                "headers": ["Metric", "Value"],
                "rows": [
                    ["Booking revenue", _display_money(booking_revenue_total)],
                    ["Event revenue", _display_money(event_revenue_total)],
                    ["POS sales", _display_money(pos_sales_total)],
                    ["Payments received", _display_money(payments_received)],
                    ["Outstanding balances", _display_money(outstanding_total)],
                ],
                "export_rows": [
                    ["Booking revenue", float(booking_revenue_total)],
                    ["Event revenue", float(event_revenue_total)],
                    ["POS sales", float(pos_sales_total)],
                    ["Payments received", float(payments_received)],
                    ["Outstanding balances", float(outstanding_total)],
                ],
                "summary_row": ["TOTALS", _display_money(total_revenue)],
                "export_summary_row": ["TOTALS", float(total_revenue)],
            },
            {
                "title": "Outstanding balances",
                "headers": ["Source", "Reference", "Guest / Event", "Status", "Total Amount", "Paid", "Balance"],
                "rows": outstanding_rows,
                "export_rows": outstanding_export_rows,
                "summary_row": ["TOTALS", "", "", "", "", "", _display_money(outstanding_total)],
                "export_summary_row": ["TOTALS", "", "", "", "", "", float(outstanding_total)],
            },
        ],
    }


def _build_housekeeping_report_section(start_date, end_date):
    quantity_field = DecimalField(max_digits=12, decimal_places=3)
    money_field = DecimalField(max_digits=12, decimal_places=2)
    logs = HousekeepingItemLog.objects.select_related("room", "item").filter(used_at__date__range=[start_date, end_date])
    current_items = list(HousekeepingItem.objects.order_by("name"))
    low_stock_items = [item for item in current_items if item.is_low_stock]

    usage_item_rows = []
    usage_item_export_rows = []
    for row in (
        logs.values("item_name", "unit")
        .annotate(
            total_used=Coalesce(Sum("quantity_used"), Value(0, output_field=quantity_field)),
            entries=Count("id"),
            current_stock=Coalesce(Max("item__quantity_in_stock"), Value(0, output_field=quantity_field)),
            initial_quantity=Coalesce(Max("item__initial_quantity"), Value(0, output_field=quantity_field)),
        )
        .order_by("item_name")
    ):
        usage_item_rows.append(
            [
                row["item_name"],
                _display_quantity(row["initial_quantity"]),
                _display_quantity(row["total_used"]),
                _display_quantity(row["current_stock"]),
                row["unit"],
                row["entries"],
            ]
        )
        usage_item_export_rows.append([row["item_name"], float(row["initial_quantity"] or 0), float(row["total_used"] or 0), float(row["current_stock"] or 0), row["unit"], row["entries"]])

    usage_room_rows = []
    usage_room_export_rows = []
    for row in (
        logs.filter(room__isnull=False)
        .values("room__room_number")
        .annotate(total_used=Coalesce(Sum("quantity_used"), Value(0, output_field=quantity_field)), entries=Count("id"))
        .order_by("room__room_number")
    ):
        usage_room_rows.append([row["room__room_number"], _display_quantity(row["total_used"]), row["entries"]])
        usage_room_export_rows.append([row["room__room_number"], float(row["total_used"] or 0), row["entries"]])

    usage_day_rows = []
    usage_day_export_rows = []
    for row in (
        logs.annotate(day=TruncDate("used_at"))
        .values("day")
        .annotate(total_used=Coalesce(Sum("quantity_used"), Value(0, output_field=quantity_field)), entries=Count("id"))
        .order_by("day")
    ):
        usage_day_rows.append([_display_date(row["day"]), _display_quantity(row["total_used"]), row["entries"]])
        usage_day_export_rows.append([row["day"], float(row["total_used"] or 0), row["entries"]])

    low_stock_rows = []
    low_stock_export_rows = []
    for item in low_stock_items:
        low_stock_rows.append([item.name, _display_quantity(item.initial_quantity), _display_quantity(item.quantity_in_stock), item.unit, _display_quantity(item.effective_low_stock_threshold)])
        low_stock_export_rows.append([item.name, float(item.initial_quantity or 0), float(item.quantity_in_stock or 0), item.unit, float(item.effective_low_stock_threshold or 0)])

    pos_sales = completed_pos_sales_queryset(start_date, end_date)
    pos_rows = []
    pos_export_rows = []
    for row in (
        pos_sales.values("payment_method")
        .annotate(sales_total=Coalesce(Sum("grand_total"), Value(0, output_field=money_field)), transactions=Count("id"))
        .order_by("payment_method")
    ):
        label = dict(Sale.PaymentMethod.choices).get(row["payment_method"], row["payment_method"] or "Other")
        pos_rows.append([label, row["transactions"], _display_money(row["sales_total"])])
        pos_export_rows.append([label, row["transactions"], float(row["sales_total"] or 0)])

    total_items_used = _quantity_total(logs, "quantity_used")
    total_stock = sum((Decimal(str(item.quantity_in_stock or 0)) for item in current_items), Decimal("0"))
    total_initial_stock = sum((Decimal(str(item.initial_quantity or 0)) for item in current_items), Decimal("0"))
    pos_sales_total = _money_total(pos_sales, "grand_total")

    return {
        "key": "housekeeping",
        "title": "Housekeeping",
        "sheet_title": "Housekeeping",
        "filename_prefix": "housekeeping",
        "subtitle": "Hotel item usage, live stock position, low-stock alerts, room-by-room consumption, and POS sales for the selected period.",
        "summary": {
            "items_used": total_items_used,
            "stock_total": total_stock,
            "initial_stock_total": total_initial_stock,
            "low_stock_count": len(low_stock_items),
            "usage_entries": logs.count(),
            "pos_sales_total": pos_sales_total,
        },
        "metrics": [
            {"label": "Items used", "value": _display_quantity(total_items_used), "export_value": float(total_items_used)},
            {"label": "Current stock", "value": _display_quantity(total_stock), "export_value": float(total_stock)},
            {"label": "Initial stock", "value": _display_quantity(total_initial_stock), "export_value": float(total_initial_stock)},
            {"label": "Low stock alerts", "value": len(low_stock_items), "export_value": len(low_stock_items)},
            {"label": "POS sales", "value": _display_money(pos_sales_total), "export_value": float(pos_sales_total)},
        ],
        "tables": [
            {
                "title": "Usage by item",
                "headers": ["Item", "Initial Qty", "Used", "Qty in Stock", "Unit", "Entries"],
                "rows": usage_item_rows,
                "export_rows": usage_item_export_rows,
                "summary_row": ["TOTALS", _display_quantity(total_initial_stock), _display_quantity(total_items_used), _display_quantity(total_stock), "", logs.count()],
                "export_summary_row": ["TOTALS", float(total_initial_stock), float(total_items_used), float(total_stock), "", logs.count()],
            },
            {
                "title": "Usage by room",
                "headers": ["Room", "Total Used", "Entries"],
                "rows": usage_room_rows,
                "export_rows": usage_room_export_rows,
                "summary_row": ["TOTALS", _display_quantity(total_items_used), logs.count()],
                "export_summary_row": ["TOTALS", float(total_items_used), logs.count()],
            },
            {
                "title": "Usage by day",
                "headers": ["Date", "Total Used", "Entries"],
                "rows": usage_day_rows,
                "export_rows": usage_day_export_rows,
                "summary_row": ["TOTALS", _display_quantity(total_items_used), logs.count()],
                "export_summary_row": ["TOTALS", float(total_items_used), logs.count()],
            },
            {
                "title": "Low stock alerts",
                "headers": ["Item", "Initial Qty", "Qty in Stock", "Unit", "Alert Threshold"],
                "rows": low_stock_rows,
                "export_rows": low_stock_export_rows,
                "summary_row": ["TOTALS", "", "", "", len(low_stock_items)],
                "export_summary_row": ["TOTALS", "", "", "", len(low_stock_items)],
            },
            {
                "title": "POS sales by payment method",
                "headers": ["Method", "Transactions", "Sales Total"],
                "rows": pos_rows,
                "export_rows": pos_export_rows,
                "summary_row": ["TOTALS", pos_sales.count(), _display_money(pos_sales_total)],
                "export_summary_row": ["TOTALS", pos_sales.count(), float(pos_sales_total)],
            },
        ],
    }


def _build_roster_report_section(start_date, end_date):
    rotas = (
        Rota.objects.select_related("employee")
        .filter(employee__isnull=False, period_end__gte=start_date, period_start__lte=end_date)
        .order_by("employee__last_name", "employee__first_name", "period_start")
    )

    employee_rollup = {}
    department_rollup = {}
    total_hours = Decimal("0")
    for rota in rotas:
        if not rota.employee_id or not rota.period_start or not rota.period_end:
            continue
        overlap_start = max(rota.period_start, start_date)
        overlap_end = min(rota.period_end, end_date)
        if overlap_start > overlap_end:
            continue
        assigned_days = (overlap_end - overlap_start).days + 1
        assigned_hours = Decimal(str(rota.daily_hours)) * Decimal(str(assigned_days))
        total_hours += assigned_hours

        employee_key = rota.employee_id
        if employee_key not in employee_rollup:
            employee_rollup[employee_key] = {
                "employee": rota.employee.full_name,
                "department": rota.employee.department or "-",
                "role": rota.employee.job_title or rota.employee.get_position_display(),
                "rota_entries": 0,
                "assigned_days": 0,
                "assigned_hours": Decimal("0"),
            }
        employee_rollup[employee_key]["rota_entries"] += 1
        employee_rollup[employee_key]["assigned_days"] += assigned_days
        employee_rollup[employee_key]["assigned_hours"] += assigned_hours

        department_key = rota.employee.department or "Unassigned"
        if department_key not in department_rollup:
            department_rollup[department_key] = {
                "department": department_key,
                "employees": set(),
                "rota_entries": 0,
                "assigned_days": 0,
                "assigned_hours": Decimal("0"),
            }
        department_rollup[department_key]["employees"].add(rota.employee_id)
        department_rollup[department_key]["rota_entries"] += 1
        department_rollup[department_key]["assigned_days"] += assigned_days
        department_rollup[department_key]["assigned_hours"] += assigned_hours

    employee_rows = []
    employee_export_rows = []
    for row in sorted(employee_rollup.values(), key=lambda item: (item["department"], item["employee"])):
        employee_rows.append([row["employee"], row["department"], row["role"], row["rota_entries"], row["assigned_days"], _display_quantity(row["assigned_hours"])])
        employee_export_rows.append([row["employee"], row["department"], row["role"], row["rota_entries"], row["assigned_days"], float(row["assigned_hours"])])

    department_rows = []
    department_export_rows = []
    for row in sorted(department_rollup.values(), key=lambda item: item["department"]):
        department_rows.append([row["department"], len(row["employees"]), row["rota_entries"], row["assigned_days"], _display_quantity(row["assigned_hours"])])
        department_export_rows.append([row["department"], len(row["employees"]), row["rota_entries"], row["assigned_days"], float(row["assigned_hours"])])

    return {
        "key": "duty-roster",
        "title": "Duty Roster",
        "sheet_title": "Duty Roster",
        "filename_prefix": "duty-roster",
        "subtitle": "Employee shift coverage and department staffing pulled from the same weekly rosters used in staff management.",
        "summary": {
            "rota_entries": rotas.count(),
            "employees_scheduled": len(employee_rollup),
            "departments_covered": len(department_rollup),
            "assigned_hours": total_hours,
        },
        "metrics": [
            {"label": "Rota entries", "value": rotas.count(), "export_value": rotas.count()},
            {"label": "Employees scheduled", "value": len(employee_rollup), "export_value": len(employee_rollup)},
            {"label": "Departments covered", "value": len(department_rollup), "export_value": len(department_rollup)},
            {"label": "Assigned hours", "value": _display_quantity(total_hours), "export_value": float(total_hours)},
        ],
        "tables": [
            {
                "title": "Shifts assigned per employee",
                "headers": ["Employee", "Department", "Role", "Rota Entries", "Assigned Days", "Assigned Hours"],
                "rows": employee_rows,
                "export_rows": employee_export_rows,
                "summary_row": ["TOTALS", "", "", rotas.count(), sum((row["assigned_days"] for row in employee_rollup.values()), 0), _display_quantity(total_hours)],
                "export_summary_row": ["TOTALS", "", "", rotas.count(), sum((row["assigned_days"] for row in employee_rollup.values()), 0), float(total_hours)],
            },
            {
                "title": "Department coverage",
                "headers": ["Department", "Employees", "Rota Entries", "Assigned Days", "Assigned Hours"],
                "rows": department_rows,
                "export_rows": department_export_rows,
                "summary_row": ["TOTALS", len(employee_rollup), rotas.count(), "", _display_quantity(total_hours)],
                "export_summary_row": ["TOTALS", len(employee_rollup), rotas.count(), "", float(total_hours)],
            },
        ],
    }


def _build_rooms_report_section(start_date, end_date, total_rooms):
    bookings = list(
        Booking.objects.select_related("room")
        .filter(check_in__lte=end_date, check_out__gte=start_date)
        .order_by("room__room_number", "check_in")
    )
    daily_rows = _daily_report_rows(start_date, end_date, total_rooms)
    occupied_room_days = sum((row["occupied_rooms"] for row in daily_rows), 0)
    average_occupancy = round(sum((row["occupancy_percent"] for row in daily_rows), 0) / len(daily_rows), 2) if daily_rows else 0
    maintenance_requests = MaintenanceRequest.objects.filter(created_at__date__range=[start_date, end_date])
    cleaned_rooms = HousekeepingItemLog.objects.filter(used_at__date__range=[start_date, end_date], room__isnull=False).values("room_id").distinct().count()

    room_booking_rollup = {}
    for booking in bookings:
        booked_nights = booking_occupied_days_in_range(
            booking.check_in,
            booking.check_out,
            start_date,
            end_date,
        )
        room_key = booking.room_id
        if room_key not in room_booking_rollup:
            room_booking_rollup[room_key] = {
                "room_number": booking.room.room_number,
                "room_type": booking.room.get_room_type_display(),
                "bookings": 0,
                "booked_nights": 0,
            }
        room_booking_rollup[room_key]["bookings"] += 1
        room_booking_rollup[room_key]["booked_nights"] += booked_nights

    ranked_rooms = sorted(
        room_booking_rollup.values(),
        key=lambda row: (row["bookings"], row["booked_nights"], row["room_number"]),
        reverse=True,
    )
    most_booked_rows = [[row["room_number"], row["room_type"], row["bookings"], row["booked_nights"]] for row in ranked_rooms[:10]]
    most_booked_export_rows = [row[:] for row in most_booked_rows]

    daily_display_rows = []
    daily_export_rows = []
    for row in daily_rows:
        daily_display_rows.append([
            _display_date(date.fromisoformat(row["date"])),
            row["occupied_rooms"],
            _display_percent(row["occupancy_percent"]),
            _display_money(row["revenue_collected"]),
        ])
        daily_export_rows.append([date.fromisoformat(row["date"]), row["occupied_rooms"], float(row["occupancy_percent"]), float(row["revenue_collected"] or 0)])

    status_rows = [
        ["Total rooms", total_rooms],
        ["Occupied room-days", occupied_room_days],
        ["Available room-days", max((total_rooms * len(daily_rows)) - occupied_room_days, 0)],
        ["Maintenance requests", maintenance_requests.count()],
        ["Rooms with housekeeping activity", cleaned_rooms],
    ]

    return {
        "key": "rooms",
        "title": "Rooms",
        "sheet_title": "Rooms",
        "filename_prefix": "rooms",
        "subtitle": "Room occupancy, maintenance activity, and the most booked rooms across the selected range.",
        "summary": {
            "average_occupancy": average_occupancy,
            "occupied_room_days": occupied_room_days,
            "maintenance_requests": maintenance_requests.count(),
            "rooms_cleaned": cleaned_rooms,
        },
        "metrics": [
            {"label": "Average occupancy", "value": _display_percent(average_occupancy), "export_value": float(average_occupancy)},
            {"label": "Occupied room-days", "value": occupied_room_days, "export_value": occupied_room_days},
            {"label": "Maintenance requests", "value": maintenance_requests.count(), "export_value": maintenance_requests.count()},
            {"label": "Rooms cleaned", "value": cleaned_rooms, "export_value": cleaned_rooms},
        ],
        "tables": [
            {
                "title": "Room status overview",
                "headers": ["Metric", "Value"],
                "rows": status_rows,
                "export_rows": status_rows,
                "summary_row": ["TOTALS", ""],
                "export_summary_row": ["TOTALS", ""],
            },
            {
                "title": "Daily occupancy",
                "headers": ["Date", "Occupied Rooms", "Occupancy %", "Revenue"],
                "rows": daily_display_rows,
                "export_rows": daily_export_rows,
                "summary_row": ["AVERAGE / TOTAL", occupied_room_days, _display_percent(average_occupancy), _display_money(sum((row["revenue_collected"] for row in daily_rows), 0))],
                "export_summary_row": ["AVERAGE / TOTAL", occupied_room_days, float(average_occupancy), float(sum((row["revenue_collected"] for row in daily_rows), 0))],
            },
            {
                "title": "Most booked rooms",
                "headers": ["Room", "Type", "Bookings", "Booked Nights"],
                "rows": most_booked_rows,
                "export_rows": most_booked_export_rows,
                "summary_row": ["TOTALS", "", len(room_booking_rollup), sum((row["booked_nights"] for row in ranked_rooms), 0)],
                "export_summary_row": ["TOTALS", "", len(room_booking_rollup), sum((row["booked_nights"] for row in ranked_rooms), 0)],
            },
        ],
    }


def _build_staff_report_section(start_date, end_date):
    active_employees = Employee.objects.filter(employment_status="active").count()
    terminated_employees = Employee.objects.filter(employment_status="terminated").count()
    on_leave_employees = Employee.objects.filter(employment_status__in=list(LEAVE_TYPE_TO_EMPLOYMENT_STATUS.values())).count()

    leave_requests = LeaveRequest.objects.filter(start_date__lte=end_date, end_date__gte=start_date)
    attendance_records = AttendanceRecord.objects.filter(work_date__range=[start_date, end_date])
    payroll_records = PayrollRecord.objects.filter(pay_period_end__gte=start_date, pay_period_start__lte=end_date)
    training_records = TrainingRecord.objects.filter(
        Q(start_date__range=[start_date, end_date])
        | Q(completion_date__range=[start_date, end_date])
        | Q(created_at__date__range=[start_date, end_date])
    )

    leave_totals = leave_requests.aggregate(total_days=Coalesce(Sum("days"), Value(0)))
    leave_rows = []
    leave_export_rows = []
    for row in (
        leave_requests.values("leave_type", "approval_status")
        .annotate(total_requests=Count("id"), total_days=Coalesce(Sum("days"), Value(0)))
        .order_by("leave_type", "approval_status")
    ):
        leave_rows.append([
            dict(LeaveRequest.LeaveType.choices).get(row["leave_type"], row["leave_type"]),
            dict(LeaveRequest.ApprovalStatus.choices).get(row["approval_status"], row["approval_status"]),
            row["total_requests"],
            row["total_days"],
        ])
        leave_export_rows.append(leave_rows[-1][:])

    attendance_rows = []
    attendance_export_rows = []
    for row in attendance_records.values("status").annotate(total_entries=Count("id")).order_by("status"):
        label = dict(AttendanceRecord.AttendanceStatus.choices).get(row["status"], row["status"])
        attendance_rows.append([label, row["total_entries"]])
        attendance_export_rows.append([label, row["total_entries"]])

    payroll_total = _money_total(payroll_records, "net_pay")

    return {
        "key": "staff-hr",
        "title": "Staff & HR",
        "sheet_title": "Staff HR",
        "filename_prefix": "staff-hr",
        "subtitle": "Headcount, leave activity, attendance, payroll, and training records for the selected period.",
        "summary": {
            "active_employees": active_employees,
            "terminated_employees": terminated_employees,
            "on_leave_employees": on_leave_employees,
            "leave_requests": leave_requests.count(),
            "attendance_entries": attendance_records.count(),
            "payroll_total": payroll_total,
            "training_records": training_records.count(),
        },
        "metrics": [
            {"label": "Active employees", "value": active_employees, "export_value": active_employees},
            {"label": "On leave", "value": on_leave_employees, "export_value": on_leave_employees},
            {"label": "Terminated", "value": terminated_employees, "export_value": terminated_employees},
            {"label": "Payroll total", "value": _display_money(payroll_total), "export_value": float(payroll_total)},
            {"label": "Training records", "value": training_records.count(), "export_value": training_records.count()},
        ],
        "tables": [
            {
                "title": "Leave summary",
                "headers": ["Leave Type", "Approval Status", "Requests", "Days"],
                "rows": leave_rows,
                "export_rows": leave_export_rows,
                "summary_row": ["TOTALS", "", leave_requests.count(), leave_totals["total_days"]],
                "export_summary_row": ["TOTALS", "", leave_requests.count(), leave_totals["total_days"]],
            },
            {
                "title": "Attendance summary",
                "headers": ["Status", "Entries"],
                "rows": attendance_rows,
                "export_rows": attendance_export_rows,
                "summary_row": ["TOTALS", attendance_records.count()],
                "export_summary_row": ["TOTALS", attendance_records.count()],
            },
        ],
    }


def _create_reports_workbook_or_none(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        messages.error(request, "openpyxl is required for report exports.")
        return None
    return Workbook()


def _write_report_section_sheet(worksheet, section, display_range):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    subheader_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet.append([section["title"]])
    worksheet["A1"].font = title_font
    worksheet.append([section["subtitle"]])
    worksheet.append([f"Range: {display_range}"])
    worksheet.append([])
    worksheet.append(["Summary", "Value"])
    for cell in worksheet[5]:
        cell.font = header_font
        cell.fill = header_fill
    for metric in section["metrics"]:
        worksheet.append([metric["label"], metric["export_value"]])

    for table in section["tables"]:
        worksheet.append([])
        worksheet.append([table["title"]])
        worksheet.cell(row=worksheet.max_row, column=1).font = subheader_font
        worksheet.append(table["headers"])
        for cell in worksheet[worksheet.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in table.get("export_rows", table["rows"]):
            worksheet.append(row)
        if table.get("export_summary_row"):
            worksheet.append(table["export_summary_row"])
        for cell in worksheet[worksheet.max_row]:
            cell.font = Font(bold=True)
            cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)


def _write_report_overview_sheet(worksheet, sections, display_range):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")

    worksheet["A1"] = "Full System Report"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {display_range}"
    worksheet["A4"] = "This workbook contains one sheet per report section."
    worksheet["A6"] = "Sheet"
    worksheet["B6"] = "Report Section"
    worksheet["C6"] = "What it includes"

    for cell in worksheet[6]:
        cell.font = header_font
        cell.fill = header_fill

    for index, section in enumerate(sections, start=7):
        worksheet.cell(row=index, column=1).value = section["sheet_title"]
        worksheet.cell(row=index, column=2).value = section["title"]
        worksheet.cell(row=index, column=3).value = section["subtitle"]

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 90
    for column in worksheet.columns:
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_owner_withdrawals_log_sheet(worksheet, withdrawals, start_date, end_date):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Sales Deposits Collections Log"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Date", "Amount Collected", "Method", "Recorded By", "Collected By"])

    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_amount = Decimal("0.00")
    for withdrawal in withdrawals:
        total_amount += Decimal(str(withdrawal.amount or 0))
        worksheet.append(
            [
                timezone.localtime(withdrawal.created_at).strftime("%d/%m/%Y %H:%M"),
                float(withdrawal.amount or 0),
                withdrawal.get_collection_method_display(),
                withdrawal.recorded_by_name,
                withdrawal.collected_by,
            ]
        )

    worksheet.append(["TOTALS", float(total_amount), f"{len(withdrawals)} entries", "", ""])
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 30)


def _write_owner_withdrawals_summary_sheet(worksheet, start_date, end_date):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Financial Summary"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Period", "Gross Revenue", "Total Withdrawals", "Net Revenue"])

    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    gross_map = daily_total_revenue_map(start_date, end_date)
    withdrawals_map = daily_owner_withdrawals_map(start_date, end_date)
    net_map = daily_net_revenue_map(start_date, end_date)
    gross_total = Decimal("0.00")
    withdrawals_total = Decimal("0.00")
    net_total = Decimal("0.00")

    for current_day in _report_days(start_date, end_date):
        gross_value = Decimal(str(gross_map.get(current_day, Decimal("0")) or 0))
        withdrawal_value = Decimal(str(withdrawals_map.get(current_day, Decimal("0")) or 0))
        net_value = Decimal(str(net_map.get(current_day, Decimal("0")) or 0))
        gross_total += gross_value
        withdrawals_total += withdrawal_value
        net_total += net_value
        worksheet.append(
            [
                current_day.strftime("%d/%m/%Y"),
                float(gross_value),
                float(withdrawal_value),
                float(net_value),
            ]
        )

    worksheet.append(["TOTALS", float(gross_total), float(withdrawals_total), float(net_total)])
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 28)


def _write_finance_revenue_sheet(worksheet, revenue_breakdown, daily_rows, start_date, end_date):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Revenue Breakdown"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Revenue Source", "Total"])
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    gross_total = Decimal("0.00")
    for row in revenue_breakdown:
        gross_total += Decimal(str(row["amount"] or 0))
        worksheet.append([row["label"], float(row["amount"] or 0)])
    worksheet.append(["TOTAL REVENUE", float(gross_total)])
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    worksheet.append([])
    worksheet.append(["Date", "Room Revenue", "POS Revenue", "Event Revenue", "Other Revenue", "Gross Revenue"])
    for cell in worksheet[worksheet.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in daily_rows:
        worksheet.append(
            [
                row["date"].strftime("%d/%m/%Y"),
                float(row["room_revenue"]),
                float(row["pos_revenue"]),
                float(row["event_revenue"]),
                float(row["other_revenue"]),
                float(row["gross_revenue"]),
            ]
        )
    worksheet.append(
        [
            "TOTALS",
            float(sum((row["room_revenue"] for row in daily_rows), Decimal("0.00"))),
            float(sum((row["pos_revenue"] for row in daily_rows), Decimal("0.00"))),
            float(sum((row["event_revenue"] for row in daily_rows), Decimal("0.00"))),
            float(sum((row["other_revenue"] for row in daily_rows), Decimal("0.00"))),
            float(sum((row["gross_revenue"] for row in daily_rows), Decimal("0.00"))),
        ]
    )
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 26)


def _write_finance_expense_sheet(worksheet, expense_breakdown_rows, expenses, start_date, end_date):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Expense Breakdown"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Category", "Entries", "Total"])
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in expense_breakdown_rows:
        worksheet.append(row)
    worksheet.append(
        [
            "TOTALS",
            expenses.count(),
            float(_money_total(expenses, "amount")),
        ]
    )
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    worksheet.append([])
    worksheet.append(["Date", "Category", "Description", "Amount", "Payment Method", "Recorded By", "Receipt"])
    for cell in worksheet[worksheet.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    daily_totals = {}
    for expense in expenses:
        daily_totals[expense.date] = daily_totals.get(expense.date, Decimal("0.00")) + Decimal(str(expense.amount or 0))
        worksheet.append(
            [
                expense.date.strftime("%d/%m/%Y"),
                expense.category,
                expense.description,
                float(expense.amount or 0),
                expense.get_payment_method_display(),
                expense.recorded_by_name,
                expense.receipt.name.split("/")[-1] if expense.receipt else "",
            ]
        )
    worksheet.append(["TOTALS", "", "", float(_money_total(expenses, "amount")), "", "", ""])
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    worksheet.append([])
    worksheet.append(["Date", "Daily Expense Total"])
    for cell in worksheet[worksheet.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for expense_date in sorted(daily_totals):
        worksheet.append([expense_date.strftime("%d/%m/%Y"), float(daily_totals[expense_date])])
    worksheet.append(["TOTALS", float(sum(daily_totals.values(), Decimal("0.00")))])
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 30)


def _write_finance_pnl_sheet(
    worksheet,
    revenue_breakdown,
    expense_category_totals,
    gross_revenue,
    total_expenses,
    sales_deposits_total,
    net_profit,
    start_date,
    end_date,
):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Profit & Loss Statement"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"Range: {start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Line Item", "Amount"])
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["Room Bookings Revenue", float(revenue_breakdown[0]["amount"])],
        ["POS Sales Revenue", float(revenue_breakdown[1]["amount"])],
        ["Event Reservations Revenue", float(revenue_breakdown[2]["amount"])],
        ["Other Revenue", float(revenue_breakdown[3]["amount"])],
        ["TOTAL REVENUE", float(gross_revenue)],
    ]
    for group_label in _finance_pnl_group_labels():
        rows.append([group_label, float(expense_category_totals[group_label])])
    rows.extend(
        [
            ["Other Expenses", float(expense_category_totals["Other"])],
            ["TOTAL EXPENSES", float(total_expenses)],
            ["SALES DEPOSITS", float(sales_deposits_total)],
            ["NET PROFIT / LOSS", float(net_profit)],
        ]
    )
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 30)


def _write_finance_balance_sheet(worksheet, balance_sheet):
    from openpyxl.styles import Alignment, Font, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="23444B", end_color="23444B", fill_type="solid")
    summary_fill = PatternFill(start_color="E8F1F2", end_color="E8F1F2", fill_type="solid")

    worksheet["A1"] = "Balance Sheet Snapshot"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"As at: {balance_sheet['snapshot_date'].strftime('%d/%m/%Y')}"
    worksheet.append([])
    worksheet.append(["Section", "Line Item", "Value", "Notes"])
    for cell in worksheet[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["ASSETS", "Cash on hand", float(balance_sheet["cash_on_hand"]), "Cash payments received minus withdrawals and cash expenses."],
        ["ASSETS", "Inventory value", float(balance_sheet["inventory_value"]), f"{balance_sheet['inventory_item_count']} active inventory items valued at current purchase cost."],
        ["LIABILITIES", "Outstanding supplier payments", "", balance_sheet["liabilities_note"]],
        ["EQUITY", "Retained earnings", float(balance_sheet["retained_earnings"]), "Accumulated net profit after expenses and owner withdrawals."],
    ]
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[worksheet.max_row]:
        cell.font = Font(bold=True)
        cell.fill = summary_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 36)


def _xlsx_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _parse_rota_range(request):
    period = request.GET.get("period", "week").strip().lower() or "week"
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    reference_date_str = request.GET.get("date")
    reference_date = timezone.localdate()
    custom_start_date = None
    custom_end_date = None

    try:
        if start_date_str:
            custom_start_date = date.fromisoformat(start_date_str)
    except ValueError:
        custom_start_date = None

    try:
        if end_date_str:
            custom_end_date = date.fromisoformat(end_date_str)
    except ValueError:
        custom_end_date = None

    try:
        if reference_date_str:
            reference_date = date.fromisoformat(reference_date_str)
    except ValueError:
        reference_date = timezone.localdate()

    if custom_start_date and custom_end_date:
        if custom_start_date > custom_end_date:
            custom_start_date, custom_end_date = custom_end_date, custom_start_date
        return period, custom_start_date, custom_start_date, custom_end_date
    if custom_start_date:
        return period, custom_start_date, custom_start_date, custom_start_date
    if custom_end_date:
        return period, custom_end_date, custom_end_date, custom_end_date

    if period == "month":
        start_date = reference_date.replace(day=1)
        end_date = reference_date.replace(day=monthrange(reference_date.year, reference_date.month)[1])
    elif period == "year":
        start_date = date(reference_date.year, 1, 1)
        end_date = date(reference_date.year, 12, 31)
    else:
        period = "week"
        start_date = reference_date - timedelta(days=reference_date.weekday())
        end_date = start_date + timedelta(days=6)

    return period, reference_date, start_date, end_date


def _rota_daily_breakdown(rota):
    if not all([rota.period_start, rota.period_end]):
        return []

    day_count = (rota.period_end - rota.period_start).days + 1
    daily_hours = rota.daily_hours if rota.opening_time and rota.closing_time else 0

    rows = []
    for day_index in range(day_count):
        current_day = rota.period_start + timedelta(days=day_index)
        rows.append(
            {
                "date": current_day,
                "day_name": current_day.strftime("%A"),
                "start_time": rota.opening_time,
                "end_time": rota.closing_time,
                "hours": daily_hours,
                "period_label": "Workday",
            }
        )

    return rows


def _rota_day_reports(rotas, start_date, end_date):
    reports = []
    days = [start_date + timedelta(days=idx) for idx in range((end_date - start_date).days + 1)]

    for current_day in days:
        active_rotas = [
            rota
            for rota in rotas
            if rota.period_start and rota.period_end and rota.period_start <= current_day <= rota.period_end
        ]
        reports.append(
            {
                "date": current_day,
                "day_name": current_day.strftime("%A"),
                "rotas": active_rotas,
                "employees": len(active_rotas),
                "hours": sum((rota.daily_hours for rota in active_rotas), 0),
            }
        )

    return reports


def _daily_report_rows(start_date, end_date, total_rooms):
    days = [start_date + timedelta(days=idx) for idx in range((end_date - start_date).days + 1)]

    bookings = Booking.objects.filter(
        status__in=[
            Booking.BookingStatus.CONFIRMED,
            Booking.BookingStatus.CHECKED_IN,
            Booking.BookingStatus.CHECKED_OUT,
        ],
        check_in__lte=end_date,
        check_out__gte=start_date,
    ).values("room_id", "check_in", "check_out")

    revenue_by_day = daily_booking_revenue_map(start_date, end_date)

    daily_rows = []
    for current_day in days:
        occupied_count = 0
        for booking in bookings:
            if booking_occupies_day(booking["check_in"], booking["check_out"], current_day):
                occupied_count += 1

        occupancy_percent = round((occupied_count / total_rooms) * 100, 2) if total_rooms else 0
        daily_rows.append(
            {
                "date": current_day.isoformat(),
                "occupied_rooms": occupied_count,
                "occupancy_percent": occupancy_percent,
                "revenue_collected": revenue_by_day.get(current_day, 0),
            }
        )

    return daily_rows


def _daily_total_revenue_rows(start_date, end_date):
    revenue_by_day = daily_total_revenue_map(start_date, end_date)
    return [
        {
            "date": current_day.isoformat(),
            "revenue_total": revenue_by_day.get(current_day, Decimal("0")),
        }
        for current_day in _report_days(start_date, end_date)
    ]


def _serialize_activity_feed(feed):
    serialized = []
    for entry in feed:
        serialized.append(
            {
                "icon": entry["icon"],
                "title": entry["title"],
                "meta": entry["meta"],
                "href": entry.get("href", ""),
                "timestamp": timezone.localtime(entry["time"]).isoformat(),
                "time_label": _relative_time(entry["time"]),
            }
        )
    return serialized


def _recent_activity_feed(limit=20):
    recent_bookings = Booking.objects.select_related("guest", "room").order_by("-created_at")[:6]
    recent_payments = Payment.objects.select_related("booking__guest", "booking__room").order_by("-paid_at")[:6]
    recent_event_bookings = EventBooking.objects.select_related("guest").order_by("-created_at")[:5]
    recent_event_payments = EventPayment.objects.select_related("event_booking__guest").order_by("-paid_at")[:5]
    recent_rooms = Room.objects.exclude(last_status_changed_at__isnull=True).order_by("-last_status_changed_at")[:6]
    recent_status_changes = StatusHistory.objects.select_related("changed_by", "content_type").order_by("-changed_at")[:10]
    recent_notifications = Notification.objects.order_by("-created_at")[:10]
    recent_housekeeping = HousekeepingItemLog.objects.select_related("item", "room").order_by("-used_at", "-created_at")[:6]
    recent_hires = Employee.objects.order_by("-created_at")[:5]
    recent_terminations = Employee.objects.filter(
        employment_status="terminated",
        termination_date__isnull=False,
    ).order_by("-updated_at")[:5]
    recent_leave_approvals = LeaveRequest.objects.select_related("employee").filter(
        approval_status=LeaveRequest.ApprovalStatus.APPROVED,
        approved_at__isnull=False,
    ).order_by("-approved_at")[:5]

    pending_booking_balances = Booking.objects.select_related("guest", "room").annotate(
        paid_total=Coalesce(
            Sum("payments__amount"),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
        ),
        balance=ExpressionWrapper(
            F("total_amount") - F("paid_total"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
    ).filter(balance__gt=0).order_by("-updated_at")[:5]

    pending_event_balances = EventBooking.objects.select_related("guest").annotate(
        paid_total=Coalesce(
            Sum("payments__amount"),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
        ),
        balance=ExpressionWrapper(
            F("total_amount") - F("paid_total"),
            output_field=DecimalField(max_digits=10, decimal_places=2),
        ),
    ).filter(balance__gt=0).order_by("-updated_at")[:5]

    feed = []
    for item in recent_bookings:
        feed.append(
            {
                "time": item.created_at,
                "icon": "calendar-range",
                "title": f"Booking created for {item.guest}",
                "meta": f"Room {item.room.room_number} · {item.get_status_display()}",
                "href": reverse("booking-detail", args=[item.pk]),
            }
        )
    for item in recent_payments:
        feed.append(
            {
                "time": item.paid_at,
                "icon": "wallet",
                "title": f"Payment received from {item.booking.guest}",
                "meta": f"Room {item.booking.room.room_number} · GHS {item.amount}",
                "href": reverse("payments-center"),
            }
        )
    for item in pending_booking_balances:
        feed.append(
            {
                "time": item.updated_at,
                "icon": "badge-alert",
                "title": f"Outstanding balance for {item.guest}",
                "meta": f"Room {item.room.room_number} · GHS {item.balance}",
                "href": reverse("payments-center"),
            }
        )
    for item in recent_event_bookings:
        feed.append(
            {
                "time": item.created_at,
                "icon": "party-popper",
                "title": f"Event booked: {item.event_title}",
                "meta": f"{item.event_space_name} · {item.expected_guests} guests",
                "href": reverse("event-booking-list"),
            }
        )
    for item in recent_event_payments:
        feed.append(
            {
                "time": item.paid_at,
                "icon": "wallet-cards",
                "title": f"Event payment received from {item.event_booking.guest}",
                "meta": f"{item.event_booking.event_title} · GHS {item.amount}",
                "href": reverse("payments-center"),
            }
        )
    for item in pending_event_balances:
        feed.append(
            {
                "time": item.updated_at,
                "icon": "receipt-text",
                "title": f"Outstanding event balance for {item.guest}",
                "meta": f"{item.event_title} · GHS {item.balance}",
                "href": reverse("payments-center"),
            }
        )
    for item in recent_rooms:
        feed.append(
            {
                "time": item.last_status_changed_at,
                "icon": "bed-double",
                "title": f"Room {item.room_number} status updated",
                "meta": f"{item.get_status_display()} · {item.get_room_type_display()}",
                "href": reverse("room-list"),
            }
        )
    for item in recent_housekeeping:
        feed.append(
            {
                "time": item.used_at,
                "icon": "sparkles",
                "title": f"Housekeeping logged {format_quantity(item.quantity_used)} {item.unit} of {item.item_name}",
                "meta": f"{'Room ' + item.room.room_number if item.room_id else 'General use'} · Stock left {format_quantity(item.quantity_in_stock)} {item.unit}",
                "href": reverse("housekeeping-dashboard"),
            }
        )
    for employee in recent_hires:
        feed.append(
            {
                "time": employee.created_at,
                "icon": "user-plus",
                "title": f"New staff profile created for {employee.first_name} {employee.last_name}",
                "meta": f"{employee.get_position_display()} · {employee.department or 'No department set'}",
                "href": reverse("hr-detail", args=[employee.pk]),
            }
        )
    for employee in recent_terminations:
        feed.append(
            {
                "time": employee.updated_at,
                "icon": "user-minus",
                "title": f"{employee.first_name} {employee.last_name} marked as terminated",
                "meta": f"{employee.termination_reason_choice or employee.termination_reason or 'Termination recorded'}",
                "href": reverse("hr-detail", args=[employee.pk]),
            }
        )
    for leave_request in recent_leave_approvals:
        feed.append(
            {
                "time": leave_request.approved_at,
                "icon": "calendar-check-2",
                "title": f"{leave_request.employee.first_name} {leave_request.employee.last_name} leave approved",
                "meta": f"{leave_request.get_leave_type_display()} · {leave_request.days} day(s)",
                "href": reverse("hr-detail", args=[leave_request.employee.pk]),
            }
        )
    for history in recent_status_changes:
        feed.append(
            {
                "time": history.changed_at,
                "icon": "refresh-cw",
                "title": f"{history.object_repr} status changed",
                "meta": f"{history.previous_status or 'Unknown'} → {history.new_status}",
                "href": reverse("notifications-center"),
            }
        )
    for notification in recent_notifications:
        feed.append(
            {
                "time": notification.created_at,
                "icon": "bell-ring",
                "title": notification.title,
                "meta": notification.message,
                "href": notification.link or reverse("notifications-center"),
            }
        )

    feed.sort(key=lambda row: row["time"], reverse=True)
    return _serialize_activity_feed(feed[:limit])
