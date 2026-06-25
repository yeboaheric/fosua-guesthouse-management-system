from io import BytesIO
import json
import shutil
import tempfile
from datetime import timedelta
from unittest.mock import patch

from axes.helpers import get_cool_off
from accounts.forms import ExpenseForm, LeaveRequestForm, StaffUserForm
from accounts.models import AuditLog, AttendanceRecord, Employee, EmployeeQualification, Expense, LeaveRequest, Notification, OwnerWithdrawal, PayrollRecord, Rota, RolePermission, TrainingRecord, UserAccessProfile
from accounts.permissions import user_has_permission
from bookings.models import Booking, EventBooking, EventPayment, Payment
from datetime import date, datetime, time, timedelta
from django.contrib.auth.models import Group, User
from django.contrib.auth.hashers import PBKDF2PasswordHasher
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.test import override_settings
from django.urls import reverse
from guests.models import Guest
from inventory.models import Sale
from openpyxl import load_workbook
from rooms.models import HousekeepingItem, HousekeepingItemLog, Room
from django.utils import timezone
from urllib.parse import quote_plus


class DashboardRoutingTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.receptionist_group = Group.objects.create(name="Receptionist")

    def test_admin_user_redirected_to_admin_dashboard(self):
        user = User.objects.create_user(username="admin1", password="pass123456")
        user.groups.add(self.admin_group)

        self.client.force_login(user)
        response = self.client.get(reverse("dashboard"))

        self.assertRedirects(response, reverse("admin-dashboard"))

    def test_receptionist_user_redirected_to_reception_dashboard(self):
        user = User.objects.create_user(username="reception1", password="pass123456")
        user.groups.add(self.receptionist_group)

        self.client.force_login(user)
        response = self.client.get(reverse("dashboard"))

        self.assertRedirects(response, reverse("reception-dashboard"))

    def test_admin_reports_access_for_admin_only(self):
        admin_user = User.objects.create_user(username="admin2", password="pass123456")
        admin_user.groups.add(self.admin_group)
        self.client.force_login(admin_user)
        response = self.client.get(reverse("admin-reports"))
        self.assertEqual(response.status_code, 200)

        receptionist_user = User.objects.create_user(
            username="reception2", password="pass123456"
        )
        receptionist_user.groups.add(self.receptionist_group)
        self.client.force_login(receptionist_user)
        response = self.client.get(reverse("admin-reports"))
        self.assertEqual(response.status_code, 403)

    def test_admin_dashboard_uses_live_metrics_and_chart_series(self):
        admin_user = User.objects.create_user(username="admin-dashboard", password="pass123456")
        admin_user.groups.add(self.admin_group)
        room = Room.objects.create(
            room_number="101",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=150,
        )
        guest = Guest.objects.create(
            first_name="Akua",
            last_name="Boateng",
            phone_number="0241111111",
        )
        Employee.objects.create(
            title="ms",
            first_name="Efua",
            last_name="Mensah",
            date_of_birth=date(1994, 4, 4),
            nationality="Ghanaian",
            ghana_card_number="GHA-123456789-1",
            contact_number="0242222222",
            start_date=date(2025, 1, 1),
            position="manager",
            employment_status="active",
            gender="female",
            marital_status="single",
        )
        Booking.objects.create(
            guest=guest,
            room=room,
            check_in=timezone.localdate(),
            check_out=timezone.localdate() + timedelta(days=2),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=admin_user,
        )

        self.client.force_login(admin_user)
        response = self.client.get(reverse("admin-dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Occupancy rate")
        self.assertContains(response, "Available rooms")
        self.assertContains(response, "Unread alerts")
        chart_series = json.loads(response.context["dashboard_chart_series_json"])
        self.assertEqual(set(chart_series.keys()), {"daily", "weekly", "monthly"})
        self.assertEqual(len(chart_series["daily"]["bookings"]), 7)
        self.assertEqual(len(response.context["dashboard_summary_cards"]), 12)
        self.assertEqual(response.context["total_rooms"], 1)
        self.assertEqual(response.context["total_bookings"], 1)
        self.assertEqual(response.context["total_staff"], 1)

    def test_admin_dashboard_activity_feed_endpoint_returns_live_items(self):
        admin_user = User.objects.create_user(username="admin-feed", password="pass123456")
        admin_user.groups.add(self.admin_group)
        room = Room.objects.create(
            room_number="102",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=180,
        )
        guest = Guest.objects.create(
            first_name="Kojo",
            last_name="Asante",
            phone_number="0243333333",
        )
        Booking.objects.create(
            guest=guest,
            room=room,
            check_in=timezone.localdate(),
            check_out=timezone.localdate() + timedelta(days=1),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=admin_user,
        )

        self.client.force_login(admin_user)
        response = self.client.get(reverse("admin-dashboard-activity-feed"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["items"])
        self.assertIn("icon", payload["items"][0])
        self.assertIn("time_label", payload["items"][0])

    def test_healthz_endpoint_available_without_login(self):
        self.client.logout()
        response = self.client.get(reverse("healthz"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], "ok")

    def test_login_lockout_applies_only_to_target_username(self):
        locked_user = User.objects.create_user(username="locked-user", password="pass123456")
        other_user = User.objects.create_user(username="other-user", password="pass123456")
        locked_user.groups.add(self.receptionist_group)
        other_user.groups.add(self.receptionist_group)

        failure_responses = []
        for _ in range(5):
            response = self.client.post(
                reverse("login"),
                {"username": "locked-user", "password": "wrong-password"},
            )
            failure_responses.append(response.status_code)

        self.assertTrue(all(status in {200, 429} for status in failure_responses))

        locked_response = self.client.post(
            reverse("login"),
            {"username": "locked-user", "password": "pass123456"},
        )
        self.assertEqual(locked_response.status_code, 429)

        other_response = self.client.post(
            reverse("login"),
            {"username": "other-user", "password": "pass123456"},
        )
        self.assertEqual(other_response.status_code, 302)
        self.assertEqual(other_response.headers["Location"], reverse("dashboard"))

    def test_login_lockout_cooloff_is_one_minute(self):
        self.assertEqual(get_cool_off(), timedelta(minutes=1))

    def test_users_roles_center_access_for_admin_only(self):
        admin_user = User.objects.create_user(username="admin-users", password="pass123456")
        admin_user.groups.add(self.admin_group)
        self.client.force_login(admin_user)
        response = self.client.get(reverse("users-roles-center"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Users and roles")
        self.assertContains(response, "Permissions")

        receptionist_user = User.objects.create_user(
            username="reception-users", password="pass123456"
        )
        receptionist_user.groups.add(self.receptionist_group)
        self.client.force_login(receptionist_user)
        response = self.client.get(reverse("users-roles-center"))
        self.assertEqual(response.status_code, 403)

    def test_receptionist_module_access_can_be_restricted(self):
        blocked_group = Group.objects.create(name="Blocked Reception")
        RolePermission.objects.create(
            role=blocked_group,
            module="rooms",
            can_view=True,
        )
        user = User.objects.create_user(username="reception-blocked", password="pass123456")
        user.groups.add(blocked_group)
        UserAccessProfile.objects.create(
            user=user,
            dashboard_access=True,
            reservations_access=False,
            rooms_access=True,
            guests_access=True,
            payments_access=True,
            services_access=True,
            housekeeping_access=True,
            notifications_access=True,
            analytics_access=True,
            reports_access=False,
            settings_access=False,
            staff_management_access=False,
            handovers_access=True,
            users_roles_access=False,
        )

        self.assertTrue(
            self.client.post(
                reverse("login"),
                {"username": "reception-blocked", "password": "pass123456"},
            ).status_code in {200, 302}
        )
        response = self.client.get(reverse("booking-list"))
        self.assertEqual(response.status_code, 403)

        response = self.client.get(reverse("room-list"))
        self.assertEqual(response.status_code, 200)

    def test_notifications_center_generates_arrival_reminders(self):
        user = User.objects.create_user(username="notify1", password="pass123456")
        user.groups.add(self.receptionist_group)
        UserAccessProfile.objects.create(
            user=user,
            dashboard_access=True,
            reservations_access=True,
            rooms_access=True,
            guests_access=True,
            payments_access=True,
            services_access=True,
            housekeeping_access=True,
            notifications_access=True,
            analytics_access=True,
            reports_access=False,
            settings_access=False,
            staff_management_access=False,
            handovers_access=True,
            users_roles_access=False,
        )
        room = Room.objects.create(
            room_number="501",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=120,
        )
        guest = Guest.objects.create(
            first_name="Nana",
            last_name="Yeboah",
            phone_number="0245555555",
        )
        now = timezone.now().replace(hour=10, minute=0, second=0, microsecond=0)
        check_in_time = (now + timedelta(minutes=30)).time().replace(second=0, microsecond=0)
        Booking.objects.create(
            guest=guest,
            room=room,
            check_in=now.date(),
            check_in_time=check_in_time,
            check_out=now.date() + timedelta(days=1),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=user,
        )
        self.client.force_login(user)
        with patch("accounts.views.timezone.localtime", return_value=now):
            response = self.client.get(reverse("notifications-center"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upcoming check-in")
        self.assertEqual(Notification.objects.filter(user=user).count(), 1)

    def test_receptionist_cannot_see_or_access_housekeeping(self):
        receptionist_user = User.objects.create_user(
            username="reception-housekeeping", password="pass123456"
        )
        receptionist_user.groups.add(self.receptionist_group)
        self.client.force_login(receptionist_user)

        nav_response = self.client.get(reverse("room-list"))
        self.assertEqual(nav_response.status_code, 200)
        self.assertNotContains(nav_response, ">Housekeeping<", html=False)
        self.assertContains(nav_response, "Operations Overview")

        access_response = self.client.get(reverse("housekeeping-dashboard"), follow=True)
        self.assertRedirects(access_response, reverse("reception-dashboard"))
        messages = [str(message) for message in access_response.context["messages"]]
        self.assertIn("You are not authorized to access Housekeeping.", messages)

        operations_response = self.client.get(reverse("operations-overview"))
        self.assertEqual(operations_response.status_code, 200)


class SalesDepositsModuleTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.reception_group = Group.objects.create(name="Receptionist")
        self.admin_user = User.objects.create_user(
            username="finance-admin",
            password="pass123456",
            first_name="Ama",
            last_name="Owusu",
        )
        self.admin_user.groups.add(self.admin_group)

        self.reception_user = User.objects.create_user(
            username="frontdesk",
            password="pass123456",
            first_name="Kojo",
            last_name="Mensah",
        )
        self.reception_user.groups.add(self.reception_group)

        self.room = Room.objects.create(
            room_number="201",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=150,
        )
        self.guest = Guest.objects.create(
            first_name="Yaa",
            last_name="Boateng",
            phone_number="0245551111",
        )
        Booking.objects.create(
            guest=self.guest,
            room=self.room,
            check_in=timezone.localdate(),
            check_out=timezone.localdate() + timedelta(days=1),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=self.admin_user,
        )
        Sale.objects.create(
            cashier=self.admin_user,
            customer_name="Walk in",
            payment_method=Sale.PaymentMethod.CASH,
            subtotal="30.00",
            grand_total="30.00",
            amount_paid="30.00",
        )

    def test_payments_access_user_can_open_page_and_log_sales_deposit(self):
        self.client.force_login(self.reception_user)
        response = self.client.post(
            reverse("sales-deposits-center"),
            {
                "created_at": timezone.localtime(timezone.now()).strftime("%Y-%m-%dT%H:%M"),
                "amount": "45.00",
                "collection_method": OwnerWithdrawal.CollectionMethod.CASH,
                "collected_by": "Owner",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sales deposit logged successfully.")
        withdrawal = OwnerWithdrawal.objects.get()
        self.assertEqual(str(withdrawal.amount), "45.00")
        self.assertEqual(withdrawal.entry_type, OwnerWithdrawal.EntryType.VISIT)
        self.assertEqual(withdrawal.collection_method, OwnerWithdrawal.CollectionMethod.CASH)
        self.assertEqual(withdrawal.recorded_by, self.reception_user)
        self.assertEqual(withdrawal.collected_by, "Owner")

    def test_staff_can_log_leftover_as_sales_deposit_entry(self):
        self.client.force_login(self.reception_user)
        response = self.client.post(
            reverse("sales-deposits-center"),
            {
                "created_at": timezone.localtime(timezone.now()).strftime("%Y-%m-%dT%H:%M"),
                "entry_type": OwnerWithdrawal.EntryType.LEFTOVER,
                "amount": "25.00",
                "collection_method": OwnerWithdrawal.CollectionMethod.BOTH,
                "collected_by": "",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sales deposit logged successfully.")
        self.assertContains(response, "Leftover")
        withdrawal = OwnerWithdrawal.objects.get()
        self.assertEqual(withdrawal.entry_type, OwnerWithdrawal.EntryType.LEFTOVER)
        self.assertEqual(withdrawal.collection_method, OwnerWithdrawal.CollectionMethod.BOTH)
        self.assertEqual(withdrawal.collected_by, "")

    def test_non_admin_cannot_edit_or_delete_sales_deposits(self):
        withdrawal = OwnerWithdrawal.objects.create(
            amount="60.00",
            reason="Supplier payment",
            collected_by="Owner",
            recorded_by=self.admin_user,
        )
        self.client.force_login(self.reception_user)

        response = self.client.get(reverse("sales-deposit-update", args=[withdrawal.pk]), follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Access Denied")

        delete_response = self.client.post(reverse("sales-deposit-delete", args=[withdrawal.pk]), follow=True)
        self.assertEqual(delete_response.status_code, 200)
        self.assertContains(delete_response, "Access Denied")
        self.assertTrue(OwnerWithdrawal.objects.filter(pk=withdrawal.pk).exists())

    def test_admin_can_export_sales_deposits_with_summary(self):
        now = timezone.localtime(timezone.now()).replace(hour=10, minute=0, second=0, microsecond=0)
        OwnerWithdrawal.objects.create(
            amount="40.00",
            reason="Bank deposit",
            collection_method=OwnerWithdrawal.CollectionMethod.MOBILE_MONEY,
            collected_by="Owner",
            recorded_by=self.admin_user,
            created_at=now,
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse("sales-deposits-export-xlsx"),
            {
                "start_date": timezone.localdate().isoformat(),
                "end_date": timezone.localdate().isoformat(),
            },
        )

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Sales Deposits Log", "Summary"])
        log_sheet = workbook["Sales Deposits Log"]
        summary_sheet = workbook["Summary"]
        self.assertEqual(log_sheet["A4"].value, "Entry Type")
        self.assertEqual(log_sheet["B4"].value, "Date")
        self.assertEqual(log_sheet["C4"].value, "Amount Collected")
        self.assertEqual(log_sheet["D5"].value, "Mobile Money")
        self.assertEqual(summary_sheet["A4"].value, "Category")
        self.assertEqual(summary_sheet["B4"].value, "Total Amount")
        self.assertEqual(summary_sheet["B5"].value, 40)

    def test_sales_deposit_filter_normalizes_inverted_dates(self):
        OwnerWithdrawal.objects.create(
            amount="60.00",
            collection_method=OwnerWithdrawal.CollectionMethod.CASH,
            collected_by="Owner",
            recorded_by=self.admin_user,
            created_at=timezone.make_aware(datetime.combine(timezone.localdate() - timedelta(days=1), datetime.min.time())),
        )
        OwnerWithdrawal.objects.create(
            amount="40.00",
            collection_method=OwnerWithdrawal.CollectionMethod.CASH,
            collected_by="Owner",
            recorded_by=self.admin_user,
            created_at=timezone.make_aware(datetime.combine(timezone.localdate(), datetime.min.time())),
        )
        self.client.force_login(self.admin_user)
        today = timezone.localdate()
        yesterday = today - timedelta(days=1)

        response = self.client.get(
            reverse("sales-deposits-center"),
            {
                "start_date": today.isoformat(),
                "end_date": yesterday.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["start_date"], yesterday.isoformat())
        self.assertEqual(response.context["end_date"], today.isoformat())
        self.assertEqual(response.context["summary_cards"][0]["label"], f"Selected range ({yesterday.strftime('%d/%m/%Y')} - {today.strftime('%d/%m/%Y')})")
        self.assertEqual(response.context["summary_cards"][0]["value"], "GHS 100.00")

    def test_revenue_analytics_does_not_include_sales_deposits(self):
        OwnerWithdrawal.objects.create(
            amount="50.00",
            reason="Personal",
            collected_by="Owner",
            recorded_by=self.admin_user,
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse("analytics-center"),
            {
                "period": "daily",
                "start_date": timezone.localdate().isoformat(),
                "end_date": timezone.localdate().isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total revenue")
        self.assertNotContains(response, "Net revenue")
        self.assertNotContains(response, "Owner withdrawals")


class FinanceModuleTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.reception_group = Group.objects.create(name="Receptionist")
        self.admin_user = User.objects.create_user(username="finance-owner", password="pass123456")
        self.admin_user.groups.add(self.admin_group)
        self.reception_user = User.objects.create_user(username="finance-frontdesk", password="pass123456")
        self.reception_user.groups.add(self.reception_group)

        self.room = Room.objects.create(
            room_number="301",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=200,
        )
        self.guest = Guest.objects.create(
            first_name="Abena",
            last_name="Adjei",
            phone_number="0244441111",
        )
        Booking.objects.create(
            guest=self.guest,
            room=self.room,
            check_in=timezone.localdate(),
            check_out=timezone.localdate() + timedelta(days=1),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=self.admin_user,
        )
        Sale.objects.create(
            cashier=self.admin_user,
            payment_method=Sale.PaymentMethod.CASH,
            subtotal="25.00",
            grand_total="25.00",
            amount_paid="25.00",
        )
        OwnerWithdrawal.objects.create(
            amount="10.00",
            reason="Owner pickup",
            collected_by="Owner",
            recorded_by=self.admin_user,
        )

    def test_finance_module_is_admin_only(self):
        self.client.force_login(self.reception_user)
        response = self.client.get(reverse("finance-center"), follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Access Denied: Finance is available to admin accounts only.")

    def test_admin_can_create_expense_and_export_finance_workbook(self):
        self.client.force_login(self.admin_user)
        create_response = self.client.post(
            reverse("finance-center"),
            {
                "date": timezone.localdate().isoformat(),
                "category": "Electricity (ECG)",
                "description": "Water bill",
                "amount": "30.00",
                "payment_method": Expense.PaymentMethod.CASH,
            },
            follow=True,
        )
        self.assertEqual(create_response.status_code, 200)
        self.assertContains(create_response, "Expense logged successfully.")
        self.assertEqual(Expense.objects.count(), 1)

        export_response = self.client.get(
            reverse("finance-export-xlsx"),
            {
                "start_date": timezone.localdate().isoformat(),
                "end_date": timezone.localdate().isoformat(),
            },
        )
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export_response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Revenue Breakdown", "Expense Breakdown", "Profit Loss", "Balance Sheet"],
        )
        pnl_sheet = workbook["Profit Loss"]
        pnl_labels = [
            pnl_sheet[f"A{row_index}"].value
            for row_index in range(1, pnl_sheet.max_row + 1)
            if pnl_sheet[f"A{row_index}"].value
        ]
        self.assertIn("Room Bookings Revenue", pnl_labels)
        self.assertIn("NET PROFIT / LOSS", pnl_labels)
        self.assertNotIn("SALES DEPOSITS", pnl_labels)

    def test_admin_can_create_expense_with_custom_category(self):
        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("finance-center"),
            {
                "date": timezone.localdate().isoformat(),
                "category": ExpenseForm.CUSTOM_CATEGORY_VALUE,
                "custom_category": "Special Repairs",
                "description": "Emergency contractor support",
                "amount": "55.00",
                "payment_method": Expense.PaymentMethod.CASH,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Expense logged successfully.")
        expense = Expense.objects.get(description="Emergency contractor support")
        self.assertEqual(expense.category, "Special Repairs")

    def test_expense_form_offers_default_and_saved_custom_categories(self):
        Expense.objects.create(
            date=timezone.localdate(),
            category="Custom Vendor Charge",
            description="One-off specialist service",
            amount="15.00",
            payment_method=Expense.PaymentMethod.CASH,
            recorded_by=self.admin_user,
        )

        form = ExpenseForm()
        grouped_choices = form.fields["category"].choices
        grouped_lookup = {
            label: choices
            for label, choices in grouped_choices
            if isinstance(choices, (list, tuple))
        }

        self.assertIn("Staffing", grouped_lookup)
        self.assertIn(
            ("Salaries & Wages", "Salaries & Wages"),
            grouped_lookup["Staffing"],
        )
        self.assertIn("Saved custom categories", grouped_lookup)
        self.assertIn(
            ("Custom Vendor Charge", "Custom Vendor Charge"),
            grouped_lookup["Saved custom categories"],
        )
        custom_form = ExpenseForm(
            data={
                "date": timezone.localdate().isoformat(),
                "category": ExpenseForm.CUSTOM_CATEGORY_VALUE,
                "custom_category": "Custom Vendor Charge",
                "description": "Repeat vendor charge",
                "amount": "10.00",
                "payment_method": Expense.PaymentMethod.CASH,
            }
        )
        self.assertTrue(custom_form.is_valid(), custom_form.errors)


class UsersRolesPermissionPropagationTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.admin_user = User.objects.create_user(username="permissions-admin", password="pass123456")
        self.admin_user.groups.add(self.admin_group)
        self.client.force_login(self.admin_user)

        self.custom_role = Group.objects.create(name="Custom Operations")
        self.staff_user = User.objects.create_user(username="permissions-user", password="pass123456")
        self.staff_user.groups.add(self.custom_role)
        UserAccessProfile.objects.create(
            user=self.staff_user,
            dashboard_access=True,
            reservations_access=False,
            rooms_access=False,
            guests_access=False,
            payments_access=False,
            services_access=False,
            housekeeping_access=False,
            inventory_access=False,
            pos_access=False,
            notifications_access=False,
            analytics_access=False,
            reports_access=False,
            settings_access=False,
            staff_management_access=False,
            handovers_access=False,
            users_roles_access=False,
        )
        self.staff_client = Client()
        self.staff_client.force_login(self.staff_user)

    def _save_role_permissions(self, **enabled_fields):
        payload = {
            "action": "save_role_permissions",
            "role_id": self.custom_role.pk,
            "role_name": self.custom_role.name,
        }
        payload.update(enabled_fields)
        return self.client.post(reverse("users-roles-center"), payload)

    def test_role_permission_changes_apply_immediately_without_relogin(self):
        denied_response = self.staff_client.get(reverse("inventory-dashboard"))
        self.assertEqual(denied_response.status_code, 403)

        update_response = self._save_role_permissions(inventory_view="on")
        self.assertRedirects(update_response, reverse("users-roles-center"))

        self.staff_user.access_profile.refresh_from_db()
        self.assertTrue(self.staff_user.access_profile.inventory_access)
        self.assertTrue(user_has_permission(self.staff_user, "inventory"))

        allowed_response = self.staff_client.get(reverse("inventory-dashboard"))
        self.assertEqual(allowed_response.status_code, 200)

        revoke_response = self._save_role_permissions()
        self.assertRedirects(revoke_response, reverse("users-roles-center"))

        self.staff_user.access_profile.refresh_from_db()
        self.assertFalse(self.staff_user.access_profile.inventory_access)
        self.assertFalse(user_has_permission(self.staff_user, "inventory"))

        denied_again_response = self.staff_client.get(reverse("inventory-dashboard"))
        self.assertEqual(denied_again_response.status_code, 403)


class AdminReportExportTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.admin_user = User.objects.create_user(username="admin3", password="pass123456")
        self.admin_user.groups.add(self.admin_group)
        self.client.force_login(self.admin_user)
        self.today = timezone.localdate()
        self.start_date = self.today - timedelta(days=1)
        self.end_date = self.today + timedelta(days=1)

        self.room = Room.objects.create(
            room_number="301",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=180,
        )
        self.guest = Guest.objects.create(
            first_name="Efua",
            last_name="Sarpong",
            phone_number="0241111111",
        )
        self.booking = Booking.objects.create(
            guest=self.guest,
            room=self.room,
            check_in=self.today,
            check_out=self.today + timedelta(days=2),
            status=Booking.BookingStatus.CONFIRMED,
            total_amount=500,
            created_by=self.admin_user,
        )
        Payment.objects.create(
            booking=self.booking,
            amount=200,
            method=Payment.PaymentMethod.CASH,
            received_by=self.admin_user,
        )
        self.event_booking = EventBooking.objects.create(
            guest=self.guest,
            event_title="Conference",
            purpose="Team training",
            expected_guests=30,
            event_start=timezone.now(),
            event_end=timezone.now() + timedelta(hours=4),
            total_amount=450,
            created_by=self.admin_user,
        )
        EventPayment.objects.create(
            event_booking=self.event_booking,
            amount=150,
            method=EventPayment.PaymentMethod.CARD,
            received_by=self.admin_user,
        )
        self.housekeeping_item = HousekeepingItem.objects.create(
            name="Laundry Detergent",
            initial_quantity=50,
            quantity_in_stock=45,
            unit="litres",
            created_by=self.admin_user,
        )
        HousekeepingItemLog.objects.create(
            item=self.housekeeping_item,
            item_name=self.housekeeping_item.name,
            initial_quantity=self.housekeeping_item.initial_quantity,
            quantity_used=5,
            quantity_in_stock=self.housekeeping_item.quantity_in_stock,
            low_stock_threshold=self.housekeeping_item.low_stock_threshold,
            unit=self.housekeeping_item.unit,
            room=self.room,
            used_at=timezone.now(),
            created_by=self.admin_user,
        )
        Sale.objects.create(
            cashier=self.admin_user,
            payment_method=Sale.PaymentMethod.CARD,
            grand_total=75,
            amount_paid=75,
            status=Sale.SaleStatus.COMPLETED,
        )
        self.employee = Employee.objects.create(
            title="mr",
            first_name="Kojo",
            last_name="Owusu",
            date_of_birth=date(1994, 4, 10),
            nationality="Ghanaian",
            ghana_card_number="GHA-123456789-0",
            contact_number="0200000000",
            department="Front Office",
            start_date=self.today - timedelta(days=120),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        Rota.objects.create(
            employee=self.employee,
            period="Kojo Owusu weekly duty roster",
            period_start=self.start_date,
            period_end=self.end_date,
            opening_time=time(8, 0),
            closing_time=time(16, 0),
        )
        LeaveRequest.objects.create(
            employee=self.employee,
            leave_type=LeaveRequest.LeaveType.ANNUAL,
            start_date=self.start_date,
            end_date=self.end_date,
            return_to_work_date=self.end_date + timedelta(days=1),
            reason="Annual leave",
            approval_status=LeaveRequest.ApprovalStatus.APPROVED,
            approving_manager=self.employee,
        )
        AttendanceRecord.objects.create(
            employee=self.employee,
            work_date=self.today,
            status=AttendanceRecord.AttendanceStatus.PRESENT,
        )
        PayrollRecord.objects.create(
            employee=self.employee,
            pay_period_start=self.start_date,
            pay_period_end=self.end_date,
            net_pay=1200,
        )
        TrainingRecord.objects.create(
            employee=self.employee,
            training_name="Guest Relations",
            provider="Fosua Academy",
            start_date=self.today,
        )

    def test_daily_report_csv_export_returns_csv(self):
        response = self.client.get(
            reverse("admin-reports-export-daily"),
            {"start_date": self.start_date.isoformat(), "end_date": self.end_date.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("occupied_rooms", response.content.decode())

    def test_balances_report_csv_export_returns_csv(self):
        response = self.client.get(reverse("admin-reports-export-balances"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode()
        self.assertIn("balance_due", content)
        self.assertIn("Efua Sarpong", content)

    def test_admin_reports_page_lists_system_sections_and_export_actions(self):
        response = self.client.get(
            reverse("admin-reports"),
            {"start_date": self.start_date.isoformat(), "end_date": self.end_date.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Export All")
        self.assertContains(response, "Bookings")
        self.assertContains(response, "Revenue &amp; Payments")
        self.assertContains(response, "Housekeeping")
        self.assertContains(response, "Duty Roster")
        self.assertContains(response, "Rooms")
        self.assertContains(response, "Staff &amp; HR")

    def test_section_excel_export_returns_selected_section_workbook(self):
        response = self.client.get(
            reverse("admin-reports-export-section", args=["bookings"]),
            {"start_date": self.start_date.isoformat(), "end_date": self.end_date.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            f"bookings-report-{self.start_date.strftime('%d-%m-%Y')}-to-{self.end_date.strftime('%d-%m-%Y')}.xlsx",
            response["Content-Disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Bookings"])
        self.assertEqual(workbook.active["A1"].value, "Bookings")
        self.assertTrue(
            any(cell.value == "Daily booking activity" for row in workbook.active.iter_rows() for cell in row)
        )

    def test_export_all_excel_creates_sheet_per_report_section(self):
        response = self.client.get(
            reverse("admin-reports-export-all"),
            {"start_date": self.start_date.isoformat(), "end_date": self.end_date.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f"full-report-{self.start_date.strftime('%d-%m-%Y')}-to-{self.end_date.strftime('%d-%m-%Y')}.xlsx",
            response["Content-Disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Overview", "Bookings", "Revenue Payments", "Housekeeping", "Duty Roster", "Rooms", "Staff HR"],
        )
        self.assertEqual(workbook["Overview"]["A1"].value, "Full System Report")
        self.assertEqual(workbook["Housekeeping"]["A1"].value, "Housekeeping")

    def test_same_day_booking_is_counted_in_rooms_report_occupancy(self):
        same_day_room = Room.objects.create(
            room_number="302",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=150,
        )
        Booking.objects.create(
            guest=self.guest,
            room=same_day_room,
            check_in=self.today,
            check_in_time=time(8, 0),
            check_out=self.today,
            check_out_time=time(17, 0),
            status=Booking.BookingStatus.CHECKED_OUT,
            created_by=self.admin_user,
        )

        response = self.client.get(
            reverse("admin-reports"),
            {"start_date": self.start_date.isoformat(), "end_date": self.end_date.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        rooms_section = next(section for section in response.context["sections"] if section["key"] == "rooms")
        daily_occupancy_table = next(table for table in rooms_section["tables"] if table["title"] == "Daily occupancy")
        today_row = next(row for row in daily_occupancy_table["export_rows"] if row[0] == self.today)
        self.assertEqual(today_row[1], 2)


class AnalyticsCenterTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.admin_user = User.objects.create_user(username="analytics_admin", password="pass123456")
        self.admin_user.groups.add(self.admin_group)
        self.client.force_login(self.admin_user)

        self.today = timezone.localdate()
        self.start_date = self.today - timedelta(days=6)
        self.end_date = self.today

        self.deluxe_room = Room.objects.create(
            room_number="401",
            room_type=Room.RoomType.DELUXE,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=300,
        )
        self.standard_room = Room.objects.create(
            room_number="201",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.CLEANING,
            base_rate=180,
        )
        self.guest_one = Guest.objects.create(
            first_name="Ama",
            last_name="Analytics",
            phone_number="0241000001",
        )
        self.guest_two = Guest.objects.create(
            first_name="Kojo",
            last_name="Returning",
            phone_number="0241000002",
        )
        self.deluxe_booking = Booking.objects.create(
            guest=self.guest_one,
            room=self.deluxe_room,
            check_in=self.today - timedelta(days=2),
            check_out=self.today + timedelta(days=1),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=self.admin_user,
        )
        self.standard_booking = Booking.objects.create(
            guest=self.guest_two,
            room=self.standard_room,
            check_in=self.today - timedelta(days=1),
            check_out=self.today + timedelta(days=2),
            status=Booking.BookingStatus.CANCELLED,
            created_by=self.admin_user,
        )
        Payment.objects.create(
            booking=self.deluxe_booking,
            amount=250,
            method=Payment.PaymentMethod.CASH,
            received_by=self.admin_user,
        )
        Payment.objects.create(
            booking=self.standard_booking,
            amount=100,
            method=Payment.PaymentMethod.CARD,
            received_by=self.admin_user,
        )
        self.event_booking = EventBooking.objects.create(
            guest=self.guest_one,
            event_title="Planning Session",
            purpose="Strategy",
            expected_guests=10,
            event_start=timezone.now() - timedelta(days=1),
            event_end=timezone.now() - timedelta(days=1) + timedelta(hours=4),
            total_amount=400,
            status=EventBooking.EventBookingStatus.CONFIRMED,
            created_by=self.admin_user,
        )
        EventPayment.objects.create(
            event_booking=self.event_booking,
            amount=150,
            method=EventPayment.PaymentMethod.MOBILE_MONEY,
            received_by=self.admin_user,
        )
        self.pos_sale = Sale.objects.create(
            cashier=self.admin_user,
            payment_method=Sale.PaymentMethod.CARD,
            grand_total=80,
            amount_paid=80,
            status=Sale.SaleStatus.COMPLETED,
        )
        self.reception_employee = Employee.objects.create(
            title="mr",
            first_name="Yaw",
            last_name="Frontdesk",
            date_of_birth=date(1992, 4, 4),
            nationality="Ghanaian",
            ghana_card_number="GHA-121212121-1",
            contact_number="0241111111",
            department="Front Desk",
            start_date=self.today - timedelta(days=2),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        self.cleaner_employee = Employee.objects.create(
            title="mrs",
            first_name="Adwoa",
            last_name="Cleaner",
            date_of_birth=date(1990, 8, 8),
            nationality="Ghanaian",
            ghana_card_number="GHA-232323232-2",
            contact_number="0242222222",
            department="Housekeeping",
            start_date=self.today - timedelta(days=30),
            position="cleaner",
            employment_status="active",
            gender="female",
            marital_status="married",
        )
        LeaveRequest.objects.create(
            employee=self.reception_employee,
            leave_type=LeaveRequest.LeaveType.ANNUAL,
            start_date=self.today - timedelta(days=1),
            end_date=self.today + timedelta(days=1),
            days=3,
            return_to_work_date=self.today + timedelta(days=2),
            reason="Break",
            approval_status=LeaveRequest.ApprovalStatus.APPROVED,
        )
        Rota.objects.create(
            employee=self.reception_employee,
            period="Front desk rota",
            period_start=self.start_date,
            period_end=self.end_date,
            opening_time=time(8, 0),
            closing_time=time(16, 0),
        )
        item = HousekeepingItem.objects.create(
            name="Bed Sheet",
            initial_quantity=20,
            quantity_in_stock=15,
            unit="pieces",
            created_by=self.admin_user,
        )
        HousekeepingItemLog.objects.create(
            item=item,
            item_name=item.name,
            initial_quantity=item.initial_quantity,
            quantity_used=5,
            quantity_in_stock=item.quantity_in_stock,
            unit=item.unit,
            room=self.deluxe_room,
            used_at=timezone.now() - timedelta(days=1),
            created_by=self.admin_user,
        )

    def test_analytics_center_renders_all_sections(self):
        response = self.client.get(
            reverse("analytics-center"),
            {
                "period": "custom",
                "start_date": self.start_date.isoformat(),
                "end_date": self.end_date.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rooms Analytics")
        self.assertContains(response, "Bookings Analytics")
        self.assertContains(response, "Revenue Analytics")
        self.assertContains(response, "Staff Analytics")
        self.assertContains(response, "Housekeeping Analytics")
        self.assertContains(response, "Duty Roster Analytics")
        self.assertContains(response, "Operations Overview Analytics")

    def test_analytics_filters_narrow_real_data(self):
        response = self.client.get(
            reverse("analytics-center"),
            {
                "period": "custom",
                "start_date": self.start_date.isoformat(),
                "end_date": self.end_date.isoformat(),
                "room_type": Room.RoomType.DELUXE,
                "department": "Front Desk",
                "staff_role": "receptionist",
            },
        )

        self.assertEqual(response.status_code, 200)
        section_map = {section["key"]: section for section in response.context["sections"]}
        self.assertEqual(section_map["rooms-analytics"]["metrics"][0]["value"], 1)
        self.assertEqual(section_map["staff-analytics"]["metrics"][0]["value"], 1)
        self.assertEqual(section_map["bookings-analytics"]["metrics"][0]["value"], 1)

    def test_revenue_totals_stay_consistent_across_dashboard_analytics_and_reports(self):
        dashboard_response = self.client.get(reverse("admin-dashboard"))
        self.assertEqual(dashboard_response.status_code, 200)

        reports_response = self.client.get(reverse("admin-reports"), {"period": "monthly"})
        self.assertEqual(reports_response.status_code, 200)
        report_section_map = {section["key"]: section for section in reports_response.context["sections"]}

        analytics_response = self.client.get(reverse("analytics-center"), {"period": "monthly"})
        self.assertEqual(analytics_response.status_code, 200)
        analytics_section_map = {section["key"]: section for section in analytics_response.context["sections"]}

        expected_booking_revenue = self.deluxe_booking.total_amount
        expected_event_revenue = self.event_booking.total_amount
        expected_pos_sales = self.pos_sale.grand_total
        expected_total_revenue = expected_booking_revenue + expected_event_revenue + expected_pos_sales

        self.assertEqual(dashboard_response.context["monthly_revenue"], expected_total_revenue)
        self.assertEqual(report_section_map["revenue-payments"]["summary"]["total_revenue"], expected_total_revenue)
        self.assertEqual(analytics_section_map["revenue-analytics"]["summary"]["total_revenue"], expected_total_revenue)
        self.assertEqual(report_section_map["revenue-payments"]["summary"]["pos_sales_total"], expected_pos_sales)

        analytics_revenue_rows = {
            row[0]: row[1]
            for row in analytics_section_map["revenue-analytics"]["tables"][1]["export_rows"]
        }
        report_revenue_rows = {
            row[0]: row[1]
            for row in report_section_map["revenue-payments"]["tables"][1]["export_rows"]
        }

        self.assertEqual(analytics_revenue_rows["Booking revenue"], float(expected_booking_revenue))
        self.assertEqual(analytics_revenue_rows["POS sales"], float(expected_pos_sales))
        self.assertEqual(report_revenue_rows["Booking revenue"], float(expected_booking_revenue))
        self.assertEqual(report_revenue_rows["POS sales"], float(expected_pos_sales))

    def test_analytics_export_xlsx_contains_all_section_sheets(self):
        response = self.client.get(
            reverse("analytics-export", args=["xlsx"]),
            {
                "period": "custom",
                "start_date": self.start_date.isoformat(),
                "end_date": self.end_date.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames[0], "Overview")
        self.assertIn("Rooms Analytics", workbook.sheetnames)
        self.assertIn("Bookings Analytics", workbook.sheetnames)
        self.assertIn("Revenue Analytics", workbook.sheetnames)
        self.assertIn("Staff Analytics", workbook.sheetnames)
        self.assertIn("Housekeeping Analytics", workbook.sheetnames)
        self.assertIn("Duty Roster Analytics", workbook.sheetnames)
        self.assertIn("Operations Analytics", workbook.sheetnames)

    def test_analytics_export_pdf_returns_pdf(self):
        response = self.client.get(
            reverse("analytics-export", args=["pdf"]),
            {
                "period": "custom",
                "start_date": self.start_date.isoformat(),
                "end_date": self.end_date.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")


class RotaViewTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.admin_user = User.objects.create_user(username="admin4", password="pass123456")
        self.admin_user.groups.add(self.admin_group)
        self.client.force_login(self.admin_user)

        self.employee = Employee.objects.create(
            title="mr",
            first_name="Kwame",
            last_name="Mensah",
            date_of_birth=date(1995, 6, 1),
            nationality="Ghanaian",
            ghana_card_number="GHA-999999999-9",
            contact_number="0249999999",
            start_date=date(2025, 1, 1),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        self.rota = Rota.objects.create(
            employee=self.employee,
            period="Kwame Mensah weekly duty roster",
            period_start=date(2026, 5, 18),
            period_end=date(2026, 5, 24),
            opening_time=time(8, 0),
            closing_time=time(16, 0),
            shift_rules="Keep the front desk fully covered and hand over notes each evening.",
        )
        self.rota.staff_members.set([self.employee])

    def test_rota_list_supports_range_filtering(self):
        response = self.client.get(
            reverse("hr-rota-list"),
            {"period": "month", "date": "2026-05-20", "q": "Kwame"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kwame Mensah")
        self.assertContains(response, "Daily roster report")

    def test_rota_list_supports_start_and_end_date_range_filtering(self):
        response = self.client.get(
            reverse("hr-rota-list"),
            {"start_date": "2026-05-18", "end_date": "2026-05-24", "employee": str(self.employee.pk)},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="start_date"', html=False)
        self.assertContains(response, 'name="end_date"', html=False)
        self.assertNotContains(response, "Reference date", html=False)
        self.assertEqual(list(response.context["rotas"]), [self.rota])

    def test_rota_detail_shows_daily_breakdown(self):
        response = self.client.get(reverse("hr-rota-detail", args=[self.rota.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Day-by-day employee account")
        self.assertContains(response, "Monday")
        self.assertContains(response, "8.00")

    def test_overnight_rota_hours_are_counted_across_midnight(self):
        overnight_rota = Rota.objects.create(
            employee=self.employee,
            period="Night shift rota",
            period_start=date(2026, 5, 18),
            period_end=date(2026, 5, 24),
            opening_time=time(22, 0),
            closing_time=time(6, 0),
        )

        self.assertEqual(overnight_rota.daily_hours, 8)
        self.assertEqual(overnight_rota.total_hours, 56)


class CustomRolePermissionTests(TestCase):
    def setUp(self):
        self.role = Group.objects.create(name="Stock Specialist")
        RolePermission.objects.create(
            role=self.role,
            module="inventory",
            can_view=True,
            can_create=True,
            can_edit=True,
            can_delete=False,
            can_approve=False,
            can_export=True,
            can_print=True,
            can_manage=False,
        )
        self.user = User.objects.create_user(username="stockuser", password="pass123456")
        self.user.groups.add(self.role)
        self.client.force_login(self.user)

    def test_custom_role_can_access_inventory_but_not_user_management(self):
        self.assertTrue(user_has_permission(self.user, "inventory", "view"))

        response = self.client.get(reverse("inventory-dashboard"))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse("users-roles-center"))
        self.assertEqual(response.status_code, 403)


class CenterFilterTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="admin5", password="pass123456")
        self.user.groups.add(self.admin_group)
        self.client.force_login(self.user)

        self.room = Room.objects.create(
            room_number="401",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=150,
            notes="Near the courtyard",
        )
        self.cleaning_room = Room.objects.create(
            room_number="402",
            room_type=Room.RoomType.DELUXE,
            status=Room.RoomStatus.CLEANING,
            base_rate=250,
            notes="Deep clean scheduled",
        )
        self.guest = Guest.objects.create(
            first_name="Abena",
            last_name="Danso",
            phone_number="0243333333",
            ghana_card_number="GHA-222222222-2",
        )
        self.event_guest = Guest.objects.create(
            first_name="Kwesi",
            last_name="Appiah",
            phone_number="0553333333",
        )
        self.booking = Booking.objects.create(
            guest=self.guest,
            room=self.room,
            check_in=date(2026, 6, 1),
            check_out=date(2026, 6, 3),
            status=Booking.BookingStatus.CONFIRMED,
            created_by=self.user,
        )
        self.event_booking = EventBooking.objects.create(
            guest=self.event_guest,
            event_space_name="Main Event Space",
            event_title="Annual Gala",
            purpose="Company celebration",
            expected_guests=80,
            event_start=timezone.localtime().replace(day=5, hour=10, minute=0, second=0, microsecond=0),
            event_end=timezone.localtime().replace(day=5, hour=15, minute=0, second=0, microsecond=0),
            status=EventBooking.EventBookingStatus.CONFIRMED,
            created_by=self.user,
        )
        self.cancelled_event = EventBooking.objects.create(
            guest=self.event_guest,
            event_space_name="Main Event Space",
            event_title="Cancelled Workshop",
            purpose="Training session",
            expected_guests=20,
            event_start=timezone.localtime().replace(day=8, hour=9, minute=0, second=0, microsecond=0),
            event_end=timezone.localtime().replace(day=8, hour=11, minute=0, second=0, microsecond=0),
            status=EventBooking.EventBookingStatus.CANCELLED,
            created_by=self.user,
        )
        self.room_payment = Payment.objects.create(
            booking=self.booking,
            amount=100,
            method=Payment.PaymentMethod.CASH,
            reference="ROOM-001",
            received_by=self.user,
        )
        self.event_payment = EventPayment.objects.create(
            event_booking=self.event_booking,
            amount=300,
            method=EventPayment.PaymentMethod.MOBILE_MONEY,
            reference="EVENT-001",
            received_by=self.user,
        )

    def test_guest_list_filter_returns_matching_guest(self):
        response = self.client.get(reverse("guest-list"), {"q": "Abena"})
        self.assertEqual(response.status_code, 200)
        guests = list(response.context["guests"])
        self.assertEqual(len(guests), 1)
        self.assertEqual(guests[0].first_name, "Abena")

    def test_booking_list_filter_combines_status_and_query(self):
        response = self.client.get(
            reverse("booking-list"),
            {"status": Booking.BookingStatus.CONFIRMED, "q": "401"},
        )
        self.assertEqual(response.status_code, 200)
        bookings = list(response.context["bookings"])
        self.assertEqual(len(bookings), 1)
        self.assertEqual(bookings[0].room.room_number, "401")

    def test_payments_center_filters_by_method_and_scope(self):
        response = self.client.get(
            reverse("payments-center"),
            {"method": Payment.PaymentMethod.CASH, "scope": "room", "q": "ROOM-001"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["room_payments"]), [self.room_payment])
        self.assertEqual(list(response.context["event_payments"]), [])

    def test_services_center_filters_by_status_and_query(self):
        response = self.client.get(
            reverse("services-center"),
            {"status": EventBooking.EventBookingStatus.CONFIRMED, "q": "Annual"},
        )
        self.assertEqual(response.status_code, 200)
        events = list(response.context["events"])
        self.assertEqual(len(events), 1)
        self.assertEqual(events[0].pk, self.event_booking.pk)

    def test_housekeeping_center_redirects_to_housekeeping_dashboard(self):
        response = self.client.get(
            reverse("housekeeping-center"),
            {"report": "weekly"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.headers["Location"],
            f"{reverse('housekeeping-dashboard')}?report=weekly",
        )


class StaffManagementTests(TestCase):
    def setUp(self):
        today = timezone.localdate()
        past_leave_start = today - timedelta(days=30)
        past_leave_end = today - timedelta(days=28)
        self.admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="hr_admin", password="pass123456")
        self.user.groups.add(self.admin_group)
        self.client.force_login(self.user)

        self.active_employee = Employee.objects.create(
            title="mr",
            first_name="Kojo",
            last_name="Active",
            date_of_birth=date(1990, 1, 1),
            nationality="Ghanaian",
            ghana_card_number="GHA-111111111-1",
            contact_number="0241111111",
            department="Front Desk",
            job_title="Lead Receptionist",
            start_date=date(2024, 1, 15),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        self.leave_employee = Employee.objects.create(
            title="mrs",
            first_name="Akosua",
            last_name="Leave",
            date_of_birth=date(1992, 2, 2),
            nationality="Ghanaian",
            ghana_card_number="GHA-222222222-2",
            contact_number="0242222222",
            department="Housekeeping",
            start_date=date(2024, 2, 1),
            position="cleaner",
            employment_status="annual_leave",
            gender="female",
            marital_status="married",
        )
        self.terminated_employee = Employee.objects.create(
            title="mr",
            first_name="Yaw",
            last_name="Terminated",
            date_of_birth=date(1988, 3, 3),
            nationality="Ghanaian",
            ghana_card_number="GHA-333333333-3",
            contact_number="0243333333",
            department="Security",
            start_date=date(2023, 3, 10),
            termination_date=date(2026, 5, 10),
            position="security",
            employment_status="terminated",
            gender="male",
            marital_status="single",
        )
        self.rota = Rota.objects.create(
            employee=self.active_employee,
            period="Kojo Active rota",
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 7),
            opening_time=time(8, 0),
            closing_time=time(16, 0),
        )
        self.rota.staff_members.set([self.active_employee])
        self.approver_employee = Employee.objects.create(
            title="mrs",
            first_name="Adwoa",
            last_name="Approver",
            date_of_birth=date(1985, 4, 4),
            nationality="Ghanaian",
            ghana_card_number="GHA-444444444-4",
            contact_number="0244444444",
            department="Management",
            start_date=date(2022, 1, 1),
            position="hotel_manager",
            employment_status="active",
            gender="female",
            marital_status="married",
        )
        EmployeeQualification.objects.create(
            employee=self.active_employee,
            qualification_name="Fire Safety",
            institution="Hotel Academy",
            certification_date=date(2025, 1, 1),
            expiry_date=timezone.localdate() + timedelta(days=10),
        )
        EmployeeQualification.objects.create(
            employee=self.terminated_employee,
            qualification_name="First Aid",
            institution="Safety Board",
            certification_date=date(2024, 1, 1),
            expiry_date=timezone.localdate() - timedelta(days=1),
        )
        self.active_leave_request = LeaveRequest.objects.create(
            employee=self.active_employee,
            leave_type=LeaveRequest.LeaveType.ANNUAL,
            start_date=past_leave_start,
            end_date=past_leave_end,
            days=3,
            return_to_work_date=past_leave_end + timedelta(days=1),
            reason="Approved annual leave",
            approval_status=LeaveRequest.ApprovalStatus.APPROVED,
            approving_manager=self.approver_employee,
        )
        self.active_pending_leave = LeaveRequest.objects.create(
            employee=self.active_employee,
            leave_type=LeaveRequest.LeaveType.SICK,
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 2),
            days=2,
            return_to_work_date=date(2026, 7, 3),
            reason="Pending sick leave",
            approval_status=LeaveRequest.ApprovalStatus.PENDING,
        )
        LeaveRequest.objects.create(
            employee=self.approver_employee,
            leave_type=LeaveRequest.LeaveType.FAMILY,
            start_date=date(2026, 6, 15),
            end_date=date(2026, 6, 16),
            days=2,
            return_to_work_date=date(2026, 6, 17),
            reason="Other employee leave",
            approval_status=LeaveRequest.ApprovalStatus.REJECTED,
        )
        self.active_present_attendance = AttendanceRecord.objects.create(
            employee=self.active_employee,
            work_date=date(2026, 6, 1),
            shift_type=AttendanceRecord.ShiftType.MORNING,
            status=AttendanceRecord.AttendanceStatus.PRESENT,
            notes="On time",
        )
        AttendanceRecord.objects.create(
            employee=self.active_employee,
            work_date=date(2026, 6, 2),
            shift_type=AttendanceRecord.ShiftType.MORNING,
            status=AttendanceRecord.AttendanceStatus.ABSENT,
            notes="Sick day",
        )
        AttendanceRecord.objects.create(
            employee=self.approver_employee,
            work_date=date(2026, 6, 1),
            shift_type=AttendanceRecord.ShiftType.MORNING,
            status=AttendanceRecord.AttendanceStatus.LATE,
            notes="Different employee",
        )

    def test_employee_edit_allows_blank_optional_gps_address(self):
        response = self.client.post(
            reverse("hr-update", args=[self.active_employee.pk]),
            {
                "employee_id": self.active_employee.employee_id or "",
                "title": self.active_employee.title,
                "first_name": "Kofi",
                "last_name": self.active_employee.last_name,
                "date_of_birth": self.active_employee.date_of_birth.isoformat(),
                "nationality": self.active_employee.nationality,
                "ghana_card_number": self.active_employee.ghana_card_number,
                "ghana_card_expiry_date": "",
                "ssnit_number": "",
                "contact_number": self.active_employee.contact_number,
                "email": "",
                "residential_address": "",
                "department": self.active_employee.department,
                "job_title": self.active_employee.job_title,
                "salary_amount": "",
                "supervisor": "",
                "gps_address": "",
                "emergency_contact_name": "",
                "next_of_kin": "",
                "next_of_kin_contact": "",
                "next_of_kin_relationship": "",
                "start_date": self.active_employee.start_date.isoformat(),
                "leave_entitlement_days": self.active_employee.leave_entitlement_days,
                "termination_date": "",
                "termination_reason_choice": "",
                "termination_approved_by": "",
                "termination_exit_interview_notes": "",
                "company_assets_returned": "",
                "termination_remarks": "",
                "termination_reason": "",
                "emergency_contact_number": "",
                "position": self.active_employee.position,
                "employment_status": self.active_employee.employment_status,
                "gender": self.active_employee.gender,
                "marital_status": self.active_employee.marital_status,
                "ethnic_origin": "",
                "religion": "",
            },
        )

        self.assertRedirects(response, reverse("hr-list"))
        self.active_employee.refresh_from_db()
        self.assertEqual(self.active_employee.first_name, "Kofi")
        self.assertEqual(self.active_employee.gps_address, "")

    def test_employee_ids_are_generated_sequentially(self):
        self.assertEqual(self.active_employee.employee_id, "EMP0001")
        self.assertEqual(self.leave_employee.employee_id, "EMP0002")
        self.assertEqual(self.terminated_employee.employee_id, "EMP0003")
        self.assertEqual(self.approver_employee.employee_id, "EMP0004")

        next_employee = Employee.objects.create(
            title="mr",
            first_name="Kwesi",
            last_name="Sequential",
            date_of_birth=date(1991, 5, 5),
            nationality="Ghanaian",
            ghana_card_number="GHA-555555555-5",
            contact_number="0245555555",
            department="Operations",
            start_date=date(2026, 1, 10),
            position="security",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        self.assertEqual(next_employee.employee_id, "EMP0005")

    def test_employee_create_ignores_manual_id_and_uses_next_sequence(self):
        response = self.client.post(
            reverse("hr-create"),
            {
                "employee_id": "MANUAL-999",
                "title": "mr",
                "first_name": "Kwaku",
                "last_name": "AutoId",
                "date_of_birth": "1994-05-01",
                "nationality": "Ghanaian",
                "ghana_card_number": "GHA-666666666-6",
                "ghana_card_expiry_date": "",
                "ssnit_number": "",
                "contact_number": "0246666666",
                "email": "",
                "residential_address": "",
                "department": "Front Desk",
                "job_title": "Reception Clerk",
                "salary_amount": "",
                "supervisor": "",
                "gps_address": "",
                "emergency_contact_name": "",
                "next_of_kin": "",
                "next_of_kin_contact": "",
                "next_of_kin_relationship": "",
                "start_date": "2026-06-07",
                "leave_entitlement_days": 21,
                "termination_date": "",
                "termination_reason_choice": "",
                "termination_approved_by": "",
                "termination_exit_interview_notes": "",
                "company_assets_returned": "",
                "termination_remarks": "",
                "termination_reason": "",
                "emergency_contact_number": "",
                "position": "receptionist",
                "employment_status": "active",
                "gender": "male",
                "marital_status": "single",
                "ethnic_origin": "",
                "religion": "",
            },
        )

        self.assertRedirects(response, reverse("hr-list"))
        created_employee = Employee.objects.get(ghana_card_number="GHA-666666666-6")
        self.assertEqual(created_employee.employee_id, "EMP0005")

    def test_employee_create_form_hides_job_title_and_termination_fields(self):
        response = self.client.get(reverse("hr-create"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Job title", html=False)
        self.assertNotContains(response, "Termination Date", html=False)
        self.assertNotContains(response, "Termination reason", html=False)
        self.assertNotContains(response, "Termination notes", html=False)
        self.assertNotContains(response, "Exit interview notes", html=False)
        self.assertNotContains(response, 'name="job_title"', html=False)
        self.assertNotContains(response, 'name="termination_date"', html=False)
        self.assertNotContains(response, 'name="termination_reason_choice"', html=False)
        self.assertNotContains(response, 'name="termination_reason"', html=False)

    def test_leave_request_form_uses_active_employees_as_approvers(self):
        form = LeaveRequestForm()
        approvers = list(form.fields["approving_manager"].queryset)
        self.assertIn(self.active_employee, approvers)
        self.assertIn(self.approver_employee, approvers)
        self.assertNotIn(self.terminated_employee, approvers)
        self.assertFalse(form.fields["days"].required)
        self.assertFalse(form.fields["return_to_work_date"].required)

    def test_leave_request_days_auto_calculate_on_save(self):
        response = self.client.post(
            reverse("hr-employee-section", args=[self.active_employee.pk, "leave"]),
            {
                "leave_type": LeaveRequest.LeaveType.ANNUAL,
                "start_date": "2026-06-10",
                "end_date": "2026-06-12",
                "days": "",
                "return_to_work_date": "",
                "reason": "Annual break",
                "approval_status": LeaveRequest.ApprovalStatus.APPROVED,
                "approving_manager": self.approver_employee.pk,
                "decision_notes": "Approved",
            },
        )

        self.assertRedirects(
            response,
            reverse("hr-employee-section", args=[self.active_employee.pk, "leave"]),
        )
        leave_request = LeaveRequest.objects.get(
            employee=self.active_employee,
            reason="Annual break",
        )
        self.assertEqual(leave_request.days, 3)
        self.assertEqual(leave_request.return_to_work_date, date(2026, 6, 13))
        self.assertEqual(leave_request.approving_manager, self.approver_employee)

    def test_approved_current_leave_updates_employee_status(self):
        today = timezone.localdate()
        response = self.client.post(
            reverse("hr-employee-section", args=[self.active_employee.pk, "leave"]),
            {
                "leave_type": LeaveRequest.LeaveType.SICK,
                "start_date": today.isoformat(),
                "end_date": today.isoformat(),
                "days": "",
                "return_to_work_date": "",
                "reason": "Medical review",
                "approval_status": LeaveRequest.ApprovalStatus.APPROVED,
                "approving_manager": self.approver_employee.pk,
                "decision_notes": "Approved",
            },
        )

        self.assertRedirects(
            response,
            reverse("hr-employee-section", args=[self.active_employee.pk, "leave"]),
        )
        self.active_employee.refresh_from_db()
        self.assertEqual(self.active_employee.employment_status, "sick_leave")

    def test_employee_detail_subsection_links_show_real_counts_and_preserve_filters(self):
        staff_filters_token = quote_plus("staff_view=active&role=receptionist")
        response = self.client.get(
            reverse("hr-detail", args=[self.active_employee.pk]),
            {"staff_filters": staff_filters_token},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["back_to_list_url"],
            f"{reverse('hr-list')}?staff_view=active&role=receptionist",
        )
        self.assertContains(response, "Annual Leave (2)")
        self.assertContains(response, "Attendance History (2)")
        leave_link = next(
            link for link in response.context["section_links"] if link["key"] == "leave"
        )
        self.assertIn(f"staff_filters={staff_filters_token}", leave_link["url"])

    def test_leave_section_filters_only_the_selected_employee_records(self):
        response = self.client.get(
            reverse("hr-employee-section", args=[self.active_employee.pk, "leave"]),
            {
                "approval_status": LeaveRequest.ApprovalStatus.APPROVED,
                "leave_type": LeaveRequest.LeaveType.ANNUAL,
                "date_from": self.active_leave_request.start_date.isoformat(),
                "date_to": self.active_leave_request.end_date.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["records"]), [self.active_leave_request])
        self.assertEqual(response.context["record_count"], 1)

    def test_attendance_section_filters_only_the_selected_employee_records(self):
        response = self.client.get(
            reverse("hr-employee-section", args=[self.active_employee.pk, "attendance"]),
            {
                "attendance_status": AttendanceRecord.AttendanceStatus.PRESENT,
                "date_from": "2026-06-01",
                "date_to": "2026-06-01",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["records"]), [self.active_present_attendance])
        self.assertEqual(response.context["record_count"], 1)

    def test_certifications_section_filters_only_the_selected_employee_records(self):
        response = self.client.get(
            reverse("hr-employee-section", args=[self.active_employee.pk, "certifications"]),
            {
                "certification_status": "expiring",
                "query": "Fire",
                "date_from": "2025-01-01",
                "date_to": "2025-12-31",
            },
        )

        self.assertEqual(response.status_code, 200)
        records = list(response.context["records"])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].employee, self.active_employee)
        self.assertEqual(records[0].qualification_name, "Fire Safety")

    def test_staff_page_defaults_to_active_staff_and_has_terminated_tab(self):
        response = self.client.get(reverse("hr-list"))
        self.assertEqual(response.status_code, 200)
        employees = list(response.context["employees"])
        self.assertIn(self.active_employee, employees)
        self.assertIn(self.leave_employee, employees)
        self.assertIn(self.approver_employee, employees)
        self.assertNotIn(self.terminated_employee, employees)
        self.assertContains(response, "Terminated Employees")
        self.assertEqual(response.context["staff_view"], "active")
        self.assertNotContains(response, 'name="hired_from"', html=False)
        self.assertNotContains(response, 'name="hired_to"', html=False)
        self.assertNotContains(response, 'name="terminated_from"', html=False)
        self.assertNotContains(response, 'name="terminated_to"', html=False)
        self.assertNotContains(response, 'name="leave"', html=False)
        self.assertIn("Front Desk", response.context["departments"])
        self.assertIn(("receptionist", "Receptionist"), response.context["role_options"])
        status_choices = dict(response.context["status_choices"])
        self.assertEqual(status_choices["active"], "Active")
        self.assertEqual(status_choices["terminated"], "Terminated")
        self.assertEqual(status_choices["annual_leave"], "Annual Leave")
        self.assertEqual(status_choices["sick_leave"], "Sick Leave")
        self.assertEqual(status_choices["family_emergency"], "Emergency Leave")
        self.assertEqual(status_choices["maternity_leave"], "Maternity Leave")
        self.assertEqual(status_choices["paternity_leave"], "Paternity Leave")
        self.assertEqual(status_choices["unpaid_leave"], "Unpaid Leave")
        self.assertEqual(status_choices["compassionate_leave"], "Compassionate Leave")
        self.assertEqual(status_choices["study_leave"], "Study Leave")
        self.assertIn(("Fire Safety", "Fire Safety"), response.context["certification_filter_options"])
        roster_values = dict(response.context["roster_options"])
        self.assertIn(f"rota:{self.rota.pk}", roster_values)

    def test_terminated_tab_returns_only_terminated_staff(self):
        response = self.client.get(reverse("hr-list"), {"staff_view": "terminated"})
        self.assertEqual(response.status_code, 200)
        employees = list(response.context["employees"])
        self.assertEqual(employees, [self.terminated_employee])

    def test_staff_filters_return_expected_results(self):
        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "department": "Front Desk"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "q": "Kojo"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "status": "annual_leave"})
        self.assertEqual(list(response.context["employees"]), [self.leave_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "role": "receptionist"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "certifications": "expiring"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "terminated", "certifications": "expired"})
        self.assertEqual(list(response.context["employees"]), [self.terminated_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "certifications": "Fire Safety"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "roster": f"rota:{self.rota.pk}"})
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(
            reverse("hr-list"),
            {
                "staff_view": "active",
                "department": "Front Desk",
                "role": "receptionist",
                "certifications": "Fire Safety",
                "roster": f"rota:{self.rota.pk}",
            },
        )
        self.assertEqual(list(response.context["employees"]), [self.active_employee])

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "status": "terminated"})
        self.assertEqual(list(response.context["employees"]), [self.terminated_employee])

    def test_status_filter_uses_live_leave_type_records(self):
        today = timezone.localdate()
        leave_employee = Employee.objects.create(
            title="mr",
            first_name="Yaw",
            last_name="Maternity",
            date_of_birth=date(1991, 6, 6),
            nationality="Ghanaian",
            ghana_card_number="GHA-777777777-7",
            contact_number="0247777777",
            department="Front Desk",
            start_date=date(2024, 3, 1),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        LeaveRequest.objects.create(
            employee=leave_employee,
            leave_type=LeaveRequest.LeaveType.MATERNITY,
            start_date=today,
            end_date=today + timedelta(days=2),
            days=3,
            return_to_work_date=today + timedelta(days=3),
            reason="Coverage test",
            approval_status=LeaveRequest.ApprovalStatus.APPROVED,
            approving_manager=self.approver_employee,
        )

        leave_employee.refresh_from_db()
        self.assertEqual(leave_employee.employment_status, "maternity_leave")

        response = self.client.get(reverse("hr-list"), {"staff_view": "active", "status": "maternity_leave"})
        self.assertEqual(list(response.context["employees"]), [leave_employee])


class PermissionSnapshotTests(TestCase):
    def test_permission_changes_take_effect_immediately_and_after_relogin(self):
        role = Group.objects.create(name="Temp Reception")
        RolePermission.objects.create(
            role=role,
            module="reservations",
            can_view=True,
            can_create=True,
        )
        user = User.objects.create_user(username="temp_reception", password="pass123456")
        user.groups.add(role)

        self.assertTrue(
            self.client.post(
                reverse("login"),
                {"username": "temp_reception", "password": "pass123456"},
            ).status_code in {200, 302}
        )
        response = self.client.get(reverse("booking-list"))
        self.assertEqual(response.status_code, 200)

        permission = RolePermission.objects.get(role=role, module="reservations")
        permission.can_view = False
        permission.can_create = False
        permission.save(update_fields=["can_view", "can_create", "updated_at"])

        response = self.client.get(reverse("booking-list"))
        self.assertEqual(response.status_code, 403)

        self.client.logout()
        self.assertTrue(
            self.client.post(
                reverse("login"),
                {"username": "temp_reception", "password": "pass123456"},
            ).status_code in {200, 302}
        )
        response = self.client.get(reverse("booking-list"))
        self.assertEqual(response.status_code, 403)


class EmployeePhotoDisplayTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(shutil.rmtree, self.media_root, ignore_errors=True)
        self.admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="photo_admin", password="pass123456")
        self.user.groups.add(self.admin_group)
        self.client.force_login(self.user)

    def test_uploaded_employee_photo_renders_as_image_in_list_and_detail(self):
        upload = SimpleUploadedFile(
            "avatar.gif",
            (
                b"GIF89a\x01\x00\x01\x00\x80\x00\x00"
                b"\x00\x00\x00\xff\xff\xff!\xf9\x04\x01"
                b"\x00\x00\x00\x00,\x00\x00\x00\x00\x01"
                b"\x00\x01\x00\x00\x02\x02D\x01\x00;"
            ),
            content_type="image/gif",
        )

        response = self.client.post(
            reverse("hr-create"),
            {
                "employee_id": "",
                "title": "mr",
                "first_name": "Yaw",
                "last_name": "Photo",
                "date_of_birth": "1994-05-01",
                "nationality": "Ghanaian",
                "ghana_card_number": "GHA-888888888-8",
                "ghana_card_expiry_date": "",
                "ssnit_number": "",
                "contact_number": "0248888888",
                "email": "",
                "residential_address": "",
                "department": "Front Desk",
                "job_title": "Reception Clerk",
                "salary_amount": "",
                "supervisor": "",
                "gps_address": "",
                "emergency_contact_name": "",
                "next_of_kin": "",
                "next_of_kin_contact": "",
                "next_of_kin_relationship": "",
                "start_date": "2026-06-07",
                "leave_entitlement_days": 21,
                "termination_date": "",
                "termination_reason_choice": "",
                "termination_approved_by": "",
                "termination_exit_interview_notes": "",
                "company_assets_returned": "",
                "termination_remarks": "",
                "termination_reason": "",
                "emergency_contact_number": "",
                "position": "receptionist",
                "employment_status": "active",
                "gender": "male",
                "marital_status": "single",
                "ethnic_origin": "",
                "religion": "",
                "passport_photo": upload,
            },
        )

        self.assertRedirects(response, reverse("hr-list"))
        employee = Employee.objects.get(ghana_card_number="GHA-888888888-8")
        self.assertTrue(employee.profile_photo_url)

        list_response = self.client.get(reverse("hr-list"))
        self.assertContains(list_response, f'src="{employee.profile_photo_url}"', html=False)
        self.assertNotContains(list_response, f">{employee.passport_photo.name}<", html=False)

        detail_response = self.client.get(reverse("hr-detail", args=[employee.pk]))
        self.assertContains(detail_response, f'src="{employee.profile_photo_url}"', html=False)
        self.assertNotContains(detail_response, f">{employee.passport_photo.name}<", html=False)

    def test_employee_without_photo_uses_placeholder_avatar(self):
        employee = Employee.objects.create(
            title="mrs",
            first_name="Akua",
            last_name="Placeholder",
            date_of_birth=date(1992, 2, 2),
            nationality="Ghanaian",
            ghana_card_number="GHA-999999998-8",
            contact_number="0249999998",
            department="Housekeeping",
            start_date=date(2026, 1, 1),
            position="cleaner",
            employment_status="active",
            gender="female",
            marital_status="single",
        )

        list_response = self.client.get(reverse("hr-list"))
        self.assertContains(list_response, "employee-avatar-placeholder")
        self.assertContains(list_response, employee.initials)

        detail_response = self.client.get(reverse("hr-detail", args=[employee.pk]))
        self.assertContains(detail_response, "employee-portrait-placeholder")
        self.assertContains(detail_response, employee.initials)

    def test_employee_edit_page_loads_with_photo_widget(self):
        employee = Employee.objects.create(
            title="mr",
            first_name="Edit",
            last_name="Page",
            date_of_birth=date(1991, 1, 1),
            nationality="Ghanaian",
            ghana_card_number="GHA-999999996-6",
            contact_number="0249999996",
            department="Front Desk",
            start_date=date(2026, 1, 1),
            position="receptionist",
            employment_status="active",
            gender="male",
            marital_status="single",
        )

        response = self.client.get(reverse("hr-update", args=[employee.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Passport photo")

    def test_missing_photo_file_falls_back_to_placeholder(self):
        employee = Employee.objects.create(
            title="mr",
            first_name="Kojo",
            last_name="Missing",
            date_of_birth=date(1991, 1, 1),
            nationality="Ghanaian",
            ghana_card_number="GHA-999999997-7",
            contact_number="0249999997",
            department="Security",
            start_date=date(2026, 1, 1),
            position="security",
            employment_status="active",
            gender="male",
            marital_status="single",
        )
        employee.passport_photo = "employee_photos/missing.gif"
        employee.save(update_fields=["passport_photo"])

        list_response = self.client.get(reverse("hr-list"))
        self.assertNotContains(list_response, 'src="/media/employee_photos/missing.gif"', html=False)
        self.assertContains(list_response, "employee-avatar-placeholder")

        detail_response = self.client.get(reverse("hr-detail", args=[employee.pk]))
        self.assertNotContains(detail_response, 'src="/media/employee_photos/missing.gif"', html=False)
        self.assertContains(detail_response, "employee-portrait-placeholder")


class UserDeletionTests(TestCase):
    def test_delete_user_removes_account(self):
        admin_group = Group.objects.create(name="Admin")
        receptionist_group = Group.objects.create(name="Receptionist")
        admin = User.objects.create_user(username="delete_admin", password="pass123456")
        admin.groups.add(admin_group)
        target = User.objects.create_user(username="delete_target", password="pass123456")
        target.groups.add(receptionist_group)
        UserAccessProfile.objects.create(user=target, dashboard_access=True)
        self.client.force_login(admin)

        response = self.client.post(
            reverse("users-roles-center"),
            {"action": "delete_user", "user_id": target.pk},
        )

        self.assertRedirects(response, reverse("users-roles-center"))
        self.assertFalse(User.objects.filter(username="delete_target").exists())


class PasswordResetViewsTests(TestCase):
    def test_password_reset_page_is_available(self):
        response = self.client.get(reverse("password_reset"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reset your password")


class SecurityHardeningTests(TestCase):
    def setUp(self):
        cache.clear()

    def test_security_headers_are_present(self):
        response = self.client.get(reverse("login"))

        self.assertIn("default-src 'self'", response.headers["Content-Security-Policy"])
        self.assertIn("script-src 'self' 'nonce-", response.headers["Content-Security-Policy"])
        self.assertNotIn("script-src 'self' 'unsafe-inline'", response.headers["Content-Security-Policy"])
        self.assertEqual(response.headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(response.headers["X-Frame-Options"], "DENY")
        self.assertEqual(response.headers["Referrer-Policy"], "strict-origin-when-cross-origin")

    def test_staff_user_form_enforces_strong_password_policy(self):
        role = Group.objects.create(name="Security Test Role")
        form = StaffUserForm(
            data={
                "username": "weak-password-user",
                "password1": "password1",
                "password2": "password1",
                "roles": [role.name],
                "is_active": True,
                "is_staff": True,
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("uppercase letter", " ".join(form.errors["password1"]))

    def test_new_passwords_use_argon2(self):
        user = User.objects.create_user(username="argon-user", password="Strong!Pass1")
        self.assertTrue(user.password.startswith("argon2$"))

    def test_legacy_pbkdf2_password_is_upgraded_after_login(self):
        encoded = PBKDF2PasswordHasher().encode("Strong!Pass1", PBKDF2PasswordHasher().salt())
        user = User.objects.create(username="legacy-user", password=encoded)

        response = self.client.post(
            reverse("login"),
            {"username": user.username, "password": "Strong!Pass1"},
        )

        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertTrue(user.password.startswith("argon2$"))

    def test_password_reset_is_rate_limited(self):
        responses = [
            self.client.post(reverse("password_reset"), {"email": "nobody@example.com"})
            for _ in range(6)
        ]
        self.assertEqual(responses[-1].status_code, 429)
        self.assertIn("Retry-After", responses[-1].headers)

    def test_failed_login_is_audited_without_password(self):
        self.client.post(
            reverse("login"),
            {"username": "audit-user", "password": "NeverLogThis!1"},
        )

        event = AuditLog.objects.filter(details__event="login_failed").latest("created_at")
        self.assertEqual(event.object_repr, "audit-user")
        self.assertNotIn("password", event.details)
        self.assertNotIn("NeverLogThis", str(event.details))

    def test_denied_backend_access_is_audited(self):
        role = Group.objects.create(name="No Reports Access")
        user = User.objects.create_user(username="denied-user", password="Strong!Pass1")
        user.groups.add(role)
        self.client.force_login(user)

        response = self.client.get(reverse("admin-reports"))

        self.assertEqual(response.status_code, 403)
        self.assertTrue(
            AuditLog.objects.filter(
                user=user,
                status_code=403,
                details__event="authorization_denied",
            ).exists(),
            list(AuditLog.objects.values("user_id", "status_code", "details", "path")),
        )
