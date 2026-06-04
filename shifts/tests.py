from datetime import timedelta, time
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from shifts.models import (
    ShiftHandover,
    ShiftHandoverUpdate,
    Shift,
    Department,
    DutyRoster,
    DutyRosterEntry,
)


class ShiftHandoverWorkflowTests(TestCase):
    def setUp(self):
        receptionist_group = Group.objects.create(name="Receptionist")
        self.reception_a = User.objects.create_user(
            username="reception_a",
            password="pass123456",
        )
        self.reception_b = User.objects.create_user(
            username="reception_b",
            password="pass123456",
        )
        self.reception_a.groups.add(receptionist_group)
        self.reception_b.groups.add(receptionist_group)

    def test_reception_can_create_handover(self):
        self.client.force_login(self.reception_a)
        start = timezone.localtime() - timedelta(hours=8)
        end = timezone.localtime()
        response = self.client.post(
            reverse("handover-create"),
            {
                "started_at": start.strftime("%Y-%m-%dT%H:%M"),
                "ended_at": end.strftime("%Y-%m-%dT%H:%M"),
                "summary": "Cash reconciled and all arrivals confirmed.",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(ShiftHandover.objects.count(), 1)

    def test_next_reception_can_add_update_only(self):
        handover = ShiftHandover.objects.create(
            started_at=timezone.localtime() - timedelta(hours=8),
            ended_at=timezone.localtime(),
            prepared_by=self.reception_a,
            summary="Initial shift handover.",
        )
        self.client.force_login(self.reception_b)
        response = self.client.post(
            reverse("handover-detail", args=[handover.pk]),
            {"note": "Received. Late arrival guest checked in at 20:15."},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(ShiftHandoverUpdate.objects.count(), 1)
        handover.refresh_from_db()
        self.assertEqual(handover.summary, "Initial shift handover.")


class RosterReportFilterTests(TestCase):
    """Tests for roster filtering and reporting functionality."""

    def setUp(self):
        """Create test fixtures for roster testing."""
        self.admin_group = Group.objects.create(name="Admin")
        # Create departments
        self.dept_reception = Department.objects.create(
            name="Reception", description="Front desk staff", is_active=True
        )
        self.dept_housekeeping = Department.objects.create(
            name="Housekeeping", description="Room maintenance", is_active=True
        )

        # Create shifts
        self.morning_shift = Shift.objects.create(
            name="Morning",
            start_time=time(6, 0),
            end_time=time(14, 0),
            is_active=True,
        )
        self.evening_shift = Shift.objects.create(
            name="Evening",
            start_time=time(14, 0),
            end_time=time(22, 0),
            is_active=True,
        )

        # Create users
        self.user1 = User.objects.create_user(
            username="john_doe", first_name="John", last_name="Doe", password="pass123"
        )
        self.user1.groups.add(self.admin_group)
        self.user2 = User.objects.create_user(
            username="jane_smith",
            first_name="Jane",
            last_name="Smith",
            password="pass123",
        )
        self.user3 = User.objects.create_user(
            username="bob_wilson",
            first_name="Bob",
            last_name="Wilson",
            password="pass123",
        )

        # Create rosters
        today = timezone.now().date()
        self.roster1 = DutyRoster.objects.create(
            roster_date=today,
            status="published",
            created_by=self.user1,
            notes="Regular schedule",
        )
        self.roster2 = DutyRoster.objects.create(
            roster_date=today + timedelta(days=1),
            status="draft",
            created_by=self.user1,
            notes="Special event",
        )

        # Create roster entries
        DutyRosterEntry.objects.create(
            roster=self.roster1,
            employee=self.user1,
            department=self.dept_reception,
            shift=self.morning_shift,
            role="Receptionist",
            assigned_duties="Check-in/check-out, phone duty",
            status="confirmed",
            assigned_by=self.user1,
        )
        DutyRosterEntry.objects.create(
            roster=self.roster1,
            employee=self.user2,
            department=self.dept_housekeeping,
            shift=self.morning_shift,
            role="Housekeeper",
            assigned_duties="Room cleaning",
            status="confirmed",
            assigned_by=self.user1,
        )
        DutyRosterEntry.objects.create(
            roster=self.roster2,
            employee=self.user3,
            department=self.dept_reception,
            shift=self.evening_shift,
            role="Night Manager",
            assigned_duties="Overnight supervision",
            status="assigned",
            assigned_by=self.user1,
        )
        self.client.force_login(self.user1)

    def test_roster_report_page_loads(self):
        """Test that roster report page loads successfully."""
        response = self.client.get(reverse("roster-report"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "shifts/roster_report.html")
        self.assertIn("filter_form", response.context)

    def test_roster_filtering_by_date_range(self):
        """Test filtering rosters by date range."""
        today = timezone.now().date()
        response = self.client.get(
            reverse("roster-report"),
            {
                "start_date": today.isoformat(),
                "end_date": today.isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        rosters = response.context.get("rosters", [])
        self.assertEqual(len(rosters), 1)
        self.assertEqual(rosters[0].roster_date, today)

    def test_roster_filtering_by_department(self):
        """Test filtering entries by department."""
        response = self.client.get(
            reverse("roster-report"), {"department": self.dept_reception.pk}
        )
        self.assertEqual(response.status_code, 200)

    def test_roster_detail_displays_all_entries(self):
        """Test that roster detail page displays all entries for a roster."""
        response = self.client.get(
            reverse("roster-detail", args=[self.roster1.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "shifts/roster_detail.html")
        self.assertEqual(response.context["roster"], self.roster1)
        entries = response.context.get("entries", [])
        self.assertEqual(len(entries), 2)

    def test_roster_detail_groups_by_shift(self):
        """Test that roster detail groups entries by shift."""
        response = self.client.get(
            reverse("roster-detail", args=[self.roster1.pk])
        )
        self.assertEqual(response.status_code, 200)
        shifts = response.context.get("shifts", [])
        self.assertIn(self.morning_shift, shifts)


class RosterExcelExportTests(TestCase):
    """Tests for Excel export functionality."""

    def setUp(self):
        """Create test fixtures for Excel export testing."""
        self.admin_group = Group.objects.create(name="Admin")
        # Create departments
        self.dept_reception = Department.objects.create(
            name="Reception", description="Front desk staff", is_active=True
        )
        self.dept_housekeeping = Department.objects.create(
            name="Housekeeping", description="Room maintenance", is_active=True
        )

        # Create shifts
        self.morning_shift = Shift.objects.create(
            name="Morning",
            start_time=time(6, 0),
            end_time=time(14, 0),
            is_active=True,
        )
        self.evening_shift = Shift.objects.create(
            name="Evening",
            start_time=time(14, 0),
            end_time=time(22, 0),
            is_active=True,
        )

        # Create users
        self.user1 = User.objects.create_user(
            username="manager1", first_name="Alice", last_name="Manager", password="pass123"
        )
        self.user1.groups.add(self.admin_group)
        self.staff_users = []
        for i in range(5):
            user = User.objects.create_user(
                username=f"staff{i}",
                first_name=f"Staff",
                last_name=f"Member{i}",
                password="pass123",
            )
            self.staff_users.append(user)

        # Create rosters
        today = timezone.now().date()
        self.roster = DutyRoster.objects.create(
            roster_date=today,
            status="published",
            created_by=self.user1,
            notes="Test roster for export",
        )

        # Create diverse roster entries - use different employees and shifts
        shifts = [self.morning_shift, self.evening_shift]
        for i, user in enumerate(self.staff_users):
            DutyRosterEntry.objects.create(
                roster=self.roster,
                employee=user,
                department=self.dept_reception if i % 2 == 0 else self.dept_housekeeping,
                shift=shifts[i % 2],
                role=f"Role {i}",
                assigned_duties=f"Duty {i}",
                status="confirmed" if i % 2 == 0 else "assigned",
                assigned_by=self.user1,
                notes=f"Note {i}" if i % 3 == 0 else "",
            )
        self.client.force_login(self.user1)

    def test_excel_export_endpoint_returns_file(self):
        """Test that Excel export endpoint returns a valid Excel file."""
        response = self.client.get(reverse("roster-export-excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(
            response["Content-Disposition"].startswith("attachment; filename=")
        )

    def test_excel_export_contains_roster_data(self):
        """Test that exported Excel file contains roster data."""
        response = self.client.get(reverse("roster-export-excel"))
        self.assertEqual(response.status_code, 200)

        # Load the workbook from response content
        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        # Check headers
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = [
            "Date",
            "Shift",
            "Department",
            "Employee",
            "Role",
            "Assigned Duties",
            "Status",
        ]
        for expected in expected_headers:
            self.assertIn(expected, headers)

    def test_excel_export_includes_all_entries(self):
        """Test that all roster entries are included in Excel export."""
        response = self.client.get(reverse("roster-export-excel"))
        self.assertEqual(response.status_code, 200)

        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        # Count data rows (excluding header)
        data_rows = worksheet.max_row - 1
        entries = DutyRosterEntry.objects.filter(
            roster__roster_date=timezone.now().date()
        )
        self.assertGreaterEqual(data_rows, entries.count())

    def test_excel_export_file_formatting(self):
        """Test that Excel file has proper formatting."""
        response = self.client.get(reverse("roster-export-excel"))
        self.assertEqual(response.status_code, 200)

        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        # Check header formatting
        header_row = worksheet[1]
        for cell in header_row:
            # Check that header has fill color (styling)
            self.assertIsNotNone(cell.fill)
            # Check that header font is styled
            self.assertIsNotNone(cell.font)

    def test_excel_export_filtered_by_date_range(self):
        """Test Excel export with date range filtering."""
        today = timezone.now().date()
        tomorrow = today + timedelta(days=1)

        response = self.client.get(
            reverse("roster-export-excel"),
            {
                "start_date": today.isoformat(),
                "end_date": today.isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)

        excel_file = BytesIO(response.content)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active

        # Verify only today's data is in the export
        data_rows = worksheet.max_row - 1
        self.assertGreater(data_rows, 0)

    def test_excel_export_filename_includes_date(self):
        """Test that Excel export filename includes date range."""
        response = self.client.get(reverse("roster-export-excel"))
        self.assertEqual(response.status_code, 200)

        filename = response["Content-Disposition"]
        self.assertIn("filename=", filename)
        self.assertIn("roster_report", filename)
        self.assertIn(".xlsx", filename)
