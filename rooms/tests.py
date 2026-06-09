from datetime import timedelta
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from rooms.models import HousekeepingItem, HousekeepingItemLog, Room


class RoomCategoryTests(TestCase):
    def setUp(self):
        admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="admin_rooms", password="pass123456")
        self.user.groups.add(admin_group)
        Room.objects.create(
            room_number="101",
            room_type=Room.RoomType.DELUXE,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=320,
        )
        Room.objects.create(
            room_number="109",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.CLEANING,
            base_rate=180,
        )

    def test_room_list_shows_standard_and_deluxe_categories(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("room-list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Deluxe Rooms")
        self.assertContains(response, "Standard Rooms")
        self.assertContains(response, "Cleaning In Progress")


class RoomTimestampTests(TestCase):
    def test_room_status_timestamps_do_not_change_for_non_status_updates(self):
        room = Room.objects.create(
            room_number="210",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=200,
        )

        original_started = room.status_started_at
        original_changed = room.last_status_changed_at

        room.base_rate = 250
        room.save()
        room.refresh_from_db()
        self.assertEqual(room.status_started_at, original_started)
        self.assertEqual(room.last_status_changed_at, original_changed)

    def test_room_status_timestamps_reset_when_status_changes(self):
        room = Room.objects.create(
            room_number="211",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=210,
        )

        original_started = room.status_started_at
        original_changed = room.last_status_changed_at

        room.status = Room.RoomStatus.CLEANING
        room.save()
        room.refresh_from_db()
        self.assertGreater(room.status_started_at, original_started)
        self.assertGreater(room.last_status_changed_at, original_changed)


class RoomEditPermissionTests(TestCase):
    def setUp(self):
        self.admin_group = Group.objects.create(name="Admin")
        self.manager_group = Group.objects.create(name="Manager")
        self.receptionist_group = Group.objects.create(name="Receptionist")
        self.room = Room.objects.create(
            room_number="250",
            room_type=Room.RoomType.DELUXE,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=275,
            notes="Ocean-facing room",
        )

    def test_receptionist_can_open_edit_form_but_only_change_status(self):
        user = User.objects.create_user(username="reception_room", password="pass123456")
        user.groups.add(self.receptionist_group)
        self.client.force_login(user)

        response = self.client.get(reverse("room-update", args=[self.room.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Only the room status can be changed", html=False)
        self.assertTrue(response.context["status_only"])
        self.assertTrue(response.context["form"].fields["room_number"].disabled)
        self.assertTrue(response.context["form"].fields["room_type"].disabled)
        self.assertTrue(response.context["form"].fields["base_rate"].disabled)
        self.assertTrue(response.context["form"].fields["notes"].disabled)
        self.assertFalse(response.context["form"].fields["status"].disabled)
        self.assertContains(response, 'name="room_number"', html=False)
        self.assertContains(response, 'disabled id="id_room_number"', html=False)
        self.assertContains(response, 'name="status"', html=False)

        response = self.client.post(
            reverse("room-update", args=[self.room.pk]),
            {
                "room_number": "999",
                "room_type": Room.RoomType.STANDARD,
                "status": Room.RoomStatus.CLEANING,
                "base_rate": "999.00",
                "notes": "Changed by receptionist",
            },
        )
        self.assertRedirects(response, reverse("room-list"))
        self.room.refresh_from_db()
        self.assertEqual(self.room.status, Room.RoomStatus.CLEANING)
        self.assertEqual(self.room.room_number, "250")
        self.assertEqual(self.room.room_type, Room.RoomType.DELUXE)
        self.assertEqual(self.room.base_rate, 275)
        self.assertEqual(self.room.notes, "Ocean-facing room")

    def test_manager_keeps_full_room_edit_access(self):
        user = User.objects.create_user(username="manager_room", password="pass123456")
        user.groups.add(self.manager_group)
        self.client.force_login(user)

        response = self.client.get(reverse("room-update", args=[self.room.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Only the room status can be changed", html=False)
        self.assertFalse(response.context["status_only"])
        self.assertFalse(response.context["form"].fields["room_number"].disabled)

        response = self.client.post(
            reverse("room-update", args=[self.room.pk]),
            {
                "room_number": "251",
                "room_type": Room.RoomType.STANDARD,
                "status": Room.RoomStatus.MAINTENANCE,
                "base_rate": "325.00",
                "notes": "Manager update",
            },
        )
        self.assertRedirects(response, reverse("room-list"))
        self.room.refresh_from_db()
        self.assertEqual(self.room.room_number, "251")
        self.assertEqual(self.room.room_type, Room.RoomType.STANDARD)
        self.assertEqual(self.room.status, Room.RoomStatus.MAINTENANCE)
        self.assertEqual(self.room.base_rate, 325)
        self.assertEqual(self.room.notes, "Manager update")


class HousekeepingUsageLoggerTests(TestCase):
    def setUp(self):
        admin_group = Group.objects.create(name="Admin")
        self.user = User.objects.create_user(username="housekeeping_admin", password="pass123456")
        self.user.groups.add(admin_group)
        self.room = Room.objects.create(
            room_number="300",
            room_type=Room.RoomType.STANDARD,
            status=Room.RoomStatus.AVAILABLE,
            base_rate=100,
        )
        self.client.force_login(self.user)

    def test_add_item_sets_initial_quantity_as_opening_stock(self):
        response = self.client.post(
            f"{reverse('housekeeping-dashboard')}?report=daily&mode=item",
            {
                "form_type": "item",
                "item-name": "Face Towel",
                "item-initial_quantity": "25.000",
                "item-unit": "pieces",
                "item-low_stock_threshold": "5.000",
            },
        )
        self.assertEqual(response.status_code, 302)
        item = HousekeepingItem.objects.get(name="Face Towel")
        self.assertEqual(str(item.initial_quantity), "25.000")
        self.assertEqual(str(item.quantity_in_stock), "25.000")

    def test_create_usage_entry_subtracts_from_item_stock(self):
        item = HousekeepingItem.objects.create(
            name="Bath Soap",
            initial_quantity="12.000",
            quantity_in_stock="12.000",
            low_stock_threshold="3.000",
            unit="bars",
            created_by=self.user,
        )

        response = self.client.post(
            f"{reverse('housekeeping-dashboard')}?report=daily&mode=log",
            {
                "form_type": "log",
                "log-item": item.pk,
                "log-quantity_used": "2.000",
                "log-room": self.room.pk,
                "log-used_at": timezone.localtime(timezone.now()).strftime("%Y-%m-%dT%H:%M"),
                "log-notes": "Restocked bathroom",
            },
        )
        self.assertEqual(response.status_code, 302)
        entry = HousekeepingItemLog.objects.get(item=item)
        self.assertEqual(str(entry.initial_quantity), "12.000")
        self.assertEqual(str(entry.quantity_used), "2.000")
        self.assertEqual(str(entry.quantity_in_stock), "10.000")
        self.assertEqual(str(entry.low_stock_threshold), "3.000")
        self.assertEqual(entry.room, self.room)
        self.assertEqual(entry.created_by, self.user)
        item.refresh_from_db()
        self.assertEqual(str(item.quantity_in_stock), "10.000")

    def test_edit_and_delete_usage_entry_recalculate_item_stock(self):
        item = HousekeepingItem.objects.create(
            name="Towel",
            initial_quantity="8.000",
            quantity_in_stock="8.000",
            unit="sheets",
            created_by=self.user,
        )
        entry = HousekeepingItemLog.objects.create(
            item=item,
            item_name="Towel",
            initial_quantity="8.000",
            quantity_used="4.000",
            quantity_in_stock="4.000",
            unit="sheets",
            room=self.room,
            used_at=timezone.now(),
            created_by=self.user,
        )

        edit_response = self.client.post(
            f"{reverse('housekeeping-log-edit', args=[entry.pk])}?report=weekly",
            {
                "form_type": "log",
                "log-item": item.pk,
                "log-quantity_used": "5.000",
                "log-room": self.room.pk,
                "log-used_at": timezone.localtime(entry.used_at).strftime("%Y-%m-%dT%H:%M"),
                "log-notes": "Updated count",
            },
        )
        self.assertEqual(edit_response.status_code, 302)
        entry.refresh_from_db()
        self.assertEqual(str(entry.quantity_used), "5.000")
        self.assertEqual(str(entry.initial_quantity), "8.000")
        self.assertEqual(str(entry.quantity_in_stock), "3.000")
        self.assertEqual(entry.notes, "Updated count")
        item.refresh_from_db()
        self.assertEqual(str(item.quantity_in_stock), "3.000")

        delete_response = self.client.post(
            f"{reverse('housekeeping-log-delete', args=[entry.pk])}?report=weekly"
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(HousekeepingItemLog.objects.filter(pk=entry.pk).exists())
        item.refresh_from_db()
        self.assertEqual(str(item.quantity_in_stock), "8.000")

    def test_daily_report_and_excel_export(self):
        now = timezone.now()
        soap = HousekeepingItem.objects.create(
            name="Bath Soap",
            initial_quantity="10.000",
            quantity_in_stock="10.000",
            unit="bars",
            created_by=self.user,
        )
        sheet = HousekeepingItem.objects.create(
            name="Bed Sheet",
            initial_quantity="12.000",
            quantity_in_stock="12.000",
            unit="sheets",
            created_by=self.user,
        )
        HousekeepingItemLog.objects.create(
            item=soap,
            item_name="Bath Soap",
            initial_quantity="10.000",
            quantity_used="2.000",
            quantity_in_stock="8.000",
            unit="bars",
            room=self.room,
            used_at=now,
            created_by=self.user,
        )
        HousekeepingItemLog.objects.create(
            item=soap,
            item_name="Bath Soap",
            initial_quantity="10.000",
            quantity_used="1.000",
            quantity_in_stock="7.000",
            unit="bars",
            room=self.room,
            used_at=now - timedelta(hours=1),
            created_by=self.user,
        )
        HousekeepingItemLog.objects.create(
            item=sheet,
            item_name="Bed Sheet",
            initial_quantity="12.000",
            quantity_used="3.000",
            quantity_in_stock="9.000",
            unit="sheets",
            room=self.room,
            used_at=now - timedelta(days=2),
            created_by=self.user,
        )

        response = self.client.get(f"{reverse('housekeeping-dashboard')}?report=daily")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Daily report")
        self.assertContains(response, "Bath Soap")
        self.assertEqual(
            [row["item_name"] for row in response.context["report_summary_rows"]],
            ["Bath Soap"],
        )
        self.assertContains(response, "Total initial stock")
        self.assertContains(response, "Total items currently in stock")
        self.assertEqual(str(response.context["summary_total_items_in_stock"]), "16")

        export_response = self.client.get(reverse("housekeeping-report-export", args=["daily"]))
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("housekeeping-report-daily-", export_response["Content-Disposition"])

        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet["A4"].value, "Item Name")
        self.assertEqual(worksheet["B4"].value, "Total Initial Qty")
        self.assertEqual(worksheet["A5"].value, "Bath Soap")
        self.assertEqual(worksheet["B5"].value, 10)
        self.assertEqual(worksheet["C5"].value, 3)
        self.assertEqual(worksheet["D5"].value, 7)
        self.assertEqual(worksheet[f"A{worksheet.max_row}"].value, "TOTALS")

    def test_low_stock_alert_uses_default_and_custom_thresholds(self):
        now = timezone.now()
        cleaner = HousekeepingItem.objects.create(
            name="Glass Cleaner",
            initial_quantity="10.000",
            quantity_in_stock="10.000",
            unit="litres",
            created_by=self.user,
        )
        laundry = HousekeepingItem.objects.create(
            name="Laundry Soap",
            initial_quantity="30.000",
            quantity_in_stock="30.000",
            low_stock_threshold="12.000",
            unit="bars",
            created_by=self.user,
        )
        HousekeepingItemLog.objects.create(
            item=cleaner,
            item_name="Glass Cleaner",
            initial_quantity="10.000",
            quantity_used="8.500",
            quantity_in_stock="1.500",
            unit="litres",
            room=self.room,
            used_at=now,
            created_by=self.user,
        )
        HousekeepingItemLog.objects.create(
            item=laundry,
            item_name="Laundry Soap",
            initial_quantity="30.000",
            quantity_used="20.000",
            quantity_in_stock="10.000",
            low_stock_threshold="12.000",
            unit="bars",
            room=self.room,
            used_at=now,
            created_by=self.user,
        )

        response = self.client.get(reverse("housekeeping-dashboard"))
        self.assertEqual(response.status_code, 200)
        low_stock_names = [entry.name for entry in response.context["low_stock_entries"]]
        self.assertEqual(low_stock_names, ["Glass Cleaner", "Laundry Soap"])

    def test_housekeeping_quantities_render_without_trailing_zeroes(self):
        item = HousekeepingItem.objects.create(
            name="Bath Soap",
            initial_quantity="10.000",
            quantity_in_stock="10.000",
            low_stock_threshold="3.000",
            unit="bars",
            created_by=self.user,
        )
        entry = HousekeepingItemLog.objects.create(
            item=item,
            item_name="Bath Soap",
            initial_quantity="10.000",
            quantity_used="2.000",
            quantity_in_stock="8.000",
            low_stock_threshold="3.000",
            unit="bars",
            room=self.room,
            used_at=timezone.now(),
            created_by=self.user,
        )

        response = self.client.get(reverse("housekeeping-dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, ">10<", html=False)
        self.assertContains(response, ">8<", html=False)
        self.assertNotContains(response, ">10.000<", html=False)
        self.assertNotContains(response, ">8.000<", html=False)
        self.assertNotContains(response, ">3.000<", html=False)

        edit_response = self.client.get(
            f"{reverse('housekeeping-log-edit', args=[entry.pk])}?report=daily"
        )
        self.assertEqual(edit_response.status_code, 200)
        self.assertContains(edit_response, 'value="2"', html=False)
        self.assertNotContains(edit_response, 'value="2.000"', html=False)
        self.assertContains(edit_response, ">10<", html=False)
        self.assertNotContains(edit_response, ">10.000<", html=False)
